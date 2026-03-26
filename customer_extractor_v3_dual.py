"""
Customer Name Extractor from Drawings - QUAD-STAGE VERSION + FILE CLASSIFICATION
==================================================================================

שיפורים בגרסה זו:
 שלב מקדים: זיהוי אוטומטי של סוג קבצים (שרטוטים vs מסמכים אחרים)
 חיסכון בעלויות - עיבוד רק של שרטוטים
 דוח מפורט של כל הקבצים בתיקייה
 חילוץ בשני שלבים נפרדים לדיוק מקסימלי
 שלב 1: זיהוי בסיסי (לקוח, פריט, שרטוט, גרסה, חומר)
 שלב 2: תהליכים ומפרטים (ציפוי, צביעה, מפרטים)
 OCR עם Tesseract + עיבוד תמונה מתקדם
 מעקב אחר עלויות מדויק

Usage:
    python customer_extractor_v3_dual.py [optional_folder_path]
"""

import os
import sys
import re
import base64
import json
import time
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
from pathlib import Path
from typing import Optional, Tuple, List, Dict, Any
from datetime import datetime
import zipfile
import subprocess

from dotenv import load_dotenv
from openai import AzureOpenAI
import pandas as pd
import pdfplumber
from src.services.ai import ModelRuntimeConfig
from src.services.image.processing import (
    _downsample_high_res_image,
    _enhance_contrast_for_title_block,
    _extract_image_smart,
    _assess_image_quality,
    _apply_rotation_angle,
    _fix_image_rotation,
)
from src.services.ai.vision_api import (
    _build_client,
    _resolve_stage_call_config,
    _calculate_stage_cost,
    _chat_create_with_token_compat,
    _call_vision_api_with_retry,
)
from src.services.extraction.filename_utils import (
    check_value_in_filename,
    check_exact_match_in_filename,
    _disambiguate_part_number,
    _normalize_item_number,
    _fuzzy_substring_match,
    extract_part_number_from_filename,
    _extract_item_number_from_filename,
)
from src.services.extraction.pn_voting import (
    deduplicate_line,
    extract_pn_dn_from_text as _extract_pn_dn_from_text,
    vote_best_pn as _vote_best_pn,
)
from src.services.extraction.sanity_checks import (
    is_cage_code,
    run_pn_sanity_checks,
    calculate_confidence,
)
from src.services.extraction.post_processing import (
    post_process_summary_from_notes as _post_process_summary_from_notes,
)
from src.services.extraction.quantity_matcher import (
    match_quantities_to_drawings as _match_quantities_to_drawings,
    extract_base_and_suffix as _extract_base_and_suffix,
    override_pn_from_email as _override_pn_from_email,
)
from src.services.reporting.b2b_export import (
    _save_text_summary_with_variants,
)
from src.services.reporting.pl_generator import (
    extract_pl_data,
)
from src.services.reporting.excel_export import (
    _save_classification_report,
    _update_pl_sheet_with_associated_items,
    _save_results_to_excel,
)
from src.services.file.file_utils import (
    _get_file_metadata,
    _build_drawing_part_map,
    _find_associated_drawing,
    _copy_folder_to_tosend,
)
from src.services.file.classifier import classify_file_type
from src.services.file.file_renamer import rename_files_by_classification as _rename_files_by_classification
from src.services.extraction.document_reader import (
    _read_email_content,
    _extract_item_details_from_documents,
)
from src.services.extraction.stage9_merge import (
    merge_descriptions as _merge_descriptions,
)

from src.utils.logger import get_logger

# ספריות עיבוד תמונה ו-OCR
import cv2
import numpy as np
from PIL import Image

# Load environment variables
load_dotenv()

# Ensure Unicode-safe console output (important on Windows cp1255/cp1252 terminals)
try:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

logger = get_logger(__name__)

MODEL_RUNTIME = ModelRuntimeConfig.from_env()
AZURE_DEPLOYMENT = MODEL_RUNTIME.deployment
DRAWING_EXTS = {".pdf", ".png", ".jpg", ".jpeg", ".tif", ".tiff"}

# Pricing from runtime configuration (.env)
MODEL_INPUT_PRICE_PER_1M = MODEL_RUNTIME.input_price_per_1m
MODEL_OUTPUT_PRICE_PER_1M = MODEL_RUNTIME.output_price_per_1m

# Per-stage model mapping (configurable in .env via STAGE_{N}_*)
STAGE_CLASSIFICATION = 0
STAGE_LAYOUT = 0
STAGE_ROTATION = 0
STAGE_BASIC_INFO = 1
STAGE_PROCESSES = 2
STAGE_NOTES = 3
STAGE_AREA = 4
STAGE_VALIDATION = 5
STAGE_PL = 6
STAGE_EMAIL_QUANTITIES = 7
STAGE_ORDER_ITEM_DETAILS = 8
STAGE_DESCRIPTION_MERGE = 9


STAGE_DISPLAY_NAMES = {
    STAGE_CLASSIFICATION: "Stage 0 (Classification/Layout/Rotation)",
    STAGE_BASIC_INFO: "Stage 1 (Basic Info)",
    STAGE_PROCESSES: "Stage 2 (Processes)",
    STAGE_NOTES: "Stage 3 (Notes)",
    STAGE_AREA: "Stage 4 (Area)",
    STAGE_VALIDATION: "Stage 5 (Validation)",
    STAGE_PL: "Stage 6 (PL)",
    STAGE_EMAIL_QUANTITIES: "Stage 7 (Email Quantities)",
    STAGE_ORDER_ITEM_DETAILS: "Stage 8 (Quote/Order Item Details)",
    STAGE_DESCRIPTION_MERGE: "Stage 9 (Description Merge)",
}
# File size limits (in MB)
MAX_FILE_SIZE_MB = 100  # Skip files larger than this
WARN_FILE_SIZE_MB = 50  # Warn about files larger than this

# Image resolution limits (in pixels)
MAX_IMAGE_DIMENSION = 4096  # Maximum width or height in pixels
TARGET_IMAGE_DIMENSION = 2048  # Target dimension for downsampling high-res images
WARN_IMAGE_DIMENSION = 3000  # Warn about images larger than this

#  Structured JSON output
RESPONSE_FORMAT = {"type": "json_object"}

# ── Internal modules (split from this file) ──
from src.services.extraction import ocr_engine as _ocr_mod
from src.services.extraction.ocr_engine import (
    debug_print,
    MultiOCREngine,
    extract_stage1_with_retry,
    DEBUG_ENABLED,
    IAI_TOP_RED_FALLBACK_ENABLED,
    set_gui_callbacks,
)
from src.core.cost_tracker import CostTracker
from src.services.extraction.stages_generic import (
    identify_drawing_layout,
    extract_basic_info,
    extract_processes_info,
    validate_notes_before_stage5,
    extract_notes_text,
    calculate_geometric_area,
)
from src.services.extraction.stages_rafael import (
    identify_drawing_layout_rafael,
    extract_basic_info_rafael,
    extract_processes_info_rafael,
    extract_processes_from_notes,
    extract_notes_text_rafael,
    extract_area_info_rafael,
)
from src.services.extraction.stages_iai import (
    _extract_iai_top_red_identifier,
    identify_drawing_layout_iai,
    extract_basic_info_iai,
    extract_processes_info_iai,
    extract_notes_text_iai,
    extract_area_info_iai,
)
from src.services.extraction.insert_validator import validate_inserts_hardware
from src.services.extraction.insert_price_lookup import enrich_inserts_with_prices


# ── Functions moved to src/services/extraction/ ──
# _post_process_summary_from_notes → post_processing.py
# _extract_pn_dn_from_text, _vote_best_pn → pn_voting.py
# extract_drawing_data, _run_with_timeout → drawing_pipeline.py


def extract_customer_name(file_path: str, client: AzureOpenAI, ocr_engine: MultiOCREngine, selected_stages: Optional[Dict[int, bool]] = None, enable_retry: bool = False, stage1_skip_retry_resolution_px: int = 8000) -> Tuple[Optional[Dict], Optional[Tuple[int, int]]]:
    """
     חילוץ מידע בארבעה שלבים
     בוחר אוטומטית בין מודל רפאל למודל סטנדרטי
    
    Args:
        file_path: Path to drawing file
        client: Azure OpenAI client
        ocr_engine: OCR engine
        selected_stages: Dict of which stages to run {1: True, 2: False, 3: True, 4: True}
    """
    from src.services.extraction.drawing_pipeline import extract_drawing_data

    # Default: run all stages
    if selected_stages is None:
        selected_stages = {1: True, 2: True, 3: True, 4: True}

    # ── Run the full pipeline (Stages 0-4 + post-processing + P.N. voting) ──
    result_data, token_counts, context = extract_drawing_data(
        file_path=file_path,
        client=client,
        ocr_engine=ocr_engine,
        selected_stages=selected_stages,
        enable_retry=enable_retry,
        stage1_skip_retry_resolution_px=stage1_skip_retry_resolution_px,
    )

    if result_data is None:
        return None, None

    # ── Sanity checks (uses context from pipeline) ───────────────
    filename = Path(file_path).stem
    result_data = run_pn_sanity_checks(
        result_data,
        filename=filename,
        file_path=file_path,
        pdfplumber_text=context.get('pdfplumber_text', ''),
        is_rafael=context.get('is_rafael', False),
        is_iai=context.get('is_iai', False),
    )

    # Add metadata
    result_data['num_pages'] = context.get('num_pages', 1)

    # Calculate confidence
    result_data = calculate_confidence(result_data, filename, file_path)

    part_in_filename = result_data.get('part_in_filename', False)
    drawing_in_filename = result_data.get('drawing_in_filename', False)

    # Display results
    if result_data.get('customer_name'):
        logger.info(f"Customer: {result_data['customer_name']}")
    if result_data.get('part_number'):
        status = "✅" if part_in_filename else "⚠️"
        quality_flag = "  [בעייתי]" if result_data.get('needs_review') else ""
        logger.info(f"{status} Part#: {result_data['part_number']}{quality_flag} {'(in filename)' if part_in_filename else '(NOT in filename!)'}")
    if result_data.get('item_name'):
        logger.info(f"Item Name: {result_data['item_name']}")
    if result_data.get('drawing_number'):
        status = "✅" if drawing_in_filename else "⚠️"
        logger.info(f"{status} Drawing#: {result_data['drawing_number']} {'(in filename)' if drawing_in_filename else '(NOT in filename!)'}")
    if result_data.get('revision'):
        logger.info(f"Revision: {result_data['revision']}")
    if result_data.get('material'):
        logger.info(f"Material: {result_data['material']}")
    if result_data.get('coating_processes'):
        logger.info(f"Coating: {result_data['coating_processes']}")
    if result_data.get('painting_processes'):
        logger.info(f"Painting: {result_data['painting_processes']}")
    if result_data.get('colors'):
        logger.info(f"Colors: {result_data['colors']}")
    if result_data.get('marking_process'):
        logger.info(f"Marking: {result_data['marking_process']}")
    if result_data.get('part_area'):
        logger.info(f"Area: {result_data['part_area']}")
    if result_data.get('specifications'):
        logger.info(f"Specifications: {result_data['specifications']}")
    if result_data.get('parts_list_page'):
        logger.info(f"Parts List: page {result_data['parts_list_page']}")
    if result_data.get('process_summary_hebrew'):
        logger.info(f"Summary: {result_data['process_summary_hebrew']}")
    if result_data.get('notes_full_text'):
        preview = result_data['notes_full_text'][:100].replace('\n', ' ')
        logger.info(f"Notes: {preview}...")

    # Clean up internal flags before returning
    result_data.pop('_searched_for_pn_field', None)
    result_data.pop('validation_warnings', None)

    # Pass accurate per-stage cost to caller
    if context and 'pipeline_cost_usd' in context:
        result_data['_pipeline_cost_usd'] = context['pipeline_cost_usd']

    return result_data, token_counts


# 
# Folder Scanning
# 
# Folder Scanning
# 

def scan_folder(
    folder_path: Path,
    recursive: bool = True,
    after_date: Optional[datetime] = None,
    date_range: Optional[tuple] = None,
    selected_stages: Optional[Dict[int, bool]] = None,
    enable_image_retry: bool = False,
    tosend_folder: Optional[str] = None,
    confidence_level: str = "LOW",
    stage1_skip_retry_resolution_px: int = 8000,
    max_file_size_mb: Optional[int] = None,
    max_image_dimension: Optional[int] = None,
) -> List[Dict[str, Any]]:
    """Scan folder for drawings
    
    Args:
        confidence_level: B2B file confidence filter - "LOW" (all rows), "MEDIUM" (MEDIUM+HIGH+FULL), "HIGH" (HIGH+FULL only)
    """
    global MAX_FILE_SIZE_MB, MAX_IMAGE_DIMENSION

    try:
        if max_file_size_mb is not None:
            MAX_FILE_SIZE_MB = max(int(max_file_size_mb), 1)
    except Exception:
        pass

    try:
        if max_image_dimension is not None:
            MAX_IMAGE_DIMENSION = max(int(max_image_dimension), 256)
    except Exception:
        pass

    # Default: run all stages
    if selected_stages is None:
        selected_stages = {1: True, 2: True, 3: True, 4: True}
    folder_path = folder_path.resolve()
    if not folder_path.exists():
        logger.info(f"Folder not found: {folder_path}")
        return [], folder_path, None, {}
    
    logger.info(f"Scanning folder: {folder_path}")
    logger.info(f"{'(including subfolders)' if recursive else '(no subfolders)'}")
    confidence_descriptions = {'LOW': 'כל השורות', 'MEDIUM': 'בינוני+גבוה+מלא', 'HIGH': 'גבוה בלבד'}
    desc = confidence_descriptions.get(confidence_level, confidence_level)
    logger.info(f"📊 B2B file confidence filter: {confidence_level} ({desc})")
    
    skip_dirs = {".venv", "venv", "env", "__pycache__", ".git", "node_modules"}
    
    # 
    # שלב 0: איסוף תת-תיקיות
    # 
    subfolders_to_process = []
    
    if recursive:
        # Collect all subfolders (including root)
        for root, dirs, _ in os.walk(folder_path):
            dirs[:] = [d for d in dirs if d.lower() not in skip_dirs]
            subfolders_to_process.append(Path(root))
    else:
        # Only root folder
        subfolders_to_process.append(folder_path)
    
    if not subfolders_to_process:
        logger.info("No folders to process")
        return [], folder_path, None, {}

    # 
    # שלב 0.5: פתיחת קובצי ZIP ו-RAR אל תוך התיקיות
    # 
    for current_folder in subfolders_to_process:
        # Extract ZIP files
        for zip_path in current_folder.glob("*.zip"):
            try:
                if not zipfile.is_zipfile(zip_path):
                    continue
                logger.info(f"Extracting ZIP: {zip_path.name} -> {current_folder}")
                with zipfile.ZipFile(zip_path, "r") as zf:
                    # Extract all files to root folder (flatten structure)
                    for member in zf.namelist():
                        # Skip directories
                        if member.endswith('/'):
                            continue
                        # Get just the filename (no path)
                        filename = Path(member).name
                        if not filename:  # Skip if empty
                            continue
                        # Extract to root folder with flattened name
                        source = zf.open(member)
                        target_path = current_folder / filename
                        with open(target_path, 'wb') as target:
                            target.write(source.read())
                        source.close()
                    logger.info(f"Extracted {len([m for m in zf.namelist() if not m.endswith('/')])} files (flattened)")
                try:
                    zip_path.unlink()
                    logger.info(f"Deleted ZIP after extract: {zip_path.name}")
                except Exception as e:
                    logger.error(f"** Failed to delete ZIP {zip_path.name}: {e}")
            except Exception as e:
                logger.error(f"** Failed to extract {zip_path.name}: {e}")
        
        # Extract RAR files using UnRAR.exe or 7z
        for rar_path in current_folder.glob("*.rar"):
            try:
                logger.info(f"Extracting RAR: {rar_path.name} -> {current_folder}")
                extracted = False

                # Locate UnRAR.exe in common WinRAR install paths
                unrar_exe = None
                for _p in (
                    Path(r"C:\Program Files\WinRAR\UnRAR.exe"),
                    Path(r"C:\Program Files (x86)\WinRAR\UnRAR.exe"),
                ):
                    if _p.exists():
                        unrar_exe = str(_p)
                        break
                if not unrar_exe:
                    import shutil as _sh
                    unrar_exe = _sh.which("UnRAR")

                if unrar_exe:
                    result = subprocess.run(
                        [unrar_exe, 'e', '-o+', '-y', str(rar_path), str(current_folder) + os.sep],
                        capture_output=True, text=True, timeout=60,
                        encoding='utf-8', errors='replace',
                    )
                    if result.returncode == 0:
                        logger.info(f"Extracted RAR via UnRAR")
                        extracted = True
                    else:
                        logger.warning(f"UnRAR failed (rc={result.returncode}): {result.stderr[:200]}")

                # Fallback: try 7z
                if not extracted:
                    _7z_exe = None
                    for _p in (
                        Path(r"C:\Program Files\7-Zip\7z.exe"),
                        Path(r"C:\Program Files (x86)\7-Zip\7z.exe"),
                    ):
                        if _p.exists():
                            _7z_exe = str(_p)
                            break
                    if not _7z_exe:
                        import shutil as _sh
                        _7z_exe = _sh.which("7z")

                    if _7z_exe:
                        result = subprocess.run(
                            [_7z_exe, 'e', str(rar_path), f'-o{current_folder}', '-aoa'],
                            capture_output=True, text=True, timeout=60,
                            encoding='utf-8', errors='replace',
                        )
                        if result.returncode == 0:
                            logger.info(f"Extracted RAR via 7z")
                            extracted = True
                        else:
                            logger.warning(f"7z failed (rc={result.returncode}): {result.stderr[:200]}")

                if extracted:
                    try:
                        rar_path.unlink()
                        logger.info(f"Deleted RAR after extract: {rar_path.name}")
                    except Exception as e:
                        logger.error(f"** Failed to delete RAR {rar_path.name}: {e}")
                else:
                    logger.warning(f"⚠️ Cannot extract {rar_path.name}: UnRAR and 7z not found")
                    logger.info(f"→ Install WinRAR or 7-Zip for RAR support")

            except Exception as e:
                logger.error(f"** Failed to extract {rar_path.name}: {e}")
    
    logger.info(f"[FOLDER] Found {len(subfolders_to_process)} folder(s) to process\n")
    
    # Create NEW FILES output folder
    project_folder = Path(__file__).parent
    output_folder = project_folder / "NEW FILES"
    output_folder.mkdir(exist_ok=True)
    
    # Initialize Azure OpenAI client once
    logger.info("[SETUP] Initializing Azure OpenAI client...")
    client = _build_client()
    logger.info(f"[AI] Azure OpenAI: {AZURE_DEPLOYMENT}\n")
    
    cost_tracker = CostTracker(MODEL_INPUT_PRICE_PER_1M, MODEL_OUTPUT_PRICE_PER_1M)
    ocr_engine = MultiOCREngine()
    
    # Track execution time
    start_time = time.time()
    
    logger.info("[TIP] Press Ctrl+C during processing to skip the current file or stop")
    
    logger.info(f"[OUTPUT] Output folder: {output_folder}")
    
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    
    # נשמור נתונים לאיחוד בסוף
    all_results = []
    all_classifications = []
    folder_classifications_map = {}  # {target_folder: file_classifications} for TOSEND copy
    total_rafael_rows = 0
    
    # Global classification tracking
    global_classification_tokens_in = 0
    global_classification_tokens_out = 0
    global_classification_time = 0  # Total classification time across all folders
    classification_folder_count = 0  # Number of folders where classification occurred
    
    # Global Stage 6 (PL extraction) tracking
    global_stage6_tokens_in = 0
    global_stage6_tokens_out = 0
    
    # Track per-folder statistics for GUI display
    folder_stats = []
    
    # Wrap main processing loop in try-except to handle Ctrl+C gracefully
    try:
        # 
        # עיבוד כל תת-תיקייה בנפרד
        # 
        for folder_idx, target_folder in enumerate(subfolders_to_process, 1):
            logger.debug(f"\n✓ DEBUG: Starting folder loop iteration {folder_idx}")
            logger.info(f"Target folder: {target_folder.name}")
            
            folder_start_time = time.time()  # Track total time for this folder
            folder_classification_cost = 0
            folder_extraction_cost = 0
            folder_extraction_cost_accurate = 0.0  # Per-stage accurate cost accumulator
            folder_extraction_tokens_in = 0  # Track extraction tokens for this folder
            folder_extraction_tokens_out = 0  # Track extraction tokens for this folder
            folder_stage6_tokens_in = 0  # Track Stage 6 (PL) tokens for this folder
            folder_stage6_tokens_out = 0  # Track Stage 6 (PL) tokens for this folder
            subfolder_results = []  # Initialize here to avoid undefined variable error
            drawing_results = subfolder_results
            pl_items_list = []  # Initialize PL items list (will be populated in Stage 6)
            
            logger.info(f"{'='*70}")
            logger.info(f"Folder {folder_idx}/{len(subfolders_to_process)}: {target_folder.name}")
            logger.info(f"{'='*70}")
            
            # 
            # שלב 1: איסוף קבצים בתיקייה זו בלבד
            # 
            folder_files = []
            skipped_by_date = 0
            
            for file_path in target_folder.iterdir():
                if file_path.is_file():
                    name = file_path.name
                    if name.startswith('.') or name.startswith('~'):
                        continue
                    # Apply date filter
                    if date_range or after_date:
                        file_mtime = datetime.fromtimestamp(file_path.stat().st_mtime)
                        if date_range:
                            date_from, date_to = date_range
                            if not (date_from <= file_mtime <= date_to):
                                skipped_by_date += 1
                                continue
                        elif after_date:
                            if file_mtime < after_date:
                                skipped_by_date += 1
                                continue
                    folder_files.append(file_path)
            
            if skipped_by_date > 0:
                logger.info(f"Skipped {skipped_by_date} files (by date filter)")
            
            if not folder_files:
                logger.info(f"No files in this folder - skipping")
                continue
            
            logger.info(f"Found {len(folder_files)} files in folder\n")
            
            # 
            # שלב 2: זיהוי סוג כל קובץ בתיקייה זו
            # 
            logger.info("PHASE 1: File Type Classification")
            logger.info("=" * 70)
            
            classification_start_time = time.time()  # Track classification time for this folder
            file_classifications = []
            classification_tokens_in = 0
            classification_tokens_out = 0
            
            for idx, file_path in enumerate(folder_files, 1):
                logger.info(f"[{idx}/{len(folder_files)}]  {file_path.name}")
                
                file_type, description, quote_number, order_number, tokens_in, tokens_out = classify_file_type(str(file_path), client)
                
                classification_tokens_in += tokens_in
                classification_tokens_out += tokens_out
                
                # Skip ARCHIVE files - they will be extracted in Phase 0.5
                if file_type == 'ARCHIVE':
                    logger.info(f"Archive file - skipped (will be extracted automatically)")
                    continue
                
                # Extract item_number and revision for DISPLAY NAME
                item_number = ''
                revision = ''
                drawing_number = ''
                part_number = ''
                display_name = ''
                associated_item = ''  # Initialize for ALL files (will be set for PARTS_LIST linking)
                
                # תיקון חזק: אם קובץ הוא תמונה בפועל (jpg, png וכו'), תיקון הסיווג ל-3D_IMAGE
                ext_lower = file_path.suffix.lower()
                if ext_lower in {'.jpg', '.jpeg', '.png', '.tif', '.tiff', '.bmp', '.gif', '.webp'}:
                    if file_type != '3D_IMAGE':
                        logger.info(f"🔧 Correcting file type for {file_path.name}: {file_type} → 3D_IMAGE")
                        file_type = '3D_IMAGE'
                
                # For DRAWING type (_30), try to find drawing_number and revision from drawing_results
                if file_type == 'DRAWING' and drawing_results:
                    file_name = file_path.name
                    for dr in drawing_results:
                        if dr.get('file_name') == file_name:
                            drawing_number = dr.get('drawing_number', '')
                            part_number = dr.get('part_number', '')
                            revision = dr.get('revision', '')
                            item_number = drawing_number
                            if drawing_number and part_number:
                                # For drawings: drawing_number + TAB + part_number + TAB + revision + TAB (no extension)
                                display_name = f"{drawing_number} \t{part_number} \t{revision} \t"
                            break
                # For other types (_99), use filename + quote_number/order_number + revision
                elif file_type in ['PURCHASE_ORDER', 'QUOTE', 'INVOICE', 'PARTS_LIST', '3D_MODEL', '3D_IMAGE', 'OTHER']:
                    if quote_number:
                        item_number = quote_number
                    elif order_number:
                        item_number = order_number
                    
                    # Try to find associated drawing for non-drawing files
                    # This gives us drawing_number and revision from the related drawing
                    if drawing_results and file_type in ['PARTS_LIST', '3D_MODEL', 'OTHER']:
                        # Build drawing map for filename matching
                        temp_drawing_map = {}
                        for dr in drawing_results:
                            dr_file = dr.get('file_name', '')
                            dr_drawing_num = dr.get('drawing_number', '')
                            dr_part_num = dr.get('part_number', '')
                            dr_rev = dr.get('revision', '')
                            if dr_file and dr_part_num:
                                temp_drawing_map[dr_file] = {
                                    'part_number': dr_part_num,
                                    'drawing_number': dr_drawing_num,
                                    'revision': dr_rev
                                }
                        
                        # Find associated drawing by filename similarity
                        if temp_drawing_map:
                            associated_part = _find_associated_drawing(file_path, file_type, temp_drawing_map)
                            if associated_part:
                                # Find the drawing with this part_number to get drawing_number and revision
                                for dr in drawing_results:
                                    if dr.get('part_number', '') == associated_part:
                                        drawing_number = dr.get('drawing_number', '')
                                        revision = dr.get('revision', '')
                                        item_number = associated_part
                                        associated_item = associated_part  # Store for PL linking (Stage 6)
                                        break
                    
                    # For non-drawing files: build display_name
                    if item_number:
                        if file_type in ('3D_MODEL', '3D_IMAGE'):
                            # Same structure as DRAWING but with file extension at the end
                            file_ext = file_path.suffix
                            display_name = f"{drawing_number} \t{item_number} \t{revision} \t{file_ext}"
                        else:
                            original_name_no_ext = file_path.stem
                            display_name = f"{original_name_no_ext} \t{item_number} \t{revision} \t"
                
                file_classifications.append({
                    'file_path': file_path,
                    'original_filename': file_path.name,
                    'file_type': file_type,
                    'description': description,
                    'quote_number': quote_number if quote_number else '',
                    'order_number': order_number if order_number else '',
                    'item_number': item_number,
                    'revision': revision,
                    'drawing_number': drawing_number,
                    'part_number': part_number,
                    'associated_item': associated_item,
                    'display_name': display_name
                })
                
                # Display classification
                icon = {
                    'DRAWING': '',
                    'PURCHASE_ORDER': '',
                    'QUOTE': '',
                    'INVOICE': '',
                    'PARTS_LIST': '',
                    '3D_MODEL': '',
                    'ARCHIVE': '📦',
                    'OTHER': ''
                }.get(file_type, '')
                
                logger.info(f"{icon} Type: {file_type}")
                if description:
                    logger.info(f"{description}")
                if file_type == 'QUOTE' and quote_number:
                    logger.info(f"Quote Number: {quote_number}")
                if file_type in ['PURCHASE_ORDER', 'INVOICE'] and order_number:
                    logger.info(f"Order Number: {order_number}")
            
            # Add to global tracking
            classification_end_time = time.time()
            classification_elapsed = classification_end_time - classification_start_time
            global_classification_tokens_in += classification_tokens_in
            global_classification_tokens_out += classification_tokens_out
            global_classification_time += classification_elapsed
            classification_folder_count += 1
            
            # Statistics for this folder
            logger.info("" + "=" * 70)
            logger.info("Classification Summary (this folder):")
            logger.info("=" * 70)
            
            from collections import Counter
            type_counts = Counter(fc['file_type'] for fc in file_classifications)
            
            for file_type, count in sorted(type_counts.items(), key=lambda x: -x[1]):
                icon = {
                    'DRAWING': '',
                    'PURCHASE_ORDER': '',
                    'QUOTE': '',
                    'INVOICE': '',
                    'PARTS_LIST': '',
                    '3D_MODEL': '',
                    'ARCHIVE': '📦',
                    'OTHER': ''
                }.get(file_type, '')
                logger.info(f"{icon} {file_type}: {count}")
            
            logger.info(f"Classification cost (this folder):")
            class_input_cost = (classification_tokens_in / 1_000_000) * MODEL_RUNTIME.get_stage_input_price(STAGE_CLASSIFICATION)
            class_output_cost = (classification_tokens_out / 1_000_000) * MODEL_RUNTIME.get_stage_output_price(STAGE_CLASSIFICATION)
            class_total_cost = class_input_cost + class_output_cost
            folder_classification_cost = class_total_cost  # Store for folder stats
            logger.info(f"Input:  {classification_tokens_in:,} tokens (${class_input_cost:.4f})")
            logger.info(f"Output: {classification_tokens_out:,} tokens (${class_output_cost:.4f})")
            logger.info(f"Total:  ${class_total_cost:.4f} USD ({class_total_cost * 3.7:.2f})")
            logger.info(f"Time: {classification_elapsed:.1f} seconds")
            logger.info("=" * 70)
            
            # 
            # שלב 3: עיבוד שרטוטים בתיקייה זו
            # 
            # Filter for drawing files - safely handle None and invalid entries
            drawing_files = [fc for fc in file_classifications if fc and isinstance(fc, dict) and fc.get('file_type') == 'DRAWING']
            
            if not drawing_files:
                logger.info(f"No drawing files in this folder - skipping to next folder")
                # Still save classification for this folder (filter out None values)
                all_classifications.extend([fc for fc in file_classifications if fc and isinstance(fc, dict)])
                
                # Collect folder statistics even if no drawings
                from collections import Counter
                file_type_counts = Counter(fc.get('file_type', 'UNKNOWN') for fc in file_classifications if fc and isinstance(fc, dict))
                
                folder_end_time = time.time()
                folder_total_time = folder_end_time - folder_start_time
                
                folder_info = {
                    'name': target_folder.name,
                    'total_files': len(file_classifications),
                    'file_types': dict(file_type_counts),
                    'total_drawings': 0,
                    'confidence_high': 0,
                    'confidence_medium': 0,
                    'confidence_low': 0,
                    'classification_cost': folder_classification_cost,
                    'extraction_cost': 0,
                    'total_cost': folder_classification_cost,
                    'processing_time': folder_total_time
                }
                folder_stats.append(folder_info)
                
                continue
            
            # Filter out files that are too large
            filtered_drawing_files = []
            skipped_large_files = []
            
            for fc in drawing_files:
                metadata = _get_file_metadata(fc['file_path'])
                if metadata['is_too_large']:
                    skipped_large_files.append({
                        'file': fc['file_path'].name,
                        'size_mb': metadata['file_size_mb']
                    })
                    logger.info(f"SKIPPING (too large): {fc['file_path'].name} ({metadata['file_size_mb']}MB)")
                else:
                    if metadata['is_large_file']:
                        logger.warning(f"WARNING (large file): {fc['file_path'].name} ({metadata['file_size_mb']}MB)")
                    filtered_drawing_files.append(fc)
            
            if not filtered_drawing_files:
                logger.info(f"No processable drawing files (all too large) - skipping to next folder")
                # Filter out None values before extending
                all_classifications.extend([fc for fc in file_classifications if fc and isinstance(fc, dict)])
                
                # Collect folder statistics even if all files too large
                from collections import Counter
                file_type_counts = Counter(fc.get('file_type', 'UNKNOWN') for fc in file_classifications if fc and isinstance(fc, dict))
                
                folder_end_time = time.time()
                folder_total_time = folder_end_time - folder_start_time
                
                folder_info = {
                    'name': target_folder.name,
                    'total_files': len(file_classifications),
                    'file_types': dict(file_type_counts),
                    'total_drawings': len(drawing_files),  # Count drawing files even if not processed
                    'confidence_high': 0,
                    'confidence_medium': 0,
                    'confidence_low': 0,
                    'classification_cost': folder_classification_cost,
                    'extraction_cost': 0,
                    'total_cost': folder_classification_cost,
                    'processing_time': folder_total_time
                }
                folder_stats.append(folder_info)
                
                continue
            
            logger.info(f"PHASE 2: Processing {len(filtered_drawing_files)} Drawing Files")
            if skipped_large_files:
                logger.info(f"Skipped: {len(skipped_large_files)} files (too large)")
            logger.info("=" * 70)
            
            logger.debug(f"\nDEBUG: About to process {len(filtered_drawing_files)} drawing files")
            logger.debug(f"DEBUG: subfolder_results before processing = {len(subfolder_results)}")
            
            # 
            # Phase 1: process drawings - extract item_numbers
            # 
            logger.info("Phase 1: Processing drawings (extracting item details)...")
            
            for idx, fc in enumerate(filtered_drawing_files, 1):
                file_path = fc['file_path']
                logger.info(f"[{idx}/{len(filtered_drawing_files)}]  {file_path.name}")
                
                # Check if GUI requested stop BEFORE starting new file
                if _ocr_mod._gui_should_stop and _ocr_mod._gui_should_stop():
                    logger.info("Stopped by user from GUI")
                    logger.info("All results so far were already saved in subfolders")
                    raise KeyboardInterrupt
                # Check if GUI requested skip
                if _ocr_mod._gui_should_skip and _ocr_mod._gui_should_skip():
                    logger.info(f"Skipping {file_path.name} (requested by GUI)")
                    # Reset skip flag for next file
                    import customer_extractor_v3_dual
                    if hasattr(customer_extractor_v3_dual, '_gui_skip_reset'):
                        customer_extractor_v3_dual._gui_skip_reset()
                    continue
                
                cost_tracker.total_files += 1
                
                # Track execution time for this file
                file_start_time = time.time()
                
                try:
                    data, usage = extract_customer_name(
                        str(file_path),
                        client,
                        ocr_engine,
                        selected_stages=selected_stages,
                        enable_retry=enable_image_retry,
                        stage1_skip_retry_resolution_px=stage1_skip_retry_resolution_px
                    )
                    
                    # Check if GUI requested skip right after processing
                    if _ocr_mod._gui_should_skip and _ocr_mod._gui_should_skip():
                        logger.info(f"Skipping {file_path.name} (requested by GUI)")
                        # Reset skip flag
                        import customer_extractor_v3_dual
                        if hasattr(customer_extractor_v3_dual, '_gui_skip_reset'):
                            customer_extractor_v3_dual._gui_skip_reset()
                        continue
                
                except KeyboardInterrupt:
                    # User pressed Ctrl+C during processing
                    logger.info("Ctrl+C pressed!")
                    logger.info("Choose an action:")
                    logger.info("[S] Skip - skip current file and continue to next")
                    logger.info("[Q] Quit - stop all processing and save results")
                    logger.info("[C] Continue - continue processing current file")
                    
                    choice = input("   Your choice (S/Q/C): ").strip().upper()
                    
                    if choice == 'S':
                        logger.info(f"Skipping {file_path.name}...")
                        continue
                    elif choice == 'Q':
                        logger.info("Stopped by user")
                        logger.info("All results so far were already saved in subfolders")
                        raise KeyboardInterrupt
                    else:
                        logger.info(f"Continuing to process {file_path.name}...")
                        # Try again
                        data, usage = extract_customer_name(
                            str(file_path),
                            client,
                            ocr_engine,
                            selected_stages=selected_stages,
                            enable_retry=enable_image_retry,
                            stage1_skip_retry_resolution_px=stage1_skip_retry_resolution_px
                        )
                
                except Exception as file_error:
                    # Catch any non-KeyboardInterrupt exception so one bad file
                    # doesn't kill processing of the entire email/folder
                    logger.error(f"❌ FAILED to process {file_path.name}: {file_error}", exc_info=True)
                    continue
                
                # Calculate execution time and cost for this file
                file_end_time = time.time()
                file_execution_time = file_end_time - file_start_time
                file_cost = 0.0
                
                if usage:
                    # Use accurate per-stage cost from pipeline when available
                    file_cost = (data.get('_pipeline_cost_usd') or 0) if data else 0
                    if not file_cost:
                        # Fallback: base-model price (for older callers)
                        file_cost = (usage[0] / 1_000_000 * MODEL_INPUT_PRICE_PER_1M) + (usage[1] / 1_000_000 * MODEL_OUTPUT_PRICE_PER_1M)
                    cost_tracker.add_usage(usage[0], usage[1], cost=file_cost)
                    # Track extraction tokens for this folder
                    folder_extraction_tokens_in += usage[0]
                    folder_extraction_tokens_out += usage[1]
                    folder_extraction_cost_accurate += file_cost
                
                # Print summary for this file
                logger.info(f"Runtime: {file_execution_time:.2f}s | Cost: ${file_cost:.4f}")
                
                try:
                    if data and isinstance(data, dict):
                        cost_tracker.successful_files += 1
                        # Add execution time and cost to data
                        data['execution_time_seconds'] = round(file_execution_time, 2)
                        data['extraction_cost_usd'] = round(file_cost, 4)
                        data.pop('_pipeline_cost_usd', None)  # internal key
                        
                        # שמור את המידע מהשרטוט (ללא כמויות עדיין)
                        result_dict = {
                            # ============ מידע מהשרטוט ============
                            "file_name": file_path.name,
                            "customer_name": data.get("customer_name") or "",
                            "part_number": data.get("part_number") or "",
                            "item_name": data.get("item_name") or "",
                            "drawing_number": data.get("drawing_number") or "",
                            "revision": data.get("revision") or "",
                            "needs_review": data.get("needs_review") or "",
                            "confidence_level": data.get("confidence_level") or "",
                            "material": data.get("material") or "",
                            "coating_processes": data.get("coating_processes") or "",
                            "painting_processes": data.get("painting_processes") or "",
                            "colors": data.get("colors") or "",
                            "part_area": data.get("part_area") or "",
                            "specifications": data.get("specifications") or "",
                            "parts_list_page": data.get("parts_list_page") or "",
                            "inserts_hardware": enrich_inserts_with_prices(
                                validate_inserts_hardware(
                                    data.get("inserts_hardware") or [],
                                    part_number=data.get("part_number") or data.get("drawing_number") or "",
                                )
                            ),
                            "process_summary_hebrew": data.get("process_summary_hebrew") or "",
                            "process_summary_hebrew_short": data.get("process_summary_hebrew_short") or "",
                            "notes_full_text": data.get("notes_full_text") or "",
                            "num_pages": data.get("num_pages") or 1,
                            # ============ מידע מהזמנות/הצעות מחיר (יתמלא בשלב 2) ============
                            "quantity": "",  # יתמלא אחר כך
                            "quantity_match_type": "",  # ספציפי למספר פריט / כללי
                            "quantity_source": "",  # מייל / הצעה/הזמנה
                            "work_description_doc": "",  # תיאור עבודה מהזמנה/הצעה (יתמלא בשלב 2)
                            "work_description_email": "",  # תיאור עבודה מהמייל (יתמלא בשלב 2)
                            # ============ מידע מהמייל ============
                            "email_from": "",  # יתמלא בשלב 2
                            "email_subject": "",  # יתמלא בשלב 2
                            # ============ מידע טכני על איכות ============
                            "subfolder": target_folder.name,
                            "validation_warnings": data.get("validation_warnings") or "",
                            "image_resolution": data.get("image_resolution") or "",
                            "drawing_layout": data.get("drawing_layout") or "",
                            "quality_issues": data.get("quality_issues") or "",
                            "execution_time_seconds": data.get("execution_time_seconds") or 0,
                            "extraction_cost_usd": data.get("extraction_cost_usd") or 0
                        }
                        
                        # שמור בזיכרון לשלב 2
                        subfolder_results.append(result_dict)
                except Exception as result_error:
                    # Catch errors in result building (e.g. validate_inserts_hardware)
                    # so one bad result doesn't kill the entire batch
                    logger.error(f"❌ FAILED to build result for {file_path.name}: {result_error}", exc_info=True)
            
            # 
            # שלב 2: התאמת כמויות ותיאורים לאחר עיבוד כל השרטוטים
            # 
            # Defaults to avoid undefined variables when no drawings/email
            email_data = {
                'found': False,
                'subject': '',
                'general_quantities': [],
                'part_quantities': {},
                'work_description': '',
                'quantity_summary': ''
            }
            item_details = {}
            if subfolder_results:
                logger.info(f"[STAGE2] Stage 2: Match quantities and descriptions ({len(subfolder_results)} items)...")
                
                # קרא קובץ email אם קיים
                email_data = _read_email_content(target_folder, client)
            
            # Step 1: חלץ כמויות ותיאורי עבודה מהזמנות/הצעות
            # Ensure file_classifications is not None before passing to extraction function
            if file_classifications is None:
                logger.debug(f"DEBUG: file_classifications is None - skipping item details extraction")
                file_classifications = []
            
            item_details = _extract_item_details_from_documents(
                target_folder, file_classifications, client, email_data=email_data
            )
            
            # Step 2: PL processing moved to Stage 6 (separate workflow)
            
            general_quantities = []
            if email_data and email_data.get('found'):
                logger.info(f"Email: {email_data.get('subject', '')[:50]}...")
                general_quantities = email_data.get('general_quantities', [])
                if general_quantities:
                    logger.info(f"General quantities from email: {', '.join(general_quantities)}")
                
                if item_details:  # Only iterate if item_details is not empty
                    for key in item_details.keys():
                        if isinstance(item_details[key], dict) and 'quantities' in item_details[key]:
                            logger.info(f"'{key}' -> Qty: {item_details[key]['quantities']}")

                

            logger.debug(f"\nDEBUG: After processing drawings, subfolder_results = {len(subfolder_results)}")
            if len(subfolder_results) == 0:
                logger.debug("DEBUG: WARNING - NO RESULTS FROM DRAWING PROCESSING!")

            items_with_specific_qty, items_with_general_qty = _match_quantities_to_drawings(
                subfolder_results, item_details, email_data, general_quantities, pl_items_list
            )
            
            #  Fallback: match by PL filename for items without match
            logger.info("Trying PL filename matching for items without a match...")
            for pl_item in pl_items_list:
                if pl_item.get('matched_drawing'):
                    continue  # כבר יש התאמה
                
                # חלץ את חלק שם הקובץ (הסר .pdf וסיומות אחרות)
                pl_filename = pl_item.get('pl_filename', '')
                # הסר סיומות קובץ ותוויות revision (_a, _b, וכו')
                pl_filename_core = re.sub(r'\.[^.]*$', '', pl_filename)  # הסר .pdf/.xlsx וכו'
                pl_filename_core = re.sub(r'_[a-z]$', '', pl_filename_core, flags=re.IGNORECASE)  # הסר _a, _b וכו'
                # נרמל את שם קובץ ה-PL
                pl_filename_normalized = _normalize_item_number(pl_filename_core)
                
                # חפש התאמה בשרטוטים
                for result_dict in subfolder_results:
                    part_num = result_dict.get('part_number', '')
                    part_num_normalized = _normalize_item_number(part_num)
                    
# אם part_number משרטוט מופיע בשם קובץ ה-PL (עם עדינויות OCR)
                    if part_num_normalized and len(part_num_normalized) >= 5:
                        # בדוק שני כיוונים עם fuzzy matching (מטפל ב-O/0, I/1, l/1):
                        if (_fuzzy_substring_match(part_num_normalized, pl_filename_normalized) or
                            _fuzzy_substring_match(pl_filename_normalized, part_num_normalized)):
                            drawing_name = result_dict.get('file_name', '')
                            pl_item['matched_drawing'] = drawing_name
                            logger.info(f"Matched by PL filename: '{pl_item['part_number']}'  '{drawing_name}' (via '{pl_filename}')")
                            break
            
            logger.info(f"{items_with_specific_qty} items with specific quantity")
            if items_with_general_qty > 0:
                logger.info(f"{items_with_general_qty} items with general quantity")
            
            # 
            # עדכן associated_item עבור PL files בـ file_classifications אחרי שsubfolder_results מלא
            # זה חיוני כי ב-Stage 0 drawing_results לא היה קיים עדיין
            # 
            logger.debug(f"DEBUG: Before PL associated_item update - subfolder_results={len(subfolder_results) if subfolder_results else 0}, file_classifications={len(file_classifications) if file_classifications else 0}")
            
            if subfolder_results and file_classifications:
                # Safely count PL files, handling None values
                pl_file_count = len([fc for fc in file_classifications if fc and isinstance(fc, dict) and fc.get('file_type') == 'PARTS_LIST'])
                if pl_file_count > 0:
                    logger.info(f"Updating file_classifications with associated_item for {pl_file_count} PL files...")
                    
                    # Build drawing_map once (not per-PL)
                    temp_drawing_map = {}
                    for dr in subfolder_results:
                        dr_file = dr.get('file_name', '')
                        dr_part_num = dr.get('part_number', '')
                        if dr_file and dr_part_num:
                            temp_drawing_map[dr_file] = dr_part_num
                            dr_item = _extract_item_number_from_filename(dr_file)
                            if dr_item:
                                temp_drawing_map[dr_item] = dr_part_num
                    
                    # Collect all PLs that need association
                    pl_files = [fc for fc in file_classifications 
                                if fc and isinstance(fc, dict) 
                                and fc.get('file_type') == 'PARTS_LIST' 
                                and not fc.get('associated_item')]
                    
                    # Score ALL PLs against ALL drawings, then assign uniquely
                    all_scores = []  # [(pl_fc, part_number, pl_filename)]
                    
                    for fc in pl_files:
                        pl_filename = fc['file_path'].name if hasattr(fc.get('file_path'), 'name') else str(fc.get('file_path'))
                        associated_part = _find_associated_drawing(fc['file_path'], 'PARTS_LIST', temp_drawing_map)
                        if associated_part:
                            all_scores.append((fc, associated_part, pl_filename))
                    
                    # ── Unique assignment: if 2+ PLs got same drawing, resolve ──
                    from collections import defaultdict
                    by_pn = defaultdict(list)
                    for fc, pn, pl_name in all_scores:
                        by_pn[pn].append((fc, pl_name))
                    
                    already_assigned = set()  # part_numbers already taken
                    
                    for pn, pl_list in by_pn.items():
                        if len(pl_list) == 1:
                            # Only one PL wants this drawing — assign
                            fc, pl_name = pl_list[0]
                            fc['associated_item'] = pn
                            already_assigned.add(pn)
                            logger.info(f"PL: {pl_name} → associated_item: {pn}")
                        else:
                            # Multiple PLs want same drawing — pick best by PL hint match
                            logger.warning(f"⚠️ {len(pl_list)} PLs compete for {pn}: {[n for _, n in pl_list]}")
                            
                            best_fc = None
                            best_name = ""
                            best_overlap = 0
                            pn_clean = re.sub(r'[^a-z0-9]', '', pn.lower())
                            
                            for fc, pl_name in pl_list:
                                # Extract PL hint
                                pl_stem = fc['file_path'].stem.upper()
                                pl_clean = re.sub(r'^PL[_\-]?', '', pl_stem)
                                pl_clean = re.sub(r'[_\-][A-Z\-]{1,3}$', '', pl_clean)
                                pl_hint = re.sub(r'[^a-z0-9]', '', pl_clean.lower())
                                
                                # Count prefix overlap with PN
                                overlap = 0
                                for a, b in zip(pl_hint, pn_clean):
                                    if a == b:
                                        overlap += 1
                                    else:
                                        break
                                
                                if overlap > best_overlap:
                                    best_overlap = overlap
                                    best_fc = fc
                                    best_name = pl_name
                            
                            if best_fc:
                                best_fc['associated_item'] = pn
                                already_assigned.add(pn)
                                logger.info(f"PL: {best_name} → associated_item: {pn} (won conflict)")
                                
                                # Remaining PLs — try to find their second-best match
                                for fc, pl_name in pl_list:
                                    if fc.get('associated_item'):
                                        continue  # Already assigned
                                    # Rebuild map without already-assigned PNs
                                    filtered_map = {k: v for k, v in temp_drawing_map.items() 
                                                    if (v if isinstance(v, str) else v.get('part_number', '')) not in already_assigned}
                                    alt = _find_associated_drawing(fc['file_path'], 'PARTS_LIST', filtered_map)
                                    if alt:
                                        fc['associated_item'] = alt
                                        already_assigned.add(alt)
                                        logger.info(f"PL: {pl_name} → associated_item: {alt} (second choice)")
                                    else:
                                        logger.warning(f"PL: {pl_name} → no unique match found")
                else:
                    logger.info(f"No PL files in file_classifications to update")
            elif subfolder_results and not file_classifications:
                logger.error(f"Note: file_classifications is None (earlier stages may have failed), skipping PL update")
            
            logger.debug(f"DEBUG: After PL associated_item update, about to start Stage 6")
            
            # 
            # STAGE 6: Extract data from PL (Parts List) PDF files
            # Process each PL file identified during classification, analyze with Azure OpenAI
            # 
            logger.debug(f"\nDEBUG: Starting Stage 6...")
            logger.debug(f"DEBUG: file_classifications type={type(file_classifications)}, length={len(file_classifications) if file_classifications else 0}")
            
            pl_items_list = []
            # Safely filter for PL files, handling None entries
            pl_files = []
            if file_classifications:
                for fc in file_classifications:
                    if fc and isinstance(fc, dict) and fc.get('file_type') == 'PARTS_LIST':
                        pl_files.append(fc)
            
            logger.debug(f"DEBUG: Found {len(pl_files)} PL files to process")
            
            if pl_files:
                logger.info(f"[STAGE6] Stage 6: Extract PL data ({len(pl_files)} parts lists)...")
                for pl_fc in pl_files:
                    pl_path = pl_fc.get('file_path')
                    if not pl_path:
                        logger.error(f"ERROR: PL file has no file_path!")
                        continue
                    logger.info(f"Processing: {pl_path.name if hasattr(pl_path, 'name') else pl_path}")
                    try:
                        pl_extracted, tokens_in, tokens_out = extract_pl_data(str(pl_path), client, file_classifications)
                        if pl_extracted:
                            pl_items_list.extend(pl_extracted)
                            folder_stage6_tokens_in += tokens_in
                            folder_stage6_tokens_out += tokens_out
                            logger.info(f"Extracted {len(pl_extracted)} items from this PL (${_calculate_stage_cost(tokens_in, tokens_out, STAGE_PL):.4f})")
                    except Exception as pl_error:
                        logger.error(f"ERROR processing PL: {str(pl_error)[:100]}")
                
                # Match PL items to drawings using associated_item from file_classifications
                if subfolder_results and pl_items_list:
                    logger.info(f"Matching {len(pl_items_list)} PL items to {len(subfolder_results)} drawings via associated_item...")
                    for pl_item in pl_items_list:
                        associated_item = pl_item.get('associated_item', '')
                        if not associated_item:
                            continue
                        
                        associated_norm = _normalize_item_number(associated_item)
                        
                        # Find matching drawing by associated_item
                        for result_dict in subfolder_results:
                            part_num = result_dict.get('part_number', '')
                            part_num_norm = _normalize_item_number(part_num)
                            
                            # Check if associated_item matches this drawing's part_number
                            if (associated_norm == part_num_norm or 
                                associated_norm in part_num_norm or 
                                part_num_norm in associated_norm):
                                pl_item['matched_item_name'] = result_dict.get('item_name', part_num)
                                pl_item['matched_drawing_part_number'] = part_num
                                logger.info(f"PL item → drawing '{result_dict.get('item_name', '')}' (associated: {associated_item})")
                                break
                
                logger.info(f"Stage 6 complete: {len(pl_items_list)} PL items extracted")
                
                # Propagate pl_main_part_number from PL items to drawing results
                if subfolder_results and pl_items_list:
                    for result_dict in subfolder_results:
                        part_num = result_dict.get('part_number', '')
                        part_num_norm = _normalize_item_number(part_num)
                        for pl_item in pl_items_list:
                            associated = pl_item.get('associated_item', '')
                            associated_norm = _normalize_item_number(associated)
                            if (associated_norm == part_num_norm or
                                associated_norm in part_num_norm or
                                part_num_norm in associated_norm):
                                pl_pn = pl_item.get('pl_main_part_number', '')
                                if pl_pn:
                                    result_dict['pl_main_part_number'] = pl_pn
                                    break
                
                # ═══════════════════════════════════════════════════════
                # PL PART NUMBER OVERRIDE
                # If PL provides a reliable part number, override OCR
                # ═══════════════════════════════════════════════════════
                if subfolder_results and pl_items_list:
                    logger.info(f"PL Part Number Override check...")
                    
                    for result_dict in subfolder_results:
                        ocr_part_number = result_dict.get('part_number', '')
                        ocr_part_normalized = _normalize_item_number(ocr_part_number)
                        
                        # Find matching PL item for this drawing
                        matched_pl = None
                        for pl_item in pl_items_list:
                            pl_associated = _normalize_item_number(pl_item.get('associated_item', ''))
                            pl_matched = _normalize_item_number(pl_item.get('matched_drawing_part_number', ''))
                            
                            if pl_associated and (pl_associated == ocr_part_normalized or 
                                                   pl_associated in ocr_part_normalized or
                                                   ocr_part_normalized in pl_associated):
                                matched_pl = pl_item
                                break
                            if pl_matched and (pl_matched == ocr_part_normalized or
                                                pl_matched in ocr_part_normalized or
                                                ocr_part_normalized in pl_matched):
                                matched_pl = pl_item
                                break
                        
                        if not matched_pl:
                            continue
                        
                        pl_main_pn = matched_pl.get('pl_main_part_number', '')
                        
                        # Skip if no PL part number or MULTIPLE
                        if not pl_main_pn or pl_main_pn == 'MULTIPLE':
                            if pl_main_pn == 'MULTIPLE':
                                result_dict['pl_part_number'] = 'MULTIPLE'
                                result_dict['pl_override_note'] = 'PL מכיל מספר פריטים מיוצרים'
                            continue
                        
                        # Validate: PL part number must look like a real P.N.
                        # Must contain at least 3 digits and be at least 6 chars
                        digit_count = sum(1 for c in pl_main_pn if c.isdigit())
                        if digit_count < 3 or len(pl_main_pn) < 6:
                            logger.warning(f"⚠️ Skipping PL PN '{pl_main_pn}' — too short or too few digits")
                            continue
                        
                        # Reject revision-like values: REV:A-001, REV:B-001, REV:-001
                        if pl_main_pn.upper().startswith('REV'):
                            logger.warning(f"⚠️ Skipping PL PN '{pl_main_pn}' — looks like revision, not P.N.")
                            continue
                        
                        # Reject common words that are not real part numbers
                        _INVALID_PN_WORDS = {
                            'category', 'description', 'part', 'number', 'rev', 'revision',
                            'catalog', 'seq', 'item', 'type', 'make', 'buy', 'qty',
                            'status', 'release', 'name', 'unit', 'level', 'none', 'null',
                        }
                        if pl_main_pn.lower() in _INVALID_PN_WORDS or len(pl_main_pn) < 3:
                            logger.warning(f"⚠️ Skipping invalid PL part number: '{pl_main_pn}'")
                            continue
                        
                        pl_main_normalized = _normalize_item_number(pl_main_pn)
                        
                        # Compare OCR vs PL
                        if ocr_part_normalized == pl_main_normalized:
                            # Normalized match — check if RAW strings differ
                            if ocr_part_number.strip() != pl_main_pn.strip():
                                # Raw differs → PL has more info (e.g. dash-number)
                                old_part = ocr_part_number
                                result_dict['part_number_ocr_original'] = old_part
                                result_dict['part_number'] = pl_main_pn
                                result_dict['pl_part_number'] = pl_main_pn
                                result_dict['pl_override_note'] = f'AS PL (OCR: {old_part})'
                                # Also fix drawing_number if it was the same bad OCR value
                                dwg = result_dict.get('drawing_number', '')
                                if dwg and _normalize_item_number(dwg) == _normalize_item_number(old_part):
                                    base, suffix = _extract_base_and_suffix(pl_main_pn)
                                    result_dict['drawing_number'] = base if suffix else pl_main_pn
                                logger.info(f"✓ OVERRIDE (dash-number): '{old_part}' → '{pl_main_pn}'")
                                continue
                            # Truly identical
                            result_dict['pl_part_number'] = pl_main_pn
                            result_dict['pl_override_note'] = 'AS PL'
                            logger.info(f"✓ CONFIRMED: '{ocr_part_number}' matches PL (AS PL)")
                            continue
                        
                        # ── OVERRIDE ──
                        old_part = ocr_part_number
                        result_dict['part_number_ocr_original'] = old_part
                        result_dict['part_number'] = pl_main_pn
                        result_dict['pl_part_number'] = pl_main_pn
                        result_dict['pl_override_note'] = f'מ-PL (OCR מקורי: {old_part})'
                        # Also fix drawing_number if it was the same bad OCR value
                        dwg = result_dict.get('drawing_number', '')
                        if dwg and _normalize_item_number(dwg) == _normalize_item_number(old_part):
                            base, suffix = _extract_base_and_suffix(pl_main_pn)
                            result_dict['drawing_number'] = base if suffix else pl_main_pn
                        
                        logger.info(f"✓ OVERRIDE: '{old_part}' → '{pl_main_pn}' (from PL)")
                    
                    # Count overrides
                    overrides = sum(1 for r in subfolder_results if r.get('part_number_ocr_original'))
                    if overrides:
                        logger.info(f"═══ {overrides} part numbers overridden from PL ═══")
                    else:
                        logger.info(f"No overrides needed (OCR matches PL)")

                    # ── EMAIL P.N. OVERRIDE ──
                    _is_iai_email = any(
                        'IAI' in str(r.get('customer_name', '')).upper()
                        for r in subfolder_results if r.get('customer_name')
                    )
                    _override_pn_from_email(subfolder_results, email_data, is_iai=_is_iai_email)
                
                # ═══════════════════════════════════════════════════════
                # UPDATE associated_item in file_classifications
                # So classification Excel + rename use corrected part number
                # ═══════════════════════════════════════════════════════
                if subfolder_results and file_classifications:
                    override_map = {}
                    for result_dict in subfolder_results:
                        ocr_original = result_dict.get('part_number_ocr_original', '')
                        if ocr_original:
                            new_pn = result_dict.get('part_number', '')
                            override_map[_normalize_item_number(ocr_original)] = new_pn
                    
                    if override_map:
                        for fc in file_classifications:
                            assoc = fc.get('associated_item', '')
                            assoc_norm = _normalize_item_number(assoc)
                            if assoc_norm in override_map:
                                old_assoc = assoc
                                fc['associated_item'] = override_map[assoc_norm]
                                logger.info(f"✓ Updated associated_item: '{old_assoc}' → '{fc['associated_item']}'")
            else:
                pl_items_list = []
            
            # ═══════════════════════════════════════════════════════
            # Propagate PL Summary Hebrew → subfolder_results
            # (needed for Stage 9 merge before Excel save)
            # ═══════════════════════════════════════════════════════
            if subfolder_results and pl_items_list:
                for result_dict in subfolder_results:
                    part_num_norm = _normalize_item_number(result_dict.get('part_number', ''))
                    ocr_orig_norm = _normalize_item_number(result_dict.get('part_number_ocr_original', ''))
                    heb_parts = []
                    pl_hw_parts = []
                    for pl_item in pl_items_list:
                        if not pl_item:
                            continue
                        assoc_norm = _normalize_item_number(str(pl_item.get('associated_item', '')))
                        # ── Match by associated_item OR fallback: single drawing gets all PL ──
                        matched = False
                        if assoc_norm and (
                            assoc_norm == part_num_norm or assoc_norm in part_num_norm or part_num_norm in assoc_norm
                            or (ocr_orig_norm and (assoc_norm == ocr_orig_norm or assoc_norm in ocr_orig_norm or ocr_orig_norm in assoc_norm))
                        ):
                            matched = True
                        elif not assoc_norm and len(subfolder_results) == 1:
                            # Fallback: PL item has no associated_item, but only 1 drawing → assign
                            matched = True
                        if matched:
                            heb = pl_item.get('pl_summary_hebrew', '')
                            if heb and heb not in heb_parts:
                                heb_parts.append(heb)
                            pl_hw = pl_item.get('pl_hardware', '')
                            if pl_hw and pl_hw not in pl_hw_parts:
                                pl_hw_parts.append(pl_hw)
                    if heb_parts:
                        result_dict['PL Summary Hebrew'] = ' | '.join(heb_parts)
                        logger.info(f"PL Summary Hebrew propagated to '{result_dict.get('part_number','')}': {len(heb_parts)} parts")
                    if pl_hw_parts:
                        result_dict['PL Hardware'] = ' | '.join(pl_hw_parts)

            # ═══════════════════════════════════════════════════════
            # STAGE 9: Smart Description Merge (o4-mini)
            # ═══════════════════════════════════════════════════════
            enable_stage9 = bool(selected_stages.get(9, True)) if selected_stages else True
            if enable_stage9 and subfolder_results and client:
                logger.info(f"[STAGE9] Stage 9: Smart merge descriptions ({len(subfolder_results)} items)...")
                try:
                    merged_count, s9_tok_in, s9_tok_out = _merge_descriptions(subfolder_results, client)
                    s9_cost = _calculate_stage_cost(s9_tok_in, s9_tok_out, STAGE_DESCRIPTION_MERGE)
                    cost_tracker.add_usage(s9_tok_in, s9_tok_out, cost=s9_cost)
                    logger.info(f"[STAGE9] Merged {merged_count} items (${s9_cost:.4f})")
                except Exception as s9_err:
                    logger.warning(f"[STAGE9] Failed: {s9_err}")
            elif not enable_stage9:
                logger.info("[STAGE9] Stage 9 disabled — skipping description merge")

            #  שמור קבצים בתוך התיקייה
            logger.debug(f"DEBUG: subfolder_results has {len(subfolder_results) if subfolder_results else 0} items")
            if subfolder_results:
                logger.info(f"Saving files in folder '{target_folder.name}'...")
                
                # יצירת שם תיקייה תקין לשם קובץ
                safe_folder_name = target_folder.name.replace(" ", "_").replace("/", "_").replace("\\", "_")
                
                # 1. קובץ ניתוח שרטוטים
                results_file = target_folder / f"drawing_results_{safe_folder_name}.xlsx"
                logger.debug(f"DEBUG: About to save Excel to {results_file}")
                _save_results_to_excel(subfolder_results, results_file, pl_items_list)
                logger.info(f"✓ Drawing analysis: {results_file.name} ({len(subfolder_results)} drawings)")
            else:
                logger.debug(f"DEBUG: subfolder_results is EMPTY - skipping drawing_results save")
            
            # 2. קובץ מיפוי קבצים (רק של תיקייה זו) - INDEPENDENT של subfolder_results!
            if file_classifications:
                logger.debug(f"\nDEBUG: About to save file_classifications ({len(file_classifications)} files)")
                # Backfill item_number/revision for drawings using actual results
                if subfolder_results:
                    drawing_index = {
                        dr.get('file_name'): dr for dr in subfolder_results
                        if dr.get('file_name')
                    }
                    for fc in file_classifications:
                        # Store original filename before rename
                        if 'original_filename' not in fc:
                            fc['original_filename'] = fc['file_path'].name
                        if fc.get('file_type') == 'DRAWING':
                            dr = drawing_index.get(fc['file_path'].name)
                            if dr:
                                fc['item_number'] = dr.get('drawing_number', '')
                                fc['revision'] = dr.get('revision', '')
                
                safe_folder_name = target_folder.name.replace(" ", "_").replace("/", "_").replace("\\", "_")
                classification_file = target_folder / f"file_classification_{safe_folder_name}.xlsx"
                drawing_map = _build_drawing_part_map(file_classifications, subfolder_results)
                _save_classification_report(file_classifications, target_folder, 
                                          0, 0,  # No tokens for individual report
                                          custom_filename=classification_file.name,
                                          drawing_map=drawing_map,
                                          drawing_results=subfolder_results)
                logger.info(f"✓ File mapping: {classification_file.name} ({len(file_classifications)} files)")
                
                # Post-process: Update PL sheet with associated_items from file_classifications
                # Classification file was just saved, so it should exist
                if subfolder_results:  # Only if we have results to update
                    results_file = target_folder / f"drawing_results_{safe_folder_name}.xlsx"
                    if classification_file.exists() and results_file.exists():
                        logger.info(f"Updating Parts_List_Items with associated_items...")
                        _update_pl_sheet_with_associated_items(results_file, classification_file)
                    elif not results_file.exists():
                        logger.info(f"Note: drawing_results file not found ({results_file.name}), skipping associated_item update")
                    else:
                        logger.info(f"Note: classification file not found, skipping associated_item update")
            else:
                logger.debug(f"DEBUG: file_classifications is EMPTY - skipping classification save")
            
            # 3. קובץ טקסט מסכם (TAB delimited) - INDEPENDENT of both files!
            if subfolder_results:
                # קרא כתובת מייל מקובץ email.txt
                sender_email = ""
                subject_text = ""
                email_file = target_folder / "email.txt"
                if email_file.exists():
                    try:
                        with open(email_file, 'r', encoding='utf-8', errors='ignore') as f:
                            lines = f.readlines()
                        if lines:
                            # בדוק אם השורה הראשונה היא כתובת מייל
                            first_line = lines[0].strip()
                            if first_line and "@" in first_line:
                                sender_email = first_line.replace("כתובת שולח:", "").replace("From:", "").strip()
                        
                        # חיפוש נוסף ב-Subject/נושא
                        for line in lines:
                            line_strip = line.strip()
                            if line_strip.lower().startswith("from:") and "@" in line_strip and not sender_email:
                                sender_email = line_strip.split(":", 1)[1].strip()
                                continue
                            if line_strip.lower().startswith("subject:"):
                                subject_text = line_strip.split(":", 1)[1].strip()
                                break
                            if line_strip.startswith("נושא:"):
                                subject_text = line_strip.split(":", 1)[1].strip()
                                break
                    except Exception:
                        pass
                
                # תחול את מידע המייל לכל הצירופים (results)
                for result_dict in subfolder_results:
                    if not result_dict.get('email_from'):  # רק אם עדיין לא הגדרנו
                        result_dict['email_from'] = sender_email
                        result_dict['email_subject'] = subject_text
                
                # חלץ מספר הצעה / הזנת רכש קודם מקובץ הסיווג (file_classifications)
                quote_or_order_number = ""
                
                if file_classifications and isinstance(file_classifications, list):
                    for fc in file_classifications:
                        if not fc or not isinstance(fc, dict):  # Skip None or invalid entries
                            continue
                        qn = str(fc.get('quote_number', '')).strip() if fc.get('quote_number') else ""
                        on = str(fc.get('order_number', '')).strip() if fc.get('order_number') else ""
                        if qn:
                            quote_or_order_number = qn
                            break
                        if on:
                            quote_or_order_number = on
                            break
                
                def _extract_number(text: str) -> str:
                    if not text:
                        return ""
                    # חיפוש לפי מילות מפתח (עברית/אנגלית)
                    patterns = [
                        r'(?:quotation|quote|rfq|הצעת מחיר|הצעה)[^0-9]{0,12}(\d{4,12})',
                        r'(?:order|po|הזמנת רכש|הזנת רכש)[^0-9]{0,12}(\d{4,12})',
                        r'B2B[_\s-]*Quotation[_\s-]*(\d{4,12})',
                        r'B2B[_\s-]*Order[_\s-]*(\d{4,12})',
                    ]
                    for pattern in patterns:
                        m = re.search(pattern, text, re.IGNORECASE)
                        if m:
                            return m.group(1)
                    # חיפוש כללי של מספר
                    m = re.search(r'(\d{5,12})', text)
                    return m.group(1) if m else ""

                def _looks_like_date(value: str) -> bool:
                    if not value or len(value) != 8 or not value.isdigit():
                        return False
                    try:
                        parsed = datetime.strptime(value, "%Y%m%d")
                        return 2000 <= parsed.year <= 2100
                    except Exception:
                        return False
                
                # אם לא נמצא בסיווג, נסה מהנושא ואז משם התיקייה
                if not quote_or_order_number:
                    candidate = _extract_number(subject_text)
                    if _looks_like_date(candidate):
                        candidate = ""
                    quote_or_order_number = candidate
                if not quote_or_order_number:
                    candidate = _extract_number(target_folder.name)
                    if _looks_like_date(candidate):
                        candidate = ""
                    quote_or_order_number = candidate
                if not quote_or_order_number:
                    quote_or_order_number = timestamp
                
                # צור שם קובץ טקסט עם מספר הצעה/הזנה
                b2b_number = "B2B-0_200002"
                text_filename = f"{b2b_number}-{quote_or_order_number}.txt"
                text_path = target_folder / text_filename
                
                # שמור קובץ טקסט עם מספר הצעה/הזנה כ-field 14 (יוצר שלושה קבצים: B2B, B2BH, B2BM)
                _save_text_summary_with_variants(subfolder_results, text_path, sender_email, b2b_number, quote_or_order_number)
                
                # שמור לאיחוד בסוף
                # Filter out None values before extending
                if file_classifications:
                    all_classifications.extend([fc for fc in file_classifications if fc and isinstance(fc, dict)])
                all_results.extend(subfolder_results)
                
                # Save file_classifications for this folder to use in TOSEND copy (filter None values)
                if file_classifications:
                    folder_classifications_map[target_folder] = [fc for fc in file_classifications if fc and isinstance(fc, dict)]
                else:
                    folder_classifications_map[target_folder] = []
                
                # Collect folder statistics for GUI (for folders with drawings that were saved)
                from collections import Counter
                file_type_counts = Counter(fc['file_type'] for fc in file_classifications)
                
                # Count confidence levels in drawings
                drawing_confidence_counts = {'HIGH': 0, 'MEDIUM': 0, 'LOW': 0}
                for result in subfolder_results:
                    conf_level = str(result.get('confidence_level', '')).strip().upper()
                    # Treat FULL as HIGH (FULL is highest confidence level)
                    if conf_level == 'FULL':
                        conf_level = 'HIGH'
                    if conf_level in drawing_confidence_counts:
                        drawing_confidence_counts[conf_level] += 1
                
                # Calculate extraction cost — prefer accurate per-stage cost
                if folder_extraction_cost_accurate > 0:
                    folder_extraction_cost = folder_extraction_cost_accurate
                else:
                    # Fallback: base-model price
                    folder_extraction_cost = ((folder_extraction_tokens_in / 1_000_000) * MODEL_INPUT_PRICE_PER_1M +
                                             (folder_extraction_tokens_out / 1_000_000) * MODEL_OUTPUT_PRICE_PER_1M)
                
                # Calculate Stage 6 cost from tracked tokens
                folder_stage6_cost = _calculate_stage_cost(folder_stage6_tokens_in, folder_stage6_tokens_out, STAGE_PL)
                
                folder_end_time = time.time()
                folder_total_time = folder_end_time - folder_start_time
                folder_total_cost = folder_classification_cost + folder_extraction_cost + folder_stage6_cost
                
                folder_info = {
                    'name': target_folder.name,
                    'total_files': len(file_classifications),
                    'file_types': dict(file_type_counts),
                    'total_drawings': len(subfolder_results),
                    'confidence_high': drawing_confidence_counts['HIGH'],
                    'confidence_medium': drawing_confidence_counts['MEDIUM'],
                    'confidence_low': drawing_confidence_counts['LOW'],
                    'classification_cost': folder_classification_cost,
                    'extraction_cost': folder_extraction_cost,
                    'stage6_cost': folder_stage6_cost,
                    'total_cost': folder_total_cost,
                    'processing_time': folder_total_time
                }
                folder_stats.append(folder_info)
                
                # Rename files in this folder based on classification and associated item BEFORE copying
                renamed_count = _rename_files_by_classification(file_classifications)
                if renamed_count > 0:
                    logger.info(f"Total renamed: {renamed_count} files")
        
        # All results processed with quantities
        if subfolder_results:
            logger.info(f"Folder '{target_folder.name}' finished - {len(subfolder_results)} drawings")
        else:
            logger.info(f"Folder '{target_folder.name}' - no drawings")
        
        # Add folder Stage 6 costs to global tracking
        global_stage6_tokens_in += folder_stage6_tokens_in
        global_stage6_tokens_out += folder_stage6_tokens_out
    
    except KeyboardInterrupt:
        # User pressed Q (Quit)
        logger.info("Stopped by user...")
        logger.info("All results so far were already saved in subfolders")
    
    # Copy all folders to TO_SEND AFTER all processing is complete
    if tosend_folder:
        logger.info(f"{'='*70}")
        logger.info("Copying all processed folders to TO_SEND...")
        logger.info(f"{'='*70}")
        for target_folder in subfolders_to_process:
            # Copy folder to TO_SEND with file_classifications from this folder
            fc_list = folder_classifications_map.get(target_folder, None)
            _copy_folder_to_tosend(target_folder, Path(tosend_folder), fc_list, confidence_level, all_results)
    
    #  איחוד כל הקבצים ל-NEW FILES (קריאה מהקבצים שנוצרו בתת-תיקיות)
    logger.info(f"{'='*70}")
    logger.info("Merging files from all folders...")
    logger.info(f"{'='*70}")
    
    # איחוד קבצי ניתוח שרטוטים - קריאה מהתת-תיקיות
    logger.info("Merging drawing_results files...")
    all_drawing_results = []
    all_pl_items = []  # Note: This is kept for debugging/logging only - actual PL data is copied directly from files
    all_file_classifications = []  # Initialize for SUMMARY post-processing
    final_classification_path = None  # Initialize
    for target_folder in subfolders_to_process:
        # חפש קבצי drawing_results בתיקייה
        drawing_files = list(target_folder.glob("drawing_results_*.xlsx"))
        for df_file in drawing_files:
            try:
                # קרא גיליון ראשי (שרטוטים)
                df = pd.read_excel(df_file, sheet_name='Sheet1')
                results = df.to_dict('records')
                all_drawing_results.extend(results)
                logger.info(f"Read {len(results)} rows from {df_file.name}")
                
                # נסה לקרוא גיליון PL Items אם קיים
                try:
                    df_pl = pd.read_excel(df_file, sheet_name='Parts_List_Items')
                    if not df_pl.empty:
                        pl_items = df_pl.to_dict('records')
                        # Normalize column names (pandas reads them with spaces)
                        # Headers are: 'PL Filename', 'Item Number', 'Description', 'Matched Item', 'Drawing Part Number', etc.
                        normalized_items = []
                        for item in pl_items:
                            # Rename columns to match our internal format
                            normalized = {
                                'item_number': item.get('Item Number') or item.get('item_number'),
                                'description': item.get('Description') or item.get('description'),
                                'quantity': item.get('Quantity') or item.get('quantity'),
                                'pl_filename': item.get('PL Filename') or item.get('pl_filename'),
                                'matched_item_name': item.get('Matched Item') or item.get('matched_item_name'),
                                'matched_drawing_part_number': item.get('Drawing Part Number') or item.get('matched_drawing_part_number')
                            }
                            normalized_items.append(normalized)
                        
                        # סנן שורות ריקות
                        pl_items_filtered = [item for item in normalized_items if item.get('item_number')]
                        if pl_items_filtered:
                            all_pl_items.extend(pl_items_filtered)
                            logger.info(f"Read {len(pl_items_filtered)} PL items from {df_file.name}")
                except Exception as e:
                    # אין גיליון PL Items בקובץ זה - עובר הלאה
                    logger.info(f"(No PL sheet in this file: {str(e)[:50]})")
                    pass
            except Exception as e:
                logger.error(f"Failed to read {df_file.name}: {e}")
    
    if all_drawing_results:
        final_results_path = output_folder / f"SUMMARY_all_results_{timestamp}.xlsx"
        
        # First save the main results sheet
        _save_results_to_excel(all_drawing_results, final_results_path, None)
        
        # Now add PL Items sheet by reading directly from sub-folder files
        logger.info(f"Adding PL Items to summary file from sub-folders...")
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            
            all_pl_rows = []
            
            # Read all PL Items sheets from drawing_results files in sub-folders
            for target_folder in subfolders_to_process:
                drawing_files = list(target_folder.glob("drawing_results_*.xlsx"))
                for df_file in drawing_files:
                    try:
                        # Try to read PL Items sheet
                        df_pl = pd.read_excel(df_file, sheet_name='Parts_List_Items')
                        if not df_pl.empty:
                            # Convert to list of lists (rows)
                            rows = df_pl.values.tolist()
                            all_pl_rows.extend(rows)
                            logger.info(f"Copied {len(rows)} rows from {df_file.name}")
                    except Exception:
                        # No PL Items sheet in this file
                        pass
            
            # If we have PL items, add them to SUMMARY file
            if all_pl_rows:
                wb = load_workbook(final_results_path)
                
                # Check if sheet exists, if not create it
                if "Parts_List_Items" in wb.sheetnames:
                    # Sheet exists - append rows
                    ws_pl = wb["Parts_List_Items"]
                    start_row = ws_pl.max_row + 1
                else:
                    # Create new sheet with headers
                    ws_pl = wb.create_sheet(title="Parts_List_Items")
                    
                    # Add headers (matching Stage 6 output format)
                    headers = ['PL Filename', 'Item Number', 'Description', 'Associated Item', 'Matched Item', 'Drawing Part Number',
                              'Quantity', 'Processes', 'Specifications', 'Product Tree', 'Item Type']
                    ws_pl.append(headers)
                    
                    # Format header row
                    from openpyxl.styles import Font, PatternFill, Alignment
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    for cell in ws_pl[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
                    start_row = 2
                
                # Add all rows
                for row in all_pl_rows:
                    ws_pl.append(row)
                
                # Set column widths
                col_widths = [20, 15, 25, 18, 20, 18, 12, 25, 30, 35, 15]
                for idx, width in enumerate(col_widths, 1):
                    ws_pl.column_dimensions[chr(64 + idx)].width = width
                    ws_pl.column_dimensions[chr(64 + idx)].alignment = Alignment(wrap_text=True)
                
                wb.save(final_results_path)
                wb.close()
                logger.info(f"Added PL Items sheet with {len(all_pl_rows)} rows to SUMMARY")
            
        except Exception as e:
            logger.error(f"Error adding PL sheet to SUMMARY: {e}")
        
        # Post-process SUMMARY file: Update PL sheet with associated_items
        if all_file_classifications and final_classification_path and final_results_path and final_results_path.exists():
            logger.info(f"Updating SUMMARY Parts_List_Items with associated_items...")
            _update_pl_sheet_with_associated_items(final_results_path, final_classification_path)
        
        logger.info(f"✅ Combined drawing analysis: {final_results_path.name} ({len(all_drawing_results)} drawings)")
        if all_pl_items:
            logger.info(f"Includes {len(all_pl_items)} combined PL items")
        
        # צור קובץ טקסט מסכם (TAB delimited)
        logger.info("📄 Creating summary text file...")
        
        # נסה לקרוא כתובת מייל מקובץ email.txt בתיקייה הראשונה
        customer_email = ""
        for target_folder in subfolders_to_process:
            email_file = target_folder / "email.txt"
            if email_file.exists():
                try:
                    with open(email_file, 'r', encoding='utf-8', errors='ignore') as f:
                        first_line = f.readline().strip()
                        # השורה הראשונה היא הכתובת
                        if first_line and "@" in first_line:
                            # נקה פורמט ישן אם קיים
                            customer_email = first_line.replace("כתובת שולח:", "").replace("From:", "").strip()
                            break
                except Exception:
                    pass
        
        # קבע את הערך לשדה 14 ושם הקובץ (quote/order number או timestamp)
        # קודם בדוק אם יש quote_number או order_number בתוצאות
        request_id = timestamp  # ברירת מחדל
        for item in all_drawing_results:
            if item.get('quote_number'):
                request_id = item['quote_number']
                break
            elif item.get('order_number'):
                request_id = item['order_number']
                break
        
        # צור שם קובץ
        b2b_number = "B2B-0_200002"  # ניתן להגדיר דינמית
        text_filename = f"{b2b_number}-{request_id}.txt"
        text_path = output_folder / text_filename
        
        # שמור קובץ טקסט (יוצר שלושה קבצים: B2B, B2BH, B2BM)
        _save_text_summary_with_variants(all_drawing_results, text_path, customer_email, b2b_number, request_id)
    else:
        final_results_path = None
        logger.warning("⚠️ No drawings to merge")
    
    # איחוד קבצי מיפוי קבצים - התחל מ-all_classifications (שכבר נבנה בלולאה)
    logger.info("Merging file_classification files...")
    
    # Use all_classifications that was already built during processing
    # but verify paths are Path objects
    for fc in all_classifications:
        # Ensure file_path is a Path object
        if 'file_path' in fc:
            if not isinstance(fc['file_path'], Path):
                fc['file_path'] = Path(fc['file_path'])
        all_file_classifications.append(fc)
    
    logger.info(f"Found {len(all_file_classifications)} file classifications from current run")
    
    # Additionally, read from existing Excel files in subfolders (for cached/previous runs)
    for target_folder in subfolders_to_process:
        # חפש קבצי file_classification בתיקייה
        classification_files = list(target_folder.glob("file_classification_*.xlsx"))
        for cf_file in classification_files:
            try:
                df = pd.read_excel(cf_file)
                # המר את השורות למבנה הנכון עם Path objects
                found_in_memory = False
                for _, row in df.iterrows():
                    file_path_str = str(row['file_path'])
                    # Check if already in all_file_classifications
                    if any(str(fc.get('file_path', '')) == file_path_str for fc in all_file_classifications):
                        found_in_memory = True
                        continue  # Skip duplicates
                    
                    classification_dict = {
                        'file_path': Path(file_path_str),
                        'file_type': row['file_type'],
                        'description': row['description'],
                        'quote_number': row.get('quote_number', ''),
                        'order_number': row.get('order_number', ''),
                        'associated_item': row.get('associated_item', ''),
                        'extension': row.get('extension', '')
                    }
                    all_file_classifications.append(classification_dict)
                if found_in_memory:
                    logger.info(f"Already in memory: {cf_file.name}")
                else:
                    logger.info(f"Read {len(df)} rows from {cf_file.name}")
            except Exception as e:
                logger.error(f"Failed to read {cf_file.name}: {e}")
    
    if all_file_classifications:
        final_classification_path = output_folder / f"SUMMARY_all_classifications_{timestamp}.xlsx"
        all_drawing_map = _build_drawing_part_map(all_file_classifications, all_results)
        _save_classification_report(all_file_classifications, output_folder, 
                                  global_classification_tokens_in, global_classification_tokens_out,
                                  custom_filename=f"SUMMARY_all_classifications_{timestamp}.xlsx",
                                  drawing_map=all_drawing_map,
                                  drawing_results=all_results)
        logger.info(f"Combined file mapping: {final_classification_path.name} ({len(all_file_classifications)} files)")
    else:
        final_classification_path = None
        logger.info("No mapping files to merge")
    
    # Calculate execution time
    end_time = time.time()
    execution_time = end_time - start_time
    
    try:
        cost_tracker.print_summary()
        logger.info(f"{ocr_engine.get_cache_stats()}")
        cost_summary = cost_tracker.get_summary()
    except Exception as e:
        logger.warning(f"Warning: Could not print cost summary: {e}")
        cost_summary = {
            'total_files': cost_tracker.total_files,
            'successful_files': cost_tracker.successful_files,
            'total_cost': 0,
            'execution_time': execution_time,
            'avg_time_per_drawing': 0
        }
    
    # Add execution time to summary
    cost_summary['execution_time'] = execution_time
    cost_summary['avg_time_per_drawing'] = execution_time / cost_summary['successful_files'] if cost_summary['successful_files'] > 0 else 0
    
    # SUMMARY file already exists and is up-to-date
    logger.info("" + "="*70)
    logger.info("The following files were saved successfully:")
    logger.info("="*70)
    if final_results_path and final_results_path.exists():
        logger.info(f"SUMMARY_all_results_{timestamp}.xlsx")
        logger.info(f"- {len(all_results)} drawings combined")
    if final_classification_path and final_classification_path.exists():
        logger.info(f"SUMMARY_all_classifications_{timestamp}.xlsx")
        logger.info(f"- {len(all_classifications)} files combined")
    logger.info("Additional files in each subfolder:")
    logger.info(f"- drawing_results_[subfolder].xlsx")
    logger.info(f"- file_classification_[subfolder].xlsx")
    logger.info("="*70)
    
    # Print final summary
    if all_results and final_results_path and final_results_path.exists():
        df = pd.DataFrame(all_results)
        
        logger.info("" + "="*70)
        logger.info(f"Final Summary Excel file: {final_results_path}")
        logger.info(f"Total files processed: {len(all_results)}")
        logger.info(f"Folders processed: {len(subfolders_to_process)}")
        if skipped_large_files:
            logger.info(f"Files skipped (too large): {len(skipped_large_files)}")
        if total_rafael_rows > 0:
            logger.info(f"Total RAFAEL rows highlighted: {total_rafael_rows}")
        
        # Statistics
        non_empty = lambda col: sum(1 for v in df[col] if v and str(v).strip())
        
        logger.info(f"Field Extraction Statistics:")
        logger.info(f"- Customer: {non_empty('customer_name')}/{len(all_results)}")
        logger.info(f"- Part#: {non_empty('part_number')}/{len(all_results)}")
        logger.info(f"- Item Name: {non_empty('item_name')}/{len(all_results)}")
        logger.info(f"- Drawing#: {non_empty('drawing_number')}/{len(all_results)}")
        logger.info(f"- Revision: {non_empty('revision')}/{len(all_results)}")
        
        # Confidence level statistics
        full_conf = sum(1 for v in df['confidence_level'] if v == 'full')
        high_conf = sum(1 for v in df['confidence_level'] if v == 'high')
        medium_conf = sum(1 for v in df['confidence_level'] if v == 'medium')
        low_conf = sum(1 for v in df['confidence_level'] if v == 'low')
        
        logger.info(f"Confidence Levels (based on filename matching):")
        logger.info(f"-  Full (exact match - both): {full_conf}")
        logger.info(f"-  High (part# in filename): {high_conf}")
        logger.info(f"-  Medium (only drawing# in filename): {medium_conf}")
        logger.info(f"-  Low (none in filename): {low_conf}")
        
        # Quality warnings
        needs_review_count = sum(1 for v in df['needs_review'] if v and str(v).strip() and "בעייתי" in str(v))
        needs_check_count = sum(1 for v in df['needs_review'] if v and str(v).strip() and "בדיקה" in str(v))
        if needs_review_count > 0 or needs_check_count > 0:
            logger.info(f"Quality Alerts:")
            if needs_review_count > 0:
                logger.info(f"- Problematic: {needs_review_count}")
            if needs_check_count > 0:
                logger.info(f"- Needs verification: {needs_check_count}")
        
        logger.info(f"- Material: {non_empty('material')}/{len(all_results)}")
        logger.info(f"- Coating: {non_empty('coating_processes')}/{len(all_results)}")
        logger.info(f"- Painting: {non_empty('painting_processes')}/{len(all_results)}")
        logger.info(f"- Colors: {non_empty('colors')}/{len(all_results)}")
        logger.info(f"- Area: {non_empty('part_area')}/{len(all_results)}")
        logger.info(f"- Specs: {non_empty('specifications')}/{len(all_results)}")
        logger.info(f"- Has PL: {non_empty('parts_list_page')}/{len(all_results)}")
        logger.info(f"- Process Summary: {non_empty('process_summary_hebrew')}/{len(all_results)}")
        logger.info(f"- Full Notes: {non_empty('notes_full_text')}/{len(all_results)}")
        
        # Execution time and performance
        logger.info(f"Execution Performance:")
        minutes = int(execution_time // 60)
        seconds = int(execution_time % 60)
        time_str = f"{minutes}m {seconds}s" if minutes > 0 else f"{seconds}s"
        logger.info(f"- Total execution time: {time_str}")
        logger.info(f"- Average per drawing: {cost_summary.get('avg_time_per_drawing', 0):.1f}s")
        
        # Cost information
        logger.info(f"API Costs:")
        logger.info(f"- Input tokens: {cost_summary.get('input_tokens', 0):,}")
        logger.info(f"- Output tokens: {cost_summary.get('output_tokens', 0):,}")
        logger.info(f"- Drawing extraction cost: ${cost_summary.get('total_cost', 0):.4f} USD ({cost_summary.get('total_cost_ils', 0):.2f})")
        logger.info(f"- Average per drawing: ${cost_summary.get('avg_cost', 0):.6f} USD ({cost_summary.get('avg_cost_ils', 0):.4f})")
        
        if cost_summary.get('classification_cost', 0) > 0:
            logger.info(f"Stage 0 (File Classification):")
            logger.info(f"- Input tokens: {cost_summary.get('classification_input_tokens', 0):,}")
            logger.info(f"- Output tokens: {cost_summary.get('classification_output_tokens', 0):,}")
            logger.info(f"- Cost: ${cost_summary.get('classification_cost', 0):.4f} USD ({cost_summary.get('classification_cost', 0)*3.7:.2f})")
        
        if cost_summary.get('stage6_cost', 0) > 0:
            logger.info(f"Stage 6 (PL Extraction):")
            logger.info(f"- Input tokens: {cost_summary.get('stage6_input_tokens', 0):,}")
            logger.info(f"- Output tokens: {cost_summary.get('stage6_output_tokens', 0):,}")
            logger.info(f"- Cost: ${cost_summary.get('stage6_cost', 0):.4f} USD ({cost_summary.get('stage6_cost', 0)*3.7:.2f})")
        
        total_all = cost_summary.get('total_cost_all', cost_summary.get('total_cost', 0))
        if total_all != cost_summary.get('total_cost', 0):
            logger.info(f"Total all stages: ${total_all:.4f} USD ({total_all*3.7:.2f})")
        
        logger.info("" + "="*70)
        logger.info(f"All files saved in: {output_folder}")
        logger.info("1 SUMMARY file (all results)")
        logger.info("1 classification file (file_classification)")
        logger.info("="*70)
    
    # Add classification costs to summary
    cost_summary['classification_input_tokens'] = global_classification_tokens_in
    cost_summary['classification_output_tokens'] = global_classification_tokens_out
    cost_summary['classification_cost'] = _calculate_stage_cost(
        global_classification_tokens_in,
        global_classification_tokens_out,
        STAGE_CLASSIFICATION,
    )
    cost_summary['total_cost_with_classification'] = cost_summary['total_cost'] + cost_summary['classification_cost']
    
    # Add Stage 6 (PL extraction) costs to summary
    cost_summary['stage6_input_tokens'] = global_stage6_tokens_in
    cost_summary['stage6_output_tokens'] = global_stage6_tokens_out
    cost_summary['stage6_cost'] = _calculate_stage_cost(
        global_stage6_tokens_in,
        global_stage6_tokens_out,
        STAGE_PL,
    )
    cost_summary['total_cost_all'] = cost_summary['total_cost'] + cost_summary['classification_cost'] + cost_summary['stage6_cost']
    
    cost_summary['classification_time'] = global_classification_time
    cost_summary['classification_folder_count'] = classification_folder_count
    cost_summary['avg_classification_time_per_folder'] = global_classification_time / classification_folder_count if classification_folder_count > 0 else 0
    cost_summary['folder_stats'] = folder_stats  # Per-folder statistics for GUI
    
    return all_results, output_folder, final_results_path, cost_summary, all_file_classifications  # Return classifications for file renaming


# Alias for GUI compatibility
scan_folder_with_stages = scan_folder


def main(folder: Optional[str] = None) -> None:
    """Main function"""
    logger.info("" + "="*70)
    logger.info("Customer Extractor V3.1 - File Classification + QUAD-STAGE")
    logger.info("="*70)
    logger.info("Features:")
    logger.info("Phase 0: Automatic file type classification")
    logger.info("Phase 1-4: Drawing processing (4 stages)")
    logger.info("Cost tracking & detailed reports")
    logger.warning(f"File size limits: Warn at {WARN_FILE_SIZE_MB}MB, Skip at {MAX_FILE_SIZE_MB}MB")
    logger.info("="*70)
    
    if folder is None:
        logger.info("Folder path:")
        folder = input("> ").strip()
        if not folder:
            folder = "."
    
    folder_path = Path(folder)
    
    # אם הטרמינל אינו אינטראקטיבי (לדוגמה בשילוב צינור/פייפ), דלג על הקלט
    import os, sys as _sys
    skip_date_prompt = os.environ.get("AI_DRAW_SKIP_DATE", "").lower() in {"1", "true", "yes"} or not _sys.stdin.isatty()
    if skip_date_prompt:
        date_input = ""
    else:
        logger.info("Filter by date? (DD/MM/YYYY or Enter to skip)")
        date_input = input("> ").strip()
    
    after_date = None
    if date_input:
        try:
            after_date = datetime.strptime(date_input, "%d/%m/%Y")
            logger.info(f"From {after_date.strftime('%d/%m/%Y')} onwards\n")
        except ValueError:
            logger.info("Invalid format. Processing all.\n")
    
    result = scan_folder(folder_path, recursive=True, after_date=after_date)
    
    if not result:
        logger.info("No results")
        return
    
    if isinstance(result, tuple) and len(result) == 5:
        results, project_folder, output_path, cost_summary, _all_file_classifications = result
    else:
        results, project_folder, output_path, cost_summary = result


if __name__ == "__main__":
    import sys
    folder = sys.argv[1] if len(sys.argv) > 1 else None
    main(folder)
