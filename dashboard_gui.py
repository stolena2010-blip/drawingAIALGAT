"""
DrawingAI Pro - Automation Dashboard
Shows statistics from automation_log.jsonl
"""
import json
import tkinter as tk
from tkinter import ttk
from pathlib import Path
from datetime import datetime, timedelta, timezone
from collections import Counter

try:
    import customtkinter as ctk
    HAS_CTK = True
except ImportError:
    HAS_CTK = False


class DashboardWindow:
    def __init__(self, parent) -> None:
        if HAS_CTK:
            self.win = ctk.CTkToplevel(parent)
        else:
            self.win = tk.Toplevel(parent)
        
        self.win.title("📊 Green Coat — DrawingAI Pro — Dashboard")
        self.win.geometry("900x700")
        self.win.minsize(800, 600)
        
        self._build_ui()
        self._load_and_display()
    
    def _load_log(self):
        """Load entries from automation_log.jsonl and all rotated/backup log files."""
        base_dir = Path(".")
        # Collect all matching log files: current + rotated + backups
        log_files = sorted(base_dir.glob("automation_log*.jsonl"))
        # Also include the .bak file if present
        bak = base_dir / "automation_log.jsonl.bak"
        if bak.exists() and bak not in log_files:
            log_files.append(bak)

        if not log_files:
            return []

        entries = []
        seen_ids = set()
        for log_path in log_files:
            try:
                with open(log_path, "r", encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()
                        if not line:
                            continue
                        try:
                            entry = json.loads(line)
                            # Deduplicate by id (same entry may appear in backup + current)
                            eid = entry.get("id")
                            if eid and eid in seen_ids:
                                continue
                            if eid:
                                seen_ids.add(eid)
                            entries.append(entry)
                        except json.JSONDecodeError:
                            continue
            except Exception:
                continue

        # Sort by timestamp ascending
        entries.sort(key=lambda e: e.get("timestamp", ""))
        return entries

    def _save_entry_field(self, entry_id: str, field: str, value) -> None:
        """Update a single field of an entry in automation_log.jsonl by its id."""
        log_path = Path("automation_log.jsonl")
        if not log_path.exists():
            return
        lines = log_path.read_text(encoding="utf-8").splitlines()
        new_lines = []
        for line in lines:
            if not line.strip():
                new_lines.append(line)
                continue
            try:
                entry = json.loads(line)
                if entry.get("id") == entry_id:
                    entry[field] = value
                    new_lines.append(json.dumps(entry, ensure_ascii=False))
                else:
                    new_lines.append(line)
            except json.JSONDecodeError:
                new_lines.append(line)
        log_path.write_text("\n".join(new_lines) + "\n", encoding="utf-8")

    @staticmethod
    def _utc_to_local(ts_str: str) -> str:
        """Convert a UTC timestamp string (ending with Z or without TZ) to local time string."""
        if not ts_str:
            return ""
        try:
            s = str(ts_str).strip()
            if s.endswith("Z"):
                s = s[:-1]
            dt_utc = datetime.fromisoformat(s).replace(tzinfo=timezone.utc)
            dt_local = dt_utc.astimezone()  # system local timezone
            return dt_local.strftime("%Y-%m-%dT%H:%M:%S")
        except Exception:
            return str(ts_str)

    def _local_ts(self, entry, field="timestamp") -> str:
        """Get a local-time timestamp string from an entry."""
        return self._utc_to_local(entry.get(field, ""))

    def _build_ui(self) -> None:
        """Build the dashboard UI."""
        # Header frame with Gollum icon
        header_container = tk.Frame(self.win, bg="#1a1a2e")
        header_container.pack(fill=tk.X)
        
        # Load company logo
        try:
            from PIL import Image, ImageTk
            import os
            icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "company_logo.png")
            if os.path.exists(icon_path):
                logo_img = Image.open(icon_path).resize((40, 40), Image.LANCZOS)
                self._logo_photo = ImageTk.PhotoImage(logo_img)
                logo_label = tk.Label(header_container, image=self._logo_photo, bg="#1a1a2e")
                logo_label.pack(side=tk.RIGHT, padx=(8, 4), pady=4)
        except Exception:
            pass
        
        if HAS_CTK:
            header = ctk.CTkLabel(
                header_container, text="Green Coat — DrawingAI Pro — Dashboard",
                font=("Arial", 20, "bold"), text_color="#00d4aa",
                fg_color="#1a1a2e", height=55, corner_radius=0
            )
        else:
            header = tk.Label(
                header_container, text="Green Coat — DrawingAI Pro — Dashboard",
                font=("Arial", 20, "bold"), bg="#1a1a2e", fg="#00d4aa", pady=14
            )
        header.pack(side=tk.RIGHT, fill=tk.X, expand=True)
        
        # Main scrollable area
        main_frame = ttk.Frame(self.win)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=16, pady=16)
        
        canvas = tk.Canvas(main_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        self.content = ttk.Frame(canvas)
        
        self.content.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=self.content, anchor="nw")
        canvas.bind("<Configure>", lambda e: canvas.itemconfigure("all", width=e.width))
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Buttons row
        btn_frame = tk.Frame(self.win)
        btn_frame.pack(pady=(0, 10), anchor="e")
        if HAS_CTK:
            ctk.CTkButton(
                btn_frame, text="🔄 רענן", command=self._load_and_display,
                fg_color="#0984e3", hover_color="#0773c5",
                width=120, height=36, corner_radius=8, font=("Arial", 13)
            ).pack(side=tk.RIGHT, padx=4)
            ctk.CTkButton(
                btn_frame, text="📊 ייצוא לאקסל", command=self._export_to_excel,
                fg_color="#00b894", hover_color="#00a381",
                width=160, height=36, corner_radius=8, font=("Arial", 13)
            ).pack(side=tk.RIGHT, padx=4)
            ctk.CTkButton(
                btn_frame, text="🗑️ איפוס סטטיסטיקה", command=self._reset_stats,
                fg_color="#7f1d1d", hover_color="#991b1b",
                width=180, height=36, corner_radius=8, font=("Arial", 13)
            ).pack(side=tk.RIGHT, padx=4)
        else:
            ttk.Button(btn_frame, text="🔄 רענן",
                      command=self._load_and_display).pack(side=tk.RIGHT, padx=4)
            ttk.Button(btn_frame, text="📊 ייצוא לאקסל",
                      command=self._export_to_excel).pack(side=tk.RIGHT, padx=4)
            ttk.Button(btn_frame, text="🗑️ איפוס סטטיסטיקה",
                      command=self._reset_stats).pack(side=tk.RIGHT, padx=4)
    
    def _load_and_display(self) -> None:
        """Load data and refresh all widgets."""
        # Clear existing content
        for widget in self.content.winfo_children():
            widget.destroy()
        
        entries = self._load_log()
        if not entries:
            ttk.Label(self.content, text="אין נתונים עדיין. הפעילי אוטומציה כדי לראות סטטיסטיקות.",
                     font=("Arial", 14)).pack(pady=40)
            return
        
        self._build_filter_bar()                   # G — filter at top
        entries = self._filter_entries(entries)     # G — apply filter
        
        self._show_summary_cards(entries)           # existing
        self._show_accuracy_metrics(entries)        # existing (Steps 1-3)
        self._show_accuracy_trend(entries)          # B — trend
        self._show_efficiency_metrics(entries)      # D — efficiency
        self._show_items_per_mail(entries)           # Items per mail stats
        self._show_pl_stats(entries)                # E — PL overrides
        self._show_customer_accuracy(entries)       # C — customer accuracy
        self._show_daily_breakdown(entries)         # existing
        self._show_top_errors(entries)              # F — errors
        self._show_top_senders(entries)             # existing
        self._show_recent_emails(entries)           # existing
    
    def _show_summary_cards(self, entries) -> None:
        """Show summary statistics cards."""
        cards_frame = ttk.LabelFrame(self.content, text="📈 סיכום כללי", padding=12)
        cards_frame.pack(fill=tk.X, pady=(0, 12))
        
        total = len(entries)
        
        # Get unique senders
        senders = set()
        for e in entries:
            sender = e.get("sender") or ""
            if sender:
                senders.add(sender.lower())
        
        # Date range
        dates = []
        for e in entries:
            ts = self._local_ts(e)
            if ts:
                dates.append(ts[:10])
        
        first_date = min(dates) if dates else "?"
        last_date = max(dates) if dates else "?"
        
        # Email accuracy (weighted average per email)
        import os as _os
        _w = {
            "full":   float(_os.getenv("ACCURACY_WEIGHT_FULL", "1.0")),
            "high":   float(_os.getenv("ACCURACY_WEIGHT_HIGH", "1.0")),
            "medium": float(_os.getenv("ACCURACY_WEIGHT_MEDIUM", "0.8")),
            "low":    float(_os.getenv("ACCURACY_WEIGHT_LOW", "0.5")),
            "none":   float(_os.getenv("ACCURACY_WEIGHT_NONE", "0.0")),
        }
        _email_scores = []
        for _e in entries:
            _acc = _e.get("accuracy_data", {})
            _tot = _acc.get("total", 0) if _acc else 0
            if _tot > 0:
                _sc = (
                    _acc.get("full", 0) * _w["full"] +
                    _acc.get("high", 0) * _w["high"] +
                    _acc.get("medium", 0) * _w["medium"] +
                    _acc.get("low", 0) * _w["low"] +
                    _acc.get("none", 0) * _w["none"]
                ) / _tot * 100
                _email_scores.append(_sc)
        email_accuracy = (sum(_email_scores) / len(_email_scores)) if _email_scores else None

        # Cost and timing aggregates
        total_cost = sum(e.get("cost_usd", 0) for e in entries if e.get("cost_usd"))
        entries_with_time = [e for e in entries if e.get("processing_time_seconds")]
        avg_time = (sum(e["processing_time_seconds"] for e in entries_with_time) / len(entries_with_time)) if entries_with_time else 0

        acc_display = f"{email_accuracy:.1f}%" if email_accuracy is not None else "—"

        # Human-verified match rate (human_verified: 1=correct, 0=incorrect, None=skip)
        verified = [e for e in entries if e.get("human_verified") in (1, 0)]
        if verified:
            match_rate = sum(1 for e in verified if e["human_verified"] == 1) / len(verified) * 100
            match_display = f"{match_rate:.1f}%"
        else:
            match_display = "—"

        stats = [
            ("📬 סה\"כ מיילים", str(total), "#0984e3"),
            ("🎯 דיוק כללי מיילים", acc_display, "#00b894"),
            ("✔️ רמת התאמה", match_display, "#6c5ce7"),
            ("📅 תקופה", f"{first_date} → {last_date}", "#fdcb6e"),
            ("💰 עלות כוללת", f"${total_cost:.2f}", "#e17055"),
            ("⏱ זמן ממוצע", f"{avg_time:.0f} שניות", "#00cec9"),
        ]
        
        for i, (label, value, color) in enumerate(stats):
            card = ttk.Frame(cards_frame, relief=tk.RIDGE, borderwidth=1)
            card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=4)
            
            tk.Label(card, text=value, font=("Arial", 22, "bold"), 
                    fg=color).pack(pady=(10, 2))
            tk.Label(card, text=label, font=("Arial", 12), 
                    fg="gray").pack(pady=(0, 10))
    
    def _show_accuracy_metrics(self, entries):
        """Show accuracy metrics: email accuracy and row accuracy."""
        import os

        # Load weights from .env (with defaults)
        weights = {
            "full":   float(os.getenv("ACCURACY_WEIGHT_FULL", "1.0")),
            "high":   float(os.getenv("ACCURACY_WEIGHT_HIGH", "1.0")),
            "medium": float(os.getenv("ACCURACY_WEIGHT_MEDIUM", "0.8")),
            "low":    float(os.getenv("ACCURACY_WEIGHT_LOW", "0.5")),
            "none":   float(os.getenv("ACCURACY_WEIGHT_NONE", "0.0")),
        }

        # Filter entries that have accuracy_data with actual rows
        entries_with_acc = [e for e in entries
                           if e.get("accuracy_data") and e["accuracy_data"].get("total", 0) > 0]

        # Always show the frame, but if no data show placeholder
        if not entries_with_acc:
            acc_frame = ttk.LabelFrame(self.content, text="\U0001F3AF מדדי דיוק חילוץ", padding=12)
            acc_frame.pack(fill=tk.X, pady=(0, 12))
            settings_btn = ttk.Button(acc_frame, text="\u2699\uFE0F", width=3,
                                       command=self._show_weights_dialog)
            settings_btn.pack(anchor="ne", pady=(0, 4))
            tk.Label(acc_frame, text="אין נתוני דיוק עדיין — הנתונים יופיעו אחרי ריצות חדשות",
                     font=("Arial", 13), fg="gray").pack(pady=12)
            return

        # Categorize entries by period
        today = datetime.now().strftime("%Y-%m-%d")
        this_month = datetime.now().strftime("%Y-%m")

        daily_entries = [e for e in entries_with_acc
                         if self._local_ts(e)[:10] == today]
        monthly_entries = [e for e in entries_with_acc
                           if self._local_ts(e)[:7] == this_month]
        all_entries = entries_with_acc

        def calc_email_accuracy(entry_list):
            """Average accuracy per email."""
            if not entry_list:
                return None
            email_scores = []
            for e in entry_list:
                acc = e.get("accuracy_data", {})
                total = acc.get("total", 0)
                if total == 0:
                    continue
                score = (
                    acc.get("full", 0) * weights["full"] +
                    acc.get("high", 0) * weights["high"] +
                    acc.get("medium", 0) * weights["medium"] +
                    acc.get("low", 0) * weights["low"] +
                    acc.get("none", 0) * weights["none"]
                ) / total * 100
                email_scores.append(score)
            return sum(email_scores) / len(email_scores) if email_scores else None

        def calc_row_accuracy(entry_list):
            """Global accuracy across all rows."""
            if not entry_list:
                return None
            total_rows = 0
            total_score = 0
            for e in entry_list:
                acc = e.get("accuracy_data", {})
                total = acc.get("total", 0)
                total_rows += total
                total_score += (
                    acc.get("full", 0) * weights["full"] +
                    acc.get("high", 0) * weights["high"] +
                    acc.get("medium", 0) * weights["medium"] +
                    acc.get("low", 0) * weights["low"] +
                    acc.get("none", 0) * weights["none"]
                )
            return (total_score / total_rows * 100) if total_rows > 0 else None

        def calc_distribution(entry_list):
            """Count total items per confidence level."""
            dist = {"full": 0, "high": 0, "medium": 0, "low": 0, "none": 0}
            for e in entry_list:
                acc = e.get("accuracy_data", {})
                for key in dist:
                    dist[key] += acc.get(key, 0)
            return dist

        # Build accuracy frame
        acc_frame = ttk.LabelFrame(self.content, text="\U0001F3AF מדדי דיוק חילוץ", padding=12)
        acc_frame.pack(fill=tk.X, pady=(0, 12))

        # Settings button
        settings_btn = ttk.Button(acc_frame, text="\u2699\uFE0F", width=3,
                                   command=self._show_weights_dialog)
        settings_btn.pack(anchor="ne", pady=(0, 4))

        # Grid-based layout for aligned columns
        grid_frame = tk.Frame(acc_frame)
        grid_frame.pack(fill=tk.X, pady=(0, 6))

        # Configure columns (RTL: col0=כללי, col1=חודשי, col2=יומי, col3=label)
        grid_frame.columnconfigure(0, weight=1, uniform="acc_col")
        grid_frame.columnconfigure(1, weight=1, uniform="acc_col")
        grid_frame.columnconfigure(2, weight=1, uniform="acc_col")
        grid_frame.columnconfigure(3, weight=1, uniform="acc_col")

        # Header row
        tk.Label(grid_frame, text="", font=("Arial", 13, "bold")).grid(
            row=0, column=3, sticky="e", padx=4)
        tk.Label(grid_frame, text="\u05D9\u05D5\u05DE\u05D9", font=("Arial", 13, "bold"),
                 anchor="center").grid(row=0, column=2, sticky="ew", padx=4)
        tk.Label(grid_frame, text="\u05D7\u05D5\u05D3\u05E9\u05D9", font=("Arial", 13, "bold"),
                 anchor="center").grid(row=0, column=1, sticky="ew", padx=4)
        tk.Label(grid_frame, text="\u05DB\u05DC\u05DC\u05D9", font=("Arial", 13, "bold"),
                 anchor="center").grid(row=0, column=0, sticky="ew", padx=4)

        # Email accuracy row
        tk.Label(grid_frame, text="\U0001F4E7 \u05D3\u05D9\u05D5\u05E7 \u05DE\u05D9\u05D9\u05DC\u05D9\u05DD:", font=("Arial", 13),
                 anchor="e").grid(row=1, column=3, sticky="e", padx=4, pady=2)

        for col_idx, period_entries in enumerate([all_entries, monthly_entries, daily_entries]):
            val = calc_email_accuracy(period_entries)
            text = f"{val:.1f}%" if val is not None else "\u2014"
            color = "#00b894" if val and val >= 85 else "#fdcb6e" if val and val >= 70 else "#e17055" if val else "gray"
            tk.Label(grid_frame, text=text, font=("Arial", 18, "bold"),
                     fg=color, anchor="center").grid(row=1, column=col_idx, sticky="ew", padx=4, pady=2)

        # Row accuracy row
        tk.Label(grid_frame, text="\U0001F4CB \u05D3\u05D9\u05D5\u05E7 \u05E9\u05D5\u05E8\u05D5\u05EA:", font=("Arial", 13),
                 anchor="e").grid(row=2, column=3, sticky="e", padx=4, pady=2)

        for col_idx, period_entries in enumerate([all_entries, monthly_entries, daily_entries]):
            val = calc_row_accuracy(period_entries)
            text = f"{val:.1f}%" if val is not None else "\u2014"
            color = "#00b894" if val and val >= 85 else "#fdcb6e" if val and val >= 70 else "#e17055" if val else "gray"
            tk.Label(grid_frame, text=text, font=("Arial", 18, "bold"),
                     fg=color, anchor="center").grid(row=2, column=col_idx, sticky="ew", padx=4, pady=2)

        # Human-verified match rate row
        def calc_match_rate(entry_list):
            """Match rate from human verification (1=correct, 0=incorrect)."""
            verified = [e for e in entry_list if e.get("human_verified") in (1, 0)]
            if not verified:
                return None
            return sum(1 for e in verified if e["human_verified"] == 1) / len(verified) * 100

        tk.Label(grid_frame, text="\u2714\uFE0F \u05E8\u05DE\u05EA \u05D4\u05EA\u05D0\u05DE\u05D4:", font=("Arial", 13),
                 anchor="e").grid(row=3, column=3, sticky="e", padx=4, pady=2)

        for col_idx, period_entries in enumerate([all_entries, monthly_entries, daily_entries]):
            val = calc_match_rate(period_entries)
            text = f"{val:.1f}%" if val is not None else "\u2014"
            color = "#00b894" if val and val >= 85 else "#fdcb6e" if val and val >= 70 else "#e17055" if val else "gray"
            tk.Label(grid_frame, text=text, font=("Arial", 18, "bold"),
                     fg=color, anchor="center").grid(row=3, column=col_idx, sticky="ew", padx=4, pady=2)

        # Distribution bar for overall
        dist = calc_distribution(all_entries)
        total = sum(dist.values())
        if total > 0:
            dist_frame = ttk.Frame(acc_frame)
            dist_frame.pack(fill=tk.X, pady=(8, 2))

            tk.Label(dist_frame, text="\u05D4\u05EA\u05E4\u05DC\u05D2\u05D5\u05EA \u05DB\u05DC\u05DC\u05D9\u05EA:", font=("Arial", 12),
                     anchor="e").pack(side=tk.RIGHT, padx=(0, 8))

            colors = {"full": "#27ae60", "high": "#2ecc71", "medium": "#f39c12",
                      "low": "#e74c3c", "none": "#95a5a6"}
            labels = {"full": "\u05DE\u05DC\u05D0", "high": "\u05D2\u05D1\u05D5\u05D4", "medium": "\u05D1\u05D9\u05E0\u05D5\u05E0\u05D9",
                      "low": "\u05E0\u05DE\u05D5\u05DA", "none": "\u05DC\u05DC\u05D0"}

            for level in ["full", "high", "medium", "low", "none"]:
                count = dist[level]
                pct = count / total * 100
                if count > 0:
                    tk.Label(dist_frame, text=f"\u25CF {labels[level]}: {count} ({pct:.0f}%)",
                             font=("Arial", 11), fg=colors[level]).pack(side=tk.RIGHT, padx=4)

        # Weights display (small, bottom)
        weights_text = f"\u05DE\u05E9\u05E7\u05D5\u05DC\u05D5\u05EA: \u05DE\u05DC\u05D0={weights['full']}, \u05D2\u05D1\u05D5\u05D4={weights['high']}, \u05D1\u05D9\u05E0\u05D5\u05E0\u05D9={weights['medium']}, \u05E0\u05DE\u05D5\u05DA={weights['low']}, \u05DC\u05DC\u05D0={weights['none']}"
        tk.Label(acc_frame, text=weights_text, font=("Arial", 10),
                 fg="#999999").pack(anchor="e", pady=(4, 0))

    def _show_weights_dialog(self) -> None:
        """Dialog to edit accuracy weights."""
        import os

        dialog = tk.Toplevel(self.win)
        dialog.title("\u2699\uFE0F \u05D4\u05D2\u05D3\u05E8\u05EA \u05DE\u05E9\u05E7\u05D5\u05DC\u05D5\u05EA \u05D3\u05D9\u05D5\u05E7")
        dialog.geometry("300x280")
        dialog.resizable(False, False)

        ttk.Label(dialog, text="\u05DE\u05E9\u05E7\u05D5\u05DC\u05D5\u05EA \u05DC\u05D7\u05D9\u05E9\u05D5\u05D1 \u05D3\u05D9\u05D5\u05E7:",
                  font=("Arial", 11, "bold")).pack(pady=(12, 8))

        levels = [
            ("\u05DE\u05DC\u05D0 (full)", "ACCURACY_WEIGHT_FULL", "1.0"),
            ("\u05D2\u05D1\u05D5\u05D4 (high)", "ACCURACY_WEIGHT_HIGH", "1.0"),
            ("\u05D1\u05D9\u05E0\u05D5\u05E0\u05D9 (medium)", "ACCURACY_WEIGHT_MEDIUM", "0.8"),
            ("\u05E0\u05DE\u05D5\u05DA (low)", "ACCURACY_WEIGHT_LOW", "0.5"),
            ("\u05DC\u05DC\u05D0 (none)", "ACCURACY_WEIGHT_NONE", "0.0"),
        ]

        weight_entries = {}
        for label, env_key, default in levels:
            row = ttk.Frame(dialog)
            row.pack(fill=tk.X, padx=20, pady=2)
            tk.Label(row, text=label, width=16, anchor="w").pack(side=tk.LEFT)

            current = os.getenv(env_key, default)
            var = tk.StringVar(value=current)
            entry = ttk.Entry(row, textvariable=var, width=8)
            entry.pack(side=tk.RIGHT)
            weight_entries[env_key] = var

        def save_weights() -> None:
            import re as _re
            env_path = Path(".env")
            if env_path.exists():
                content = env_path.read_text(encoding="utf-8")
            else:
                content = ""

            for env_key, var in weight_entries.items():
                value = var.get().strip()
                pattern = rf"^{env_key}=.*$"
                replacement = f"{env_key}={value}"
                if _re.search(pattern, content, _re.MULTILINE):
                    content = _re.sub(pattern, replacement, content, flags=_re.MULTILINE)
                else:
                    content += f"\n{replacement}"
                os.environ[env_key] = value

            env_path.write_text(content, encoding="utf-8")
            dialog.destroy()
            self._load_and_display()  # Refresh dashboard with new weights

        ttk.Button(dialog, text="\U0001F4BE \u05E9\u05DE\u05D5\u05E8", command=save_weights).pack(pady=12)
        ttk.Label(dialog, text="\u05E9\u05D9\u05E0\u05D5\u05D9\u05D9\u05DD \u05D9\u05D9\u05DB\u05E0\u05E1\u05D5 \u05DC\u05EA\u05D5\u05E7\u05E3 \u05DE\u05D9\u05D9\u05D3\u05D9\u05EA",
                  font=("Arial", 8), foreground="gray").pack()

    # ── Enhancement G: Date Range Filter ─────────────────────────────

    def _build_filter_bar(self) -> None:
        """Add date range filter at top."""
        filter_frame = ttk.Frame(self.content)
        filter_frame.pack(fill=tk.X, pady=(0, 12))

        tk.Label(filter_frame, text="📅 סינון:", font=("Arial", 13)).pack(side=tk.RIGHT, padx=4)

        if not hasattr(self, 'filter_var'):
            self.filter_var = tk.StringVar(value="all")

        periods = [
            ("היום", "today"),
            ("שבוע", "week"),
            ("חודש", "month"),
            ("הכל", "all"),
            ("טווח...", "custom"),
        ]

        for label, value in periods:
            ttk.Radiobutton(filter_frame, text=label, variable=self.filter_var,
                            value=value, command=self._on_filter_change).pack(side=tk.RIGHT, padx=4)

        # Custom date range entries (hidden by default)
        self.custom_frame = ttk.Frame(filter_frame)

        tk.Label(self.custom_frame, text="עד:").pack(side=tk.RIGHT, padx=2)
        self.date_to_entry = ttk.Entry(self.custom_frame, width=12)
        from datetime import datetime as _dt
        self.date_to_entry.insert(0, _dt.now().strftime("%Y-%m-%d"))
        self.date_to_entry.pack(side=tk.RIGHT, padx=2)

        tk.Label(self.custom_frame, text="מ:").pack(side=tk.RIGHT, padx=2)
        self.date_from_entry = ttk.Entry(self.custom_frame, width=12)
        self.date_from_entry.insert(0, _dt.now().strftime("%Y-%m-%d"))
        self.date_from_entry.pack(side=tk.RIGHT, padx=2)

        ttk.Button(self.custom_frame, text="סנן", width=5,
                   command=self._load_and_display).pack(side=tk.RIGHT, padx=4)

    def _on_filter_change(self) -> None:
        """Show/hide custom date range inputs."""
        if hasattr(self, 'custom_frame'):
            if self.filter_var.get() == "custom":
                self.custom_frame.pack(side=tk.RIGHT, padx=8)
            else:
                self.custom_frame.pack_forget()
        self._load_and_display()

    def _filter_entries(self, entries: list) -> list:
        """Filter entries based on selected period."""
        from datetime import timedelta

        if not hasattr(self, 'filter_var'):
            return entries

        period = self.filter_var.get()
        if period == "all":
            return entries

        now = datetime.now()
        if period == "today":
            cutoff = now.strftime("%Y-%m-%d")
            return [e for e in entries if self._local_ts(e)[:10] == cutoff]
        elif period == "week":
            cutoff = (now - timedelta(days=7)).strftime("%Y-%m-%dT%H:%M:%S")
            return [e for e in entries if self._local_ts(e) >= cutoff]
        elif period == "month":
            cutoff = now.strftime("%Y-%m")
            return [e for e in entries if self._local_ts(e)[:7] == cutoff]
        elif period == "custom":
            date_from = getattr(self, 'date_from_entry', None)
            date_to = getattr(self, 'date_to_entry', None)
            if date_from and date_to:
                from_val = date_from.get().strip()
                to_val = date_to.get().strip()
                return [e for e in entries
                        if from_val <= self._local_ts(e)[:10] <= to_val]
        return entries

    # ── Enhancement B: Accuracy Trend Graph ──────────────────────────

    def _show_accuracy_trend(self, entries) -> None:
        """Show daily accuracy trend."""
        import os
        from collections import defaultdict

        weights = {
            "full":   float(os.getenv("ACCURACY_WEIGHT_FULL", "1.0")),
            "high":   float(os.getenv("ACCURACY_WEIGHT_HIGH", "1.0")),
            "medium": float(os.getenv("ACCURACY_WEIGHT_MEDIUM", "0.8")),
            "low":    float(os.getenv("ACCURACY_WEIGHT_LOW", "0.5")),
            "none":   float(os.getenv("ACCURACY_WEIGHT_NONE", "0.0")),
        }

        entries_with_acc = [e for e in entries
                           if e.get("accuracy_data") and e["accuracy_data"].get("total", 0) > 0]
        if not entries_with_acc:
            return

        daily_scores = defaultdict(list)
        for e in entries_with_acc:
            day = self._local_ts(e)[:10]
            acc = e.get("accuracy_data", {})
            total = acc.get("total", 0)
            if total == 0:
                continue
            score = sum(acc.get(k, 0) * weights.get(k, 0) for k in weights) / total * 100
            daily_scores[day].append(score)

        if not daily_scores:
            return

        trend_frame = ttk.LabelFrame(self.content, text="\U0001F4C8 \u05DE\u05D2\u05DE\u05EA \u05D3\u05D9\u05D5\u05E7 \u05D9\u05D5\u05DE\u05D9\u05EA", padding=12)
        trend_frame.pack(fill=tk.X, pady=(0, 12))

        days_sorted = sorted(daily_scores.keys())

        for day in sorted(daily_scores.keys(), reverse=True)[:14]:
            scores = daily_scores[day]
            avg = sum(scores) / len(scores)

            bar_len = int(avg / 2)
            bar = "\u2588" * bar_len

            color = "#00b894" if avg >= 85 else "#fdcb6e" if avg >= 70 else "#e17055"

            # Trend arrow
            trend_icon = ""
            day_idx = days_sorted.index(day) if day in days_sorted else -1
            if day_idx > 0:
                prev_day = days_sorted[day_idx - 1]
                prev_avg = sum(daily_scores[prev_day]) / len(daily_scores[prev_day])
                if avg > prev_avg + 2:
                    trend_icon = " \u2191"
                elif avg < prev_avg - 2:
                    trend_icon = " \u2193"

            row = ttk.Frame(trend_frame)
            row.pack(fill=tk.X, pady=1)
            tk.Label(row, text=day, font=("Courier", 12), width=12,
                    anchor="e").pack(side=tk.RIGHT)
            tk.Label(row, text=bar, font=("Courier", 12), fg=color,
                    anchor="e").pack(side=tk.RIGHT)
            tk.Label(row, text=f" {avg:.0f}%{trend_icon}", font=("Courier", 12, "bold"),
                    fg=color, anchor="e").pack(side=tk.RIGHT)
            tk.Label(row, text=f"  ({len(scores)} \u05DE\u05D9\u05D9\u05DC\u05D9\u05DD)", font=("Courier", 11),
                    fg="gray").pack(side=tk.RIGHT)

    # ── Enhancement C: Accuracy per Customer (Top 10) ────────────────

    def _show_customer_accuracy(self, entries) -> None:
        """Show accuracy breakdown by customer with % of totals."""
        import os
        from collections import defaultdict

        weights = {
            "full":   float(os.getenv("ACCURACY_WEIGHT_FULL", "1.0")),
            "high":   float(os.getenv("ACCURACY_WEIGHT_HIGH", "1.0")),
            "medium": float(os.getenv("ACCURACY_WEIGHT_MEDIUM", "0.8")),
            "low":    float(os.getenv("ACCURACY_WEIGHT_LOW", "0.5")),
            "none":   float(os.getenv("ACCURACY_WEIGHT_NONE", "0.0")),
        }

        entries_with_acc = [e for e in entries
                           if e.get("accuracy_data") and e.get("customers")
                           and e["accuracy_data"].get("total", 0) > 0]
        if not entries_with_acc:
            return

        customer_data = defaultdict(lambda: {"total_score": 0, "total_items": 0, "emails": 0})

        for e in entries_with_acc:
            acc = e.get("accuracy_data", {})
            customers = e.get("customers", [])
            total = acc.get("total", 0)
            if total == 0 or not customers:
                continue

            score = sum(acc.get(k, 0) * weights.get(k, 0) for k in weights)
            cust = customers[0].upper()
            customer_data[cust]["total_score"] += score
            customer_data[cust]["total_items"] += total
            customer_data[cust]["emails"] += 1

        if not customer_data:
            return

        # Grand totals for percentage calculation
        grand_total_items = sum(d["total_items"] for d in customer_data.values())
        grand_total_emails = sum(d["emails"] for d in customer_data.values())

        cust_frame = ttk.LabelFrame(self.content, text="\U0001F3E2 \u05D3\u05D9\u05D5\u05E7 \u05DC\u05E4\u05D9 \u05DC\u05E7\u05D5\u05D7 (Top 10)", padding=12)
        cust_frame.pack(fill=tk.X, pady=(0, 12))

        sorted_customers = sorted(customer_data.items(),
                                   key=lambda x: x[1]["total_items"], reverse=True)[:10]

        for cust, data in sorted_customers:
            acc_pct = (data["total_score"] / data["total_items"] * 100) if data["total_items"] > 0 else 0
            items_pct = (data["total_items"] / grand_total_items * 100) if grand_total_items > 0 else 0
            emails_pct = (data["emails"] / grand_total_emails * 100) if grand_total_emails > 0 else 0
            color = "#00b894" if acc_pct >= 85 else "#fdcb6e" if acc_pct >= 70 else "#e17055"

            row = ttk.Frame(cust_frame)
            row.pack(fill=tk.X, pady=1)
            tk.Label(row, text=f"{acc_pct:5.1f}%", font=("Courier", 13, "bold"),
                    fg=color, width=7).pack(side=tk.RIGHT)
            tk.Label(row, text=cust[:25], font=("Courier", 12),
                    width=25, anchor="e").pack(side=tk.RIGHT)
            tk.Label(row, text=f"{data['total_items']} ({items_pct:.1f}%)",
                    font=("Courier", 11), fg="gray", width=14,
                    anchor="e").pack(side=tk.RIGHT)
            tk.Label(row, text=f"{data['emails']} ({emails_pct:.1f}%)",
                    font=("Courier", 11), fg="gray", width=14,
                    anchor="e").pack(side=tk.RIGHT)

    # ── Enhancement D: Cost per Item + Time per Item ─────────────────

    def _show_efficiency_metrics(self, entries):
        """Show cost per item and time per item."""
        def _get_items(e):
            ic = e.get("items_count", 0)
            if ic and ic > 0:
                return ic
            return e.get("accuracy_data", {}).get("total", 0)

        entries_with_items = [e for e in entries if _get_items(e) > 0]
        if not entries_with_items:
            return

        total_items = sum(_get_items(e) for e in entries_with_items)
        total_cost = sum(e.get("cost_usd", 0) for e in entries_with_items)
        total_time = sum(e.get("processing_time_seconds", 0) for e in entries_with_items)

        cost_per_item = total_cost / total_items if total_items > 0 else 0
        time_per_item = total_time / total_items if total_items > 0 else 0

        eff_frame = ttk.LabelFrame(self.content, text="\u26A1 \u05D9\u05E2\u05D9\u05DC\u05D5\u05EA", padding=12)
        eff_frame.pack(fill=tk.X, pady=(0, 12))

        metrics = [
            ("\U0001F4B0 \u05E2\u05DC\u05D5\u05EA \u05DC\u05E4\u05E8\u05D9\u05D8", f"${cost_per_item:.4f}", "#e17055"),
            ("\u23F1 \u05D6\u05DE\u05DF \u05DC\u05E4\u05E8\u05D9\u05D8", f"{time_per_item:.1f} \u05E9\u05E0\u05D9\u05D5\u05EA", "#00cec9"),
            ("\U0001F4CB \u05E1\u05D4\"\u05DB \u05E4\u05E8\u05D9\u05D8\u05D9\u05DD", str(total_items), "#0984e3"),
        ]

        for label, value, color in metrics:
            row = ttk.Frame(eff_frame)
            row.pack(fill=tk.X, pady=2)
            tk.Label(row, text=label, font=("Arial", 13), width=16,
                    anchor="e").pack(side=tk.RIGHT)
            tk.Label(row, text=value, font=("Arial", 16, "bold"),
                    fg=color).pack(side=tk.RIGHT, padx=8)

    # ── Enhancement E: PL Override Stats ─────────────────────────────

    def _show_pl_stats(self, entries) -> None:
        """Show PL override statistics."""
        total_pl_overrides = sum(e.get("pl_overrides", 0) for e in entries)
        total_items = sum(
            e.get("items_count", 0) or e.get("accuracy_data", {}).get("total", 0)
            for e in entries
        )
        emails_with_pl = sum(1 for e in entries if e.get("pl_overrides", 0) > 0)

        if total_pl_overrides == 0:
            return

        pl_frame = ttk.LabelFrame(self.content, text="\U0001F4CB PL Override \u2014 \u05EA\u05D9\u05E7\u05D5\u05E0\u05D9 OCR", padding=12)
        pl_frame.pack(fill=tk.X, pady=(0, 12))

        pct = (total_pl_overrides / total_items * 100) if total_items > 0 else 0

        metrics = [
            ("\U0001F527 \u05E4\u05E8\u05D9\u05D8\u05D9\u05DD \u05E9\u05EA\u05D5\u05E7\u05E0\u05D5", str(total_pl_overrides), "#3b82f6"),
            ("\U0001F4E7 \u05DE\u05D9\u05D9\u05DC\u05D9\u05DD \u05E2\u05DD PL", str(emails_with_pl), "#6c5ce7"),
            ("\U0001F4CA \u05D0\u05D7\u05D5\u05D6 \u05EA\u05D9\u05E7\u05D5\u05DF", f"{pct:.1f}%", "#00b894"),
        ]

        for label, value, color in metrics:
            row = ttk.Frame(pl_frame)
            row.pack(fill=tk.X, pady=2)
            tk.Label(row, text=label, font=("Arial", 13), width=18,
                    anchor="e").pack(side=tk.RIGHT)
            tk.Label(row, text=value, font=("Arial", 16, "bold"),
                    fg=color).pack(side=tk.RIGHT, padx=8)

    # ── Enhancement F: Top Errors ────────────────────────────────────

    def _show_top_errors(self, entries) -> None:
        """Show most common error patterns."""
        all_errors = []
        for e in entries:
            all_errors.extend(e.get("error_types", []))

        if not all_errors:
            return

        error_counts = Counter(all_errors)

        err_frame = ttk.LabelFrame(self.content, text="\u26A0\uFE0F \u05E9\u05D2\u05D9\u05D0\u05D5\u05EA \u05E0\u05E4\u05D5\u05E6\u05D5\u05EA", padding=12)
        err_frame.pack(fill=tk.X, pady=(0, 12))

        ERROR_LABELS = {
            "missing_part_number": "\u05D7\u05E1\u05E8 \u05DE\u05E1\u05E4\u05E8 \u05E4\u05E8\u05D9\u05D8",
            "low_confidence": "\u05D1\u05D9\u05D8\u05D7\u05D5\u05DF \u05E0\u05DE\u05D5\u05DA (\u05DC\u05D0 \u05D1\u05E9\u05DD \u05E7\u05D5\u05D1\u05E5)",
        }

        for error_type, count in error_counts.most_common(5):
            label = ERROR_LABELS.get(error_type, error_type)
            row = ttk.Frame(err_frame)
            row.pack(fill=tk.X, pady=1)
            tk.Label(row, text=f"{count:4d}", font=("Courier", 13, "bold"),
                    fg="#e17055", width=5).pack(side=tk.RIGHT)
            tk.Label(row, text=label, font=("Arial", 12)).pack(side=tk.RIGHT, padx=4)

    def _show_items_per_mail(self, entries: list) -> None:
        """Show items-per-mail statistics and distribution."""
        def _get_items(e):
            ic = e.get("items_count", 0)
            if ic and ic > 0:
                return ic
            ad = e.get("accuracy_data", {})
            return sum(ad.get(k, 0) for k in ['full', 'high', 'medium', 'low', 'none'])

        items_list = [_get_items(e) for e in entries if _get_items(e) > 0]
        if not items_list:
            return

        total_items = sum(items_list)
        avg_items = total_items / len(items_list)
        sorted_items = sorted(items_list)
        median_items = sorted_items[len(sorted_items) // 2]
        max_items = max(items_list)
        min_items = min(items_list)

        frame = ttk.LabelFrame(self.content, text="📦 פריטים למייל", padding=12)
        frame.pack(fill=tk.X, pady=(0, 12))

        # Summary metrics row
        metrics = [
            ("📊 ממוצע", f"{avg_items:.1f}", "#0984e3"),
            ("📐 חציון", str(median_items), "#00b894"),
            ("⬆️ מקסימום", str(max_items), "#e17055"),
            ("⬇️ מינימום", str(min_items), "#636e72"),
            ("📧 מיילים", str(len(items_list)), "#6c5ce7"),
            ("📋 סה\"כ פריטים", str(total_items), "#00cec9"),
        ]

        metrics_row = ttk.Frame(frame)
        metrics_row.pack(fill=tk.X, pady=(0, 8))
        for label, value, color in metrics:
            cell = ttk.Frame(metrics_row)
            cell.pack(side=tk.RIGHT, padx=12)
            tk.Label(cell, text=value, font=("Arial", 18, "bold"), fg=color).pack()
            tk.Label(cell, text=label, font=("Arial", 12), fg="gray").pack()

        # Distribution bars
        dist_frame = ttk.Frame(frame)
        dist_frame.pack(fill=tk.X, pady=(4, 0))

        tk.Label(dist_frame, text="התפלגות:", font=("Arial", 13, "bold")).pack(anchor="e")

        buckets = [
            ("1 פריט", lambda n: n == 1),
            ("2-3 פריטים", lambda n: 2 <= n <= 3),
            ("4-5 פריטים", lambda n: 4 <= n <= 5),
            ("6-10 פריטים", lambda n: 6 <= n <= 10),
            ("11+ פריטים", lambda n: n >= 11),
        ]

        colors = ["#00b894", "#00cec9", "#0984e3", "#6c5ce7", "#e17055"]

        for i, (label, condition) in enumerate(buckets):
            count = sum(1 for n in items_list if condition(n))
            pct = count / len(items_list) * 100 if items_list else 0

            row = ttk.Frame(dist_frame)
            row.pack(fill=tk.X, pady=1)

            tk.Label(row, text=f"{label}", font=("Arial", 12), width=12,
                    anchor="e").pack(side=tk.RIGHT)
            
            # Visual bar
            bar_canvas = tk.Canvas(row, height=20, highlightthickness=0)
            bar_canvas.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(4, 4))
            
            bar_canvas.update_idletasks()
            canvas_width = max(bar_canvas.winfo_width(), 300)
            bar_width = int(canvas_width * pct / 100)
            if bar_width < 2 and count > 0:
                bar_width = 2
            bar_canvas.create_rectangle(canvas_width - bar_width, 1, canvas_width, 15,
                                       fill=colors[i], outline="")
            
            tk.Label(row, text=f"{count} ({pct:.0f}%)", font=("Arial", 12),
                    width=10).pack(side=tk.RIGHT)

    def _show_daily_breakdown(self, entries) -> None:
        """Show emails per day."""
        daily_frame = ttk.LabelFrame(self.content, text="📅 מיילים לפי יום", padding=12)
        daily_frame.pack(fill=tk.X, pady=(0, 12))
        
        day_counts = Counter()
        day_costs = {}
        for e in entries:
            ts = self._local_ts(e)
            day = ts[:10]
            if day and day != "":
                day_counts[day] += 1
                day_costs.setdefault(day, 0.0)
                day_costs[day] += e.get("cost_usd", 0) or 0
        
        if not day_counts:
            ttk.Label(daily_frame, text="אין נתוני תאריך").pack()
            return
        
        # Show last 14 days
        for day in sorted(day_counts.keys(), reverse=True)[:14]:
            count = day_counts[day]
            cost = day_costs.get(day, 0)
            bar = "█" * min(count, 40) + f" ({count})"
            cost_str = f"  ${cost:.2f}" if cost > 0 else ""
            row = ttk.Frame(daily_frame)
            row.pack(fill=tk.X, pady=1)
            tk.Label(row, text=day, font=("Courier", 12), width=12, 
                    anchor="e").pack(side=tk.RIGHT)
            tk.Label(row, text=bar, font=("Courier", 12), fg="#00b894", 
                    anchor="e").pack(side=tk.RIGHT)
            if cost_str:
                tk.Label(row, text=cost_str, font=("Courier", 12), fg="#e17055", 
                        anchor="e").pack(side=tk.RIGHT)
    
    def _show_top_senders(self, entries) -> None:
        """Show top 10 senders."""
        senders_frame = ttk.LabelFrame(self.content, text="👤 שולחים מובילים (Top 10)", padding=12)
        senders_frame.pack(fill=tk.X, pady=(0, 12))
        
        sender_counts = Counter()
        for e in entries:
            sender = e.get("sender") or ""
            if sender:
                sender_counts[sender.lower()] += 1
        
        if not sender_counts:
            ttk.Label(senders_frame, text="אין נתוני שולחים").pack()
            return
        
        for sender, count in sender_counts.most_common(10):
            row = ttk.Frame(senders_frame)
            row.pack(fill=tk.X, pady=1)
            tk.Label(row, text=f"{count:4d}", font=("Courier", 12, "bold"), 
                    fg="#6c5ce7", width=5).pack(side=tk.RIGHT)
            tk.Label(row, text=sender, font=("Courier", 12), 
                    anchor="e").pack(side=tk.RIGHT, fill=tk.X)
    
    def _show_recent_emails(self, entries) -> None:
        """Show last 100 processed emails with scrollbar."""
        recent_frame = ttk.LabelFrame(self.content, text="📬 100 מיילים אחרונים", padding=12)
        recent_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 12))
        
        # Headers
        header_frame = tk.Frame(recent_frame, bg="#2c3e50")
        header_frame.pack(fill=tk.X, pady=(0, 2))
        
        tk.Label(header_frame, text="תאריך", font=("Arial", 12, "bold"), 
                bg="#2c3e50", fg="white", width=18, anchor="e").pack(side=tk.RIGHT, padx=2, pady=3)
        tk.Label(header_frame, text="שולח", font=("Arial", 12, "bold"), 
                bg="#2c3e50", fg="white", width=32, anchor="e").pack(side=tk.RIGHT, padx=2, pady=3)
        tk.Label(header_frame, text="קבצים", font=("Arial", 12, "bold"), 
                bg="#2c3e50", fg="white", width=6, anchor="center").pack(side=tk.RIGHT, padx=2, pady=3)
        tk.Label(header_frame, text="עלות", font=("Arial", 12, "bold"), 
                bg="#2c3e50", fg="white", width=8, anchor="center").pack(side=tk.RIGHT, padx=2, pady=3)
        tk.Label(header_frame, text="זמן", font=("Arial", 12, "bold"), 
                bg="#2c3e50", fg="white", width=8, anchor="center").pack(side=tk.RIGHT, padx=2, pady=3)
        tk.Label(header_frame, text="סטטוס", font=("Arial", 12, "bold"), 
                bg="#2c3e50", fg="white", width=6, anchor="center").pack(side=tk.RIGHT, padx=2, pady=3)
        tk.Label(header_frame, text="אימות", font=("Arial", 12, "bold"), 
                bg="#2c3e50", fg="white", width=6, anchor="center").pack(side=tk.RIGHT, padx=2, pady=3)
        
        # Scrollable area for rows
        scroll_container = tk.Frame(recent_frame)
        scroll_container.pack(fill=tk.BOTH, expand=True)
        
        canvas = tk.Canvas(scroll_container, highlightthickness=0, bg="#ffffff")
        scrollbar = ttk.Scrollbar(scroll_container, orient=tk.VERTICAL, command=canvas.yview)
        rows_frame = tk.Frame(canvas, bg="#ffffff")
        
        rows_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=rows_frame, anchor="nw", tags="rows_window")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Keep rows_frame width synced with canvas
        def _on_canvas_configure(event):
            canvas.itemconfig("rows_window", width=event.width)
        canvas.bind("<Configure>", _on_canvas_configure)
        
        scrollbar.pack(side=tk.LEFT, fill=tk.Y)
        canvas.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        
        # Mouse wheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind("<MouseWheel>", _on_mousewheel)
        rows_frame.bind("<MouseWheel>", _on_mousewheel)
        
        # Last 100 entries (reversed = newest first)
        for idx, entry in enumerate(reversed(entries[-100:])):
            bg = "#f0f0f0" if idx % 2 == 0 else "#ffffff"
            
            ts = self._utc_to_local(entry.get("received") or entry.get("timestamp") or "")
            sender = entry.get("sender") or ""
            attachments = entry.get("attachments_display_names") or []
            attachments = len(attachments) if isinstance(attachments, list) else 0
            status_ok = entry.get("sent", False)
            status_icon = "✅" if status_ok else "❌"
            
            row = tk.Frame(rows_frame, bg=bg)
            row.pack(fill=tk.X, pady=0)
            row.bind("<MouseWheel>", _on_mousewheel)
            
            for widget_args in [
                {"text": str(ts)[:16], "font": ("Courier", 11), "width": 18, "anchor": "e"},
                {"text": sender[:32], "font": ("Courier", 11), "width": 32, "anchor": "e"},
                {"text": str(attachments), "font": ("Arial", 12), "width": 6, "anchor": "center", "fg": "#0984e3"},
            ]:
                lbl = tk.Label(row, bg=bg, **widget_args)
                lbl.pack(side=tk.RIGHT, padx=2, pady=2)
                lbl.bind("<MouseWheel>", _on_mousewheel)

            cost = entry.get("cost_usd", 0)
            cost_str = f"${cost:.3f}" if cost else "-"
            proc_time = entry.get("processing_time_seconds", 0)
            time_str = f"{proc_time:.0f}s" if proc_time else "-"

            for widget_args in [
                {"text": cost_str, "font": ("Courier", 11), "width": 8, "anchor": "center", "fg": "#e17055"},
                {"text": time_str, "font": ("Courier", 11), "width": 8, "anchor": "center", "fg": "#00cec9"},
                {"text": status_icon, "font": ("Arial", 13), "width": 6, "anchor": "center"},
            ]:
                lbl = tk.Label(row, bg=bg, **widget_args)
                lbl.pack(side=tk.RIGHT, padx=2, pady=2)
                lbl.bind("<MouseWheel>", _on_mousewheel)

            # Human verification dropdown (1 = correct, 0 = incorrect, empty = not reviewed)
            entry_id = entry.get("id", "")
            current_val = entry.get("human_verified")
            # Default to 1 if not yet reviewed
            if current_val is None:
                current_val = 1
                self._save_entry_field(entry_id, "human_verified", 1)
            var = tk.StringVar(value=str(current_val) if current_val in (1, 0) else "1")

            def _on_verify_change(eid=entry_id, v=var):
                sel = v.get()
                if sel == "1":
                    self._save_entry_field(eid, "human_verified", 1)
                elif sel == "0":
                    self._save_entry_field(eid, "human_verified", 0)
                else:
                    self._save_entry_field(eid, "human_verified", None)

            combo_frame = tk.Frame(row, bg=bg, width=60)
            combo_frame.pack(side=tk.RIGHT, padx=2, pady=2)
            combo_frame.pack_propagate(False)
            combo_frame.configure(height=24)
            combo = ttk.Combobox(combo_frame, textvariable=var, values=["", "1", "0"],
                                 width=3, state="readonly", justify="center")
            combo.pack(expand=True)
            combo.bind("<<ComboboxSelected>>", lambda e, cb=_on_verify_change: cb())
            combo.bind("<MouseWheel>", _on_mousewheel)

    def _export_to_excel(self) -> None:
        """Export dashboard data to Excel file with full statistics."""
        from tkinter import filedialog, messagebox

        entries = self._load_log()
        if not entries:
            messagebox.showinfo("ייצוא", "אין נתונים לייצוא")
            return

        default_name = f"dashboard_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=default_name,
            title="שמור דוח Dashboard",
        )
        if not filepath:
            return

        try:
            import os
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from collections import defaultdict

            weights = {
                "full":   float(os.getenv("ACCURACY_WEIGHT_FULL", "1.0")),
                "high":   float(os.getenv("ACCURACY_WEIGHT_HIGH", "1.0")),
                "medium": float(os.getenv("ACCURACY_WEIGHT_MEDIUM", "0.8")),
                "low":    float(os.getenv("ACCURACY_WEIGHT_LOW", "0.5")),
                "none":   float(os.getenv("ACCURACY_WEIGHT_NONE", "0.0")),
            }

            wb = openpyxl.Workbook()

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_align = Alignment(horizontal="center", vertical="center")
            green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

            def write_header(ws, columns, row=1):
                for col_idx, col_name in enumerate(columns, 1):
                    cell = ws.cell(row=row, column=col_idx, value=col_name)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_align
                    cell.border = thin_border

            def auto_width(ws, min_w=10, max_w=35):
                for col in ws.columns:
                    max_len = max((len(str(cell.value or "")) for cell in col), default=0)
                    ws.column_dimensions[col[0].column_letter].width = max(min(max_len + 3, max_w), min_w)

            def _get_items(e):
                ic = e.get("items_count", 0)
                if ic and ic > 0:
                    return ic
                return e.get("accuracy_data", {}).get("total", 0)

            def _calc_accuracy(acc_data):
                total = acc_data.get("total", 0)
                if total == 0:
                    return None
                score = sum(acc_data.get(k, 0) * weights.get(k, 0) for k in weights)
                return score / total * 100

            def _color_accuracy(ws, cell, val):
                if val is None:
                    return
                if val >= 85:
                    cell.fill = green_fill
                elif val >= 70:
                    cell.fill = yellow_fill
                else:
                    cell.fill = red_fill

            # ── Sheet 1: סיכום כללי ──
            ws = wb.active
            ws.title = "סיכום כללי"
            ws.sheet_view.rightToLeft = True

            total = len(entries)
            sent = sum(1 for e in entries if e.get("sent"))
            costs = [e.get("cost_usd", 0) for e in entries if e.get("cost_usd")]
            total_cost = sum(costs)
            avg_cost = total_cost / len(costs) if costs else 0
            times = [e.get("processing_time_seconds", 0) for e in entries if e.get("processing_time_seconds")]
            avg_time = sum(times) / len(times) if times else 0
            total_time = sum(times)

            acc = Counter()
            for e in entries:
                ad = e.get("accuracy_data", {})
                for level in ["full", "high", "medium", "low", "none"]:
                    acc[level] += ad.get(level, 0)
            total_items = sum(acc.values())
            success = acc["full"] + acc["high"]
            success_rate = success / total_items * 100 if total_items else 0

            unique_senders = len(set(e.get("sender", "").lower() for e in entries if e.get("sender")))
            unique_customers = len(set(
                c.upper() for e in entries for c in (e.get("customers") or [])
            ))

            summary_data = [
                ["מדד", "ערך"],
                ["סה\"כ מיילים", total],
                ["נשלחו בהצלחה", f"{sent} ({sent/total*100:.1f}%)" if total else "0"],
                ["שולחים ייחודיים", unique_senders],
                ["לקוחות ייחודיים", unique_customers],
                ["סה\"כ פריטים (שורות)", total_items],
                ["סה\"כ קבצים", sum(e.get("files_processed", 0) for e in entries)],
                ["דיוק כולל (full+high)", f"{success_rate:.1f}%"],
                ["Full", acc["full"]],
                ["High", acc["high"]],
                ["Medium", acc["medium"]],
                ["Low", acc["low"]],
                ["None", acc["none"]],
                ["עלות כוללת", f"${total_cost:.2f}"],
                ["עלות ממוצעת/מייל", f"${avg_cost:.4f}"],
                ["עלות לפריט", f"${total_cost/total_items:.4f}" if total_items else "$0"],
                ["זמן כולל (דקות)", f"{total_time/60:.1f}"],
                ["זמן ממוצע/מייל (שניות)", f"{avg_time:.0f}"],
                ["זמן לפריט (שניות)", f"{total_time/total_items:.1f}" if total_items else "0"],
                ["PL Overrides", sum(e.get("pl_overrides", 0) for e in entries)],
            ]

            for row_idx, row in enumerate(summary_data, 1):
                for col_idx, val in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border
                    if row_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_align
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 22

            # ── Sheet 2: נתוני מיילים ──
            ws2 = wb.create_sheet("נתוני מיילים")
            ws2.sheet_view.rightToLeft = True
            columns = [
                "תאריך", "שולח", "לקוחות", "קבצים", "שורות",
                "עלות $", "זמן (שניות)", "נשלח", "דיוק %",
                "Full", "High", "Medium", "Low", "None",
                "PL Overrides", "שגיאות",
            ]
            write_header(ws2, columns)

            for row_idx, entry in enumerate(entries, 2):
                ad = entry.get("accuracy_data", {})
                acc_val = _calc_accuracy(ad)
                items = _get_items(entry)

                ws2.cell(row=row_idx, column=1, value=self._local_ts(entry)[:19]).border = thin_border
                ws2.cell(row=row_idx, column=2, value=entry.get("sender", "")).border = thin_border
                ws2.cell(row=row_idx, column=3, value=", ".join(entry.get("customers", []))).border = thin_border
                ws2.cell(row=row_idx, column=4, value=entry.get("files_processed", 0)).border = thin_border
                ws2.cell(row=row_idx, column=5, value=items).border = thin_border
                ws2.cell(row=row_idx, column=6, value=entry.get("cost_usd", 0)).border = thin_border
                ws2.cell(row=row_idx, column=7, value=entry.get("processing_time_seconds", 0)).border = thin_border
                ws2.cell(row=row_idx, column=8, value="כן" if entry.get("sent") else "לא").border = thin_border

                acc_cell = ws2.cell(row=row_idx, column=9, value=f"{acc_val:.1f}%" if acc_val is not None else "—")
                acc_cell.border = thin_border
                _color_accuracy(ws2, acc_cell, acc_val)

                ws2.cell(row=row_idx, column=10, value=ad.get("full", 0)).border = thin_border
                ws2.cell(row=row_idx, column=11, value=ad.get("high", 0)).border = thin_border
                ws2.cell(row=row_idx, column=12, value=ad.get("medium", 0)).border = thin_border
                ws2.cell(row=row_idx, column=13, value=ad.get("low", 0)).border = thin_border
                ws2.cell(row=row_idx, column=14, value=ad.get("none", 0)).border = thin_border
                ws2.cell(row=row_idx, column=15, value=entry.get("pl_overrides", 0)).border = thin_border
                ws2.cell(row=row_idx, column=16, value=", ".join(entry.get("error_types", []))).border = thin_border

            auto_width(ws2)

            # ── Sheet 3: סיכום יומי ──
            ws3 = wb.create_sheet("סיכום יומי")
            ws3.sheet_view.rightToLeft = True

            daily = defaultdict(lambda: {
                "count": 0, "cost": 0, "sent": 0, "items": 0,
                "score": 0, "time": 0, "files": 0
            })
            for e in entries:
                ts = self._local_ts(e)[:10]
                if ts:
                    daily[ts]["count"] += 1
                    daily[ts]["cost"] += e.get("cost_usd", 0)
                    daily[ts]["sent"] += 1 if e.get("sent") else 0
                    daily[ts]["time"] += e.get("processing_time_seconds", 0)
                    daily[ts]["files"] += e.get("files_processed", 0)
                    ad = e.get("accuracy_data", {})
                    itm = sum(ad.get(k, 0) for k in ["full", "high", "medium", "low", "none"])
                    daily[ts]["items"] += itm
                    daily[ts]["score"] += sum(ad.get(k, 0) * weights[k] for k in weights)

            daily_cols = ["תאריך", "מיילים", "נשלחו", "קבצים", "שורות",
                          "עלות $", "זמן (דקות)", "דיוק %"]
            write_header(ws3, daily_cols)

            for row_idx, (date, data) in enumerate(sorted(daily.items()), 2):
                rate = data["score"] / data["items"] * 100 if data["items"] else 0
                ws3.cell(row=row_idx, column=1, value=date).border = thin_border
                ws3.cell(row=row_idx, column=2, value=data["count"]).border = thin_border
                ws3.cell(row=row_idx, column=3, value=data["sent"]).border = thin_border
                ws3.cell(row=row_idx, column=4, value=data["files"]).border = thin_border
                ws3.cell(row=row_idx, column=5, value=data["items"]).border = thin_border
                ws3.cell(row=row_idx, column=6, value=round(data["cost"], 2)).border = thin_border
                ws3.cell(row=row_idx, column=7, value=round(data["time"] / 60, 1)).border = thin_border
                acc_cell = ws3.cell(row=row_idx, column=8, value=f"{rate:.1f}%")
                acc_cell.border = thin_border
                _color_accuracy(ws3, acc_cell, rate)

            auto_width(ws3)

            # ── Sheet 4: סטטיסטיקת לקוחות ──
            ws4 = wb.create_sheet("סטטיסטיקת לקוחות")
            ws4.sheet_view.rightToLeft = True

            cust_data = defaultdict(lambda: {
                "emails": 0, "files": 0, "items": 0,
                "score": 0, "cost": 0, "time": 0,
                "sent": 0, "full": 0, "high": 0,
                "medium": 0, "low": 0, "none": 0,
                "senders": set(),
            })
            for e in entries:
                customers = e.get("customers") or []
                if not customers:
                    customers = ["לא ידוע"]
                ad = e.get("accuracy_data", {})
                itm = _get_items(e)
                score = sum(ad.get(k, 0) * weights.get(k, 0) for k in weights)
                sender = (e.get("sender") or "").lower()

                for cust_name in customers:
                    cust = cust_name.upper()
                    cust_data[cust]["emails"] += 1
                    cust_data[cust]["files"] += e.get("files_processed", 0)
                    cust_data[cust]["items"] += itm
                    cust_data[cust]["score"] += score
                    cust_data[cust]["cost"] += e.get("cost_usd", 0)
                    cust_data[cust]["time"] += e.get("processing_time_seconds", 0)
                    cust_data[cust]["sent"] += 1 if e.get("sent") else 0
                    cust_data[cust]["full"] += ad.get("full", 0)
                    cust_data[cust]["high"] += ad.get("high", 0)
                    cust_data[cust]["medium"] += ad.get("medium", 0)
                    cust_data[cust]["low"] += ad.get("low", 0)
                    cust_data[cust]["none"] += ad.get("none", 0)
                    if sender:
                        cust_data[cust]["senders"].add(sender)

            cust_cols = [
                "לקוח", "מיילים", "% מיילים", "קבצים", "שורות", "% שורות",
                "דיוק %", "Full", "High", "Medium", "Low", "None",
                "עלות $", "עלות/פריט $", "זמן (דקות)", "זמן/פריט (שניות)",
                "שולחים ייחודיים", "שיעור הצלחה",
            ]
            write_header(ws4, cust_cols)

            grand_emails = sum(d["emails"] for d in cust_data.values())
            grand_items = sum(d["items"] for d in cust_data.values())

            sorted_custs = sorted(cust_data.items(), key=lambda x: x[1]["items"], reverse=True)
            for row_idx, (cust, d) in enumerate(sorted_custs, 2):
                acc_pct = (d["score"] / d["items"] * 100) if d["items"] > 0 else 0
                emails_pct = (d["emails"] / grand_emails * 100) if grand_emails else 0
                items_pct = (d["items"] / grand_items * 100) if grand_items else 0
                cost_per_item = d["cost"] / d["items"] if d["items"] else 0
                time_per_item = d["time"] / d["items"] if d["items"] else 0
                success_rate = (d["sent"] / d["emails"] * 100) if d["emails"] else 0

                ws4.cell(row=row_idx, column=1, value=cust).border = thin_border
                ws4.cell(row=row_idx, column=2, value=d["emails"]).border = thin_border
                ws4.cell(row=row_idx, column=3, value=f"{emails_pct:.1f}%").border = thin_border
                ws4.cell(row=row_idx, column=4, value=d["files"]).border = thin_border
                ws4.cell(row=row_idx, column=5, value=d["items"]).border = thin_border
                ws4.cell(row=row_idx, column=6, value=f"{items_pct:.1f}%").border = thin_border

                acc_cell = ws4.cell(row=row_idx, column=7, value=f"{acc_pct:.1f}%")
                acc_cell.border = thin_border
                _color_accuracy(ws4, acc_cell, acc_pct)

                ws4.cell(row=row_idx, column=8, value=d["full"]).border = thin_border
                ws4.cell(row=row_idx, column=9, value=d["high"]).border = thin_border
                ws4.cell(row=row_idx, column=10, value=d["medium"]).border = thin_border
                ws4.cell(row=row_idx, column=11, value=d["low"]).border = thin_border
                ws4.cell(row=row_idx, column=12, value=d["none"]).border = thin_border
                ws4.cell(row=row_idx, column=13, value=round(d["cost"], 2)).border = thin_border
                ws4.cell(row=row_idx, column=14, value=round(cost_per_item, 4)).border = thin_border
                ws4.cell(row=row_idx, column=15, value=round(d["time"] / 60, 1)).border = thin_border
                ws4.cell(row=row_idx, column=16, value=round(time_per_item, 1)).border = thin_border
                ws4.cell(row=row_idx, column=17, value=len(d["senders"])).border = thin_border
                ws4.cell(row=row_idx, column=18, value=f"{success_rate:.1f}%").border = thin_border

            # Totals row for customers
            tot_row = len(sorted_custs) + 2
            tot_font = Font(bold=True, size=11)
            ws4.cell(row=tot_row, column=1, value="סה\"כ").font = tot_font
            ws4.cell(row=tot_row, column=1).border = thin_border
            ws4.cell(row=tot_row, column=2, value=grand_emails).font = tot_font
            ws4.cell(row=tot_row, column=2).border = thin_border
            ws4.cell(row=tot_row, column=5, value=grand_items).font = tot_font
            ws4.cell(row=tot_row, column=5).border = thin_border
            ws4.cell(row=tot_row, column=13, value=round(sum(d["cost"] for d in cust_data.values()), 2)).font = tot_font
            ws4.cell(row=tot_row, column=13).border = thin_border

            auto_width(ws4)

            # ── Sheet 5: סטטיסטיקת שולחים ──
            ws5 = wb.create_sheet("סטטיסטיקת שולחים")
            ws5.sheet_view.rightToLeft = True

            sender_data = defaultdict(lambda: {
                "emails": 0, "files": 0, "items": 0,
                "score": 0, "cost": 0, "time": 0,
                "sent": 0, "customers": set(),
            })
            for e in entries:
                sender = (e.get("sender") or "").lower()
                if not sender:
                    sender = "לא ידוע"
                ad = e.get("accuracy_data", {})
                itm = _get_items(e)
                score = sum(ad.get(k, 0) * weights.get(k, 0) for k in weights)

                sender_data[sender]["emails"] += 1
                sender_data[sender]["files"] += e.get("files_processed", 0)
                sender_data[sender]["items"] += itm
                sender_data[sender]["score"] += score
                sender_data[sender]["cost"] += e.get("cost_usd", 0)
                sender_data[sender]["time"] += e.get("processing_time_seconds", 0)
                sender_data[sender]["sent"] += 1 if e.get("sent") else 0
                for c in (e.get("customers") or []):
                    sender_data[sender]["customers"].add(c.upper())

            sender_cols = [
                "שולח", "לקוחות", "מיילים", "קבצים", "שורות",
                "דיוק %", "עלות $", "עלות/פריט $",
                "זמן (דקות)", "זמן/פריט (שניות)", "שיעור הצלחה",
            ]
            write_header(ws5, sender_cols)

            sorted_senders = sorted(sender_data.items(), key=lambda x: x[1]["emails"], reverse=True)
            for row_idx, (snd, d) in enumerate(sorted_senders, 2):
                acc_pct = (d["score"] / d["items"] * 100) if d["items"] > 0 else 0
                cost_per_item = d["cost"] / d["items"] if d["items"] else 0
                time_per_item = d["time"] / d["items"] if d["items"] else 0
                s_rate = (d["sent"] / d["emails"] * 100) if d["emails"] else 0

                ws5.cell(row=row_idx, column=1, value=snd).border = thin_border
                ws5.cell(row=row_idx, column=2, value=", ".join(sorted(d["customers"]))).border = thin_border
                ws5.cell(row=row_idx, column=3, value=d["emails"]).border = thin_border
                ws5.cell(row=row_idx, column=4, value=d["files"]).border = thin_border
                ws5.cell(row=row_idx, column=5, value=d["items"]).border = thin_border

                acc_cell = ws5.cell(row=row_idx, column=6, value=f"{acc_pct:.1f}%")
                acc_cell.border = thin_border
                _color_accuracy(ws5, acc_cell, acc_pct)

                ws5.cell(row=row_idx, column=7, value=round(d["cost"], 2)).border = thin_border
                ws5.cell(row=row_idx, column=8, value=round(cost_per_item, 4)).border = thin_border
                ws5.cell(row=row_idx, column=9, value=round(d["time"] / 60, 1)).border = thin_border
                ws5.cell(row=row_idx, column=10, value=round(time_per_item, 1)).border = thin_border
                ws5.cell(row=row_idx, column=11, value=f"{s_rate:.1f}%").border = thin_border

            auto_width(ws5)

            wb.save(filepath)
            messagebox.showinfo("ייצוא", f"הדוח נשמר בהצלחה עם 5 גליונות:\n{filepath}")

        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה בייצוא:\n{str(e)}")

    def _reset_stats(self) -> None:
        """Reset statistics — archive old entries and keep only from selected date."""
        from tkinter import messagebox
        from datetime import datetime
        import shutil

        log_path = Path("automation_log.jsonl")
        if not log_path.exists():
            messagebox.showinfo("איפוס", "אין קובץ לוג לאיפוס")
            return

        # Ask user: reset all or keep from date?
        reset_win = tk.Toplevel(self.win)
        reset_win.title("איפוס סטטיסטיקה")
        reset_win.geometry("380x280")
        reset_win.transient(self.win)
        reset_win.grab_set()

        # RTL
        try:
            reset_win.tk.call('tk', 'windowingsystem')
        except Exception:
            pass

        tk.Label(reset_win, text="איפוס סטטיסטיקה", font=("Arial", 14, "bold")).pack(pady=(16, 8))
        tk.Label(reset_win, text="בחר אופן איפוס:", font=("Arial", 10)).pack(pady=4)

        mode_var = tk.StringVar(value="from_date")

        ttk.Radiobutton(reset_win, text="מחק הכל — התחל מאפס",
                        variable=mode_var, value="all").pack(anchor="e", padx=20, pady=2)
        ttk.Radiobutton(reset_win, text="שמור רק מתאריך:",
                        variable=mode_var, value="from_date").pack(anchor="e", padx=20, pady=2)

        # Date + time picker
        date_frame = ttk.Frame(reset_win)
        date_frame.pack(pady=8)
        tk.Label(date_frame, text="מתאריך:").pack(side=tk.RIGHT, padx=4)
        date_entry = ttk.Entry(date_frame, width=12)
        date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
        date_entry.pack(side=tk.RIGHT, padx=2)
        
        tk.Label(date_frame, text="שעה:").pack(side=tk.RIGHT, padx=4)
        time_entry = ttk.Entry(date_frame, width=6)
        time_entry.insert(0, "00:00")
        time_entry.pack(side=tk.RIGHT, padx=2)

        # Info
        entries = self._load_log()
        tk.Label(reset_win, text=f"סה\"כ {len(entries)} רשומות בלוג",
                 font=("Arial", 9), foreground="gray").pack(pady=4)

        def do_reset():
            mode = mode_var.get()

            # Backup first
            backup_name = f"automation_log_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jsonl"
            backup_path = log_path.parent / backup_name
            try:
                shutil.copy2(log_path, backup_path)
            except Exception as e:
                messagebox.showerror("שגיאה", f"שגיאה בגיבוי:\n{e}")
                return

            if mode == "all":
                # Delete everything
                try:
                    log_path.write_text("", encoding="utf-8")
                    messagebox.showinfo("איפוס", f"הלוג אופס.\nגיבוי נשמר: {backup_name}")
                except Exception as e:
                    messagebox.showerror("שגיאה", f"שגיאה באיפוס:\n{e}")
                    return
            else:
                # Keep only from date
                date_val = date_entry.get().strip()
                time_val = time_entry.get().strip()
                
                if not date_val or len(date_val) != 10:
                    messagebox.showerror("שגיאה", "תאריך לא תקין. פורמט: YYYY-MM-DD")
                    return
                
                # Validate time
                if not time_val or len(time_val) < 4:
                    time_val = "00:00"
                if ':' not in time_val:
                    time_val = "00:00"
                
                cutoff = f"{date_val}T{time_val}"
                
                kept = [e for e in entries if str(e.get("timestamp", "")) >= cutoff]
                removed = len(entries) - len(kept)

                try:
                    with open(log_path, "w", encoding="utf-8") as f:
                        for entry in kept:
                            f.write(json.dumps(entry, ensure_ascii=False) + "\n")
                    messagebox.showinfo("איפוס",
                        f"נמחקו {removed} רשומות ישנות.\n"
                        f"נשארו {len(kept)} רשומות מ-{date_val} {time_val}.\n"
                        f"גיבוי: {backup_name}")
                except Exception as e:
                    messagebox.showerror("שגיאה", f"שגיאה בכתיבה:\n{e}")
                    return

            reset_win.destroy()
            self._load_and_display()  # Refresh dashboard

        ttk.Button(reset_win, text="✓ אפס", command=do_reset).pack(pady=12)
        ttk.Button(reset_win, text="ביטול", command=reset_win.destroy).pack()
