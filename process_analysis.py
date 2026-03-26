"""
ניתוח תהליכים, חומרים, קשיחים וצבעים ממאגר הריצות.

שימוש:
  python process_analysis.py              → חלונית GUI לבחירת תאריך
  python process_analysis.py 01/03/2026   → שורת פקודה, ייצור אקסל מ-01/03/2026

הפלט: ANALYSIS_processes_YYYYMMDD.xlsx ב-NEW FILES
"""

import sys
import re
import glob
from pathlib import Path
from datetime import datetime
from collections import Counter

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Configuration ──────────────────────────────────────────────
NEW_FILES_DIR = Path(__file__).parent / "NEW FILES"
OUTPUT_DIR = NEW_FILES_DIR


# ── Helpers ────────────────────────────────────────────────────

def _parse_date(text: str) -> datetime:
    """Parse dd/mm/yyyy or yyyy-mm-dd."""
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%Y %H:%M"):
        try:
            return datetime.strptime(text.strip(), fmt)
        except ValueError:
            continue
    raise ValueError(f"Cannot parse date: {text}")


def _timestamp_from_filename(name: str) -> datetime | None:
    """Extract timestamp from SUMMARY_all_results_YYYYMMDDHHMMSS.xlsx."""
    m = re.search(r'(\d{14})', name)
    if m:
        return datetime.strptime(m.group(1), "%Y%m%d%H%M%S")
    m = re.search(r'(\d{8})', name)
    if m:
        return datetime.strptime(m.group(1), "%Y%m%d")
    return None


def _safe_str(val) -> str:
    """Convert to stripped string, empty for NaN."""
    if pd.isna(val):
        return ""
    s = str(val).strip()
    return "" if s.lower() == "nan" else s


def _split_pipe(text: str) -> list[str]:
    """Split by | and strip."""
    return [p.strip() for p in text.split("|") if p.strip()]


def _extract_inserts_from_bom(bom_text: str) -> list[str]:
    """
    Extract individual insert part numbers from merged_bom.
    
    Format examples:
      קשיחים [2]:
      שרטוט: 1011K500-001 | MS51835
      PL: MS124655 ×12 ×0.45₪ | MS51831-203 ×4
    """
    inserts = []
    for line in bom_text.split("\n"):
        line = line.strip()
        if not line or line.startswith("קשיחים"):
            continue
        # Remove prefix (שרטוט:, PL:, עץ מוצר:, עץ:)
        line = re.sub(r'^(שרטוט|PL|עץ מוצר|עץ)\s*:\s*', '', line)
        # Split by |
        for part in line.split("|"):
            part = part.strip()
            if not part:
                continue
            # Remove quantity and price: ×12 ×0.45₪, (חלופי) etc
            clean = re.sub(r'\s*×[\d.,]+[₪$]?', '', part)
            clean = re.sub(r'\s*\(חלופי\)', '', clean)
            clean = clean.strip().rstrip(',').strip()
            if clean:
                inserts.append(clean)
    return inserts


def _extract_inserts_from_summary(text: str) -> list[str]:
    """Extract inserts from process_summary_hebrew 'קשיחים: XX×N, YY×M'."""
    m = re.search(r'קשיחים\s*:\s*(.+)', text)
    if not m:
        return []
    inserts = []
    for part in m.group(1).split(","):
        part = part.strip()
        clean = re.sub(r'×\d+', '', part).strip()
        if clean:
            inserts.append(clean)
    return inserts


# ── Main analysis ──────────────────────────────────────────────

def run_analysis(from_date: datetime) -> Path:
    """
    Collect all SUMMARY_all_results from from_date onward,
    extract processes/materials/inserts/colors, count occurrences,
    and write to Excel.
    
    Returns: Path to output Excel file.
    """
    # ── Collect files ──
    pattern = str(NEW_FILES_DIR / "SUMMARY_all_results_*.xlsx")
    all_files = glob.glob(pattern)
    
    selected = []
    for fpath in all_files:
        ts = _timestamp_from_filename(Path(fpath).name)
        if ts and ts >= from_date:
            selected.append(fpath)
    
    selected.sort()
    
    if not selected:
        raise FileNotFoundError(
            f"No SUMMARY_all_results files found from {from_date.strftime('%d/%m/%Y')} onward.\n"
            f"Total files in folder: {len(all_files)}"
        )
    
    # ── Load all data ──
    dfs = []
    errors = 0
    for f in selected:
        try:
            df = pd.read_excel(f)
            dfs.append(df)
        except Exception:
            errors += 1
    
    data = pd.concat(dfs, ignore_index=True)
    total_drawings = len(data)
    
    # ── Extract counts ──
    processes_counter: Counter = Counter()
    materials_counter: Counter = Counter()
    inserts_counter: Counter = Counter()
    colors_counter: Counter = Counter()
    
    for _, row in data.iterrows():
        # ── PROCESSES ──
        # Prefer merged_processes, fall back to process_summary_hebrew
        proc_text = _safe_str(row.get("merged_processes")) or _safe_str(row.get("process_summary_hebrew", ""))
        if proc_text:
            parts = _split_pipe(proc_text)
            # First segment is often the material — skip it
            _material_re = re.compile(
                r'(?i)^(אלומיניום|פלדה|נירוסטה|פליז|נחושת|טיטניום|'
                r'aluminum|steel|stainless|brass|copper|titanium|'
                r'al[- ]?\d|ss[- ]?\d|aisi|inconel|invar|kovar|'
                r'\d{4}-[HT])', re.IGNORECASE
            )
            for i, p in enumerate(parts):
                p_clean = p.strip()
                if not p_clean:
                    continue
                # Skip insert lines embedded in process summary
                if p_clean.startswith("קשיחים"):
                    continue
                # Skip first segment if it looks like material
                if i == 0 and _material_re.search(p_clean):
                    continue
                processes_counter[p_clean] += 1
        
        # ── MATERIALS ──
        mat = _safe_str(row.get("material"))
        if mat:
            materials_counter[mat] += 1
        
        # ── INSERTS/HARDWARE ──
        # From merged_bom (structured)
        bom = _safe_str(row.get("merged_bom"))
        if bom:
            for insert in _extract_inserts_from_bom(bom):
                inserts_counter[insert] += 1
        else:
            # Fallback: from process_summary_hebrew
            psh = _safe_str(row.get("process_summary_hebrew", ""))
            if "קשיחים" in psh:
                for insert in _extract_inserts_from_summary(psh):
                    inserts_counter[insert] += 1
        
        # Also from inserts_hardware column (older format)
        hw = _safe_str(row.get("inserts_hardware"))
        if hw:
            for part in _split_pipe(hw):
                clean = re.sub(r'×\d+', '', part).strip()
                if clean:
                    inserts_counter[clean] += 1
        
        # ── COLORS ──
        color = _safe_str(row.get("colors"))
        if color:
            # Sometimes multiple colors separated by comma
            for c in color.split(","):
                c = c.strip()
                if c:
                    colors_counter[c] += 1
        
        # Also extract RAL/color from painting_processes
        paint = _safe_str(row.get("painting_processes"))
        if paint:
            # Find RAL codes
            ral_matches = re.findall(r'RAL\s*\d{4}', paint, re.IGNORECASE)
            for ral in ral_matches:
                colors_counter[ral.upper().replace(" ", "")] += 1
            # Find FED-STD color numbers (e.g., #26231, FED-STD-595 36320)
            fed_matches = re.findall(r'(?:FED[- ]?STD[- ]?595[, ]*)?#(\d{5})', paint)
            for fed in fed_matches:
                colors_counter[f"#{fed}"] += 1
            # Standalone 5-digit FED color codes (but NOT MIL spec numbers)
            fed2 = re.findall(r'\b(\d{5})\b', paint)
            for code in fed2:
                # Skip if preceded by MIL-PRF- or MIL-DTL- (these are spec numbers)
                if re.search(rf'MIL[- ](?:PRF|DTL|C|P)[- ]\d*{code}', paint):
                    continue
                colors_counter[f"#{code}"] += 1
    
    # ── Build output DataFrames ──
    def _counter_to_df(counter: Counter, col_name: str) -> pd.DataFrame:
        items = counter.most_common()
        return pd.DataFrame(items, columns=[col_name, "מופעים"])
    
    df_processes = _counter_to_df(processes_counter, "תהליך")
    df_materials = _counter_to_df(materials_counter, "חומר")
    df_inserts = _counter_to_df(inserts_counter, "קשיח")
    df_colors = _counter_to_df(colors_counter, "צבע")
    
    # ── Write Excel ──
    date_str = from_date.strftime("%Y%m%d")
    output_name = f"ANALYSIS_processes_{date_str}.xlsx"
    output_path = OUTPUT_DIR / output_name
    
    # Summary sheet data
    summary_data = {
        "נתון": ["תאריך התחלה", "קבצים שנסרקו", "שגיאות קריאה", "סה\"כ שרטוטים",
                  "תהליכים ייחודיים", "חומרים ייחודיים", "קשיחים ייחודיים", "צבעים ייחודיים"],
        "ערך": [from_date.strftime("%d/%m/%Y"), len(selected), errors, total_drawings,
                len(processes_counter), len(materials_counter), len(inserts_counter), len(colors_counter)]
    }
    df_summary = pd.DataFrame(summary_data)
    
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="סיכום", index=False)
        df_processes.to_excel(writer, sheet_name="תהליכים", index=False)
        df_materials.to_excel(writer, sheet_name="חומרים", index=False)
        df_inserts.to_excel(writer, sheet_name="קשיחים", index=False)
        df_colors.to_excel(writer, sheet_name="צבעים", index=False)
    
    # ── Format Excel ──
    wb = load_workbook(output_path)
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    sheet_colors = {
        "סיכום": "F2F2F2",
        "תהליכים": "E2EFDA",
        "חומרים": "DAEEF3",
        "קשיחים": "FCE4D6",
        "צבעים": "E4DFEC",
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        fill_color = sheet_colors.get(sheet_name, "FFFFFF")
        data_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # Format headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        # Format data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.fill = data_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        
        # Auto-width columns
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)
        
        # Set RTL
        ws.sheet_view.rightToLeft = True
    
    wb.save(output_path)
    wb.close()
    
    return output_path


# ── GUI ────────────────────────────────────────────────────────

def run_gui():
    """Simple tkinter dialog to pick a date and run analysis."""
    import tkinter as tk
    from tkinter import ttk, messagebox
    import subprocess
    
    root = tk.Tk()
    root.title("ניתוח תהליכים — AI DRAW")
    root.geometry("420x220")
    root.resizable(False, False)
    root.configure(bg="#F0F0F0")
    
    # RTL support
    try:
        root.tk.call('tk', 'scaling', 1.25)
    except Exception:
        pass
    
    frame = ttk.Frame(root, padding=20)
    frame.pack(fill="both", expand=True)
    
    ttk.Label(frame, text="ניתוח תהליכים, חומרים, קשיחים וצבעים",
              font=("Arial", 13, "bold")).pack(pady=(0, 15))
    
    date_frame = ttk.Frame(frame)
    date_frame.pack(pady=5)
    
    ttk.Label(date_frame, text="מתאריך (dd/mm/yyyy):  ", font=("Arial", 11)).pack(side="left")
    date_var = tk.StringVar(value="01/03/2026")
    date_entry = ttk.Entry(date_frame, textvariable=date_var, width=15, font=("Arial", 11), justify="center")
    date_entry.pack(side="left")
    
    status_var = tk.StringVar(value="")
    status_label = ttk.Label(frame, textvariable=status_var, font=("Arial", 10), foreground="gray")
    status_label.pack(pady=10)
    
    def on_run():
        date_text = date_var.get().strip()
        if not date_text:
            messagebox.showerror("שגיאה", "יש להזין תאריך")
            return
        try:
            from_date = _parse_date(date_text)
        except ValueError as e:
            messagebox.showerror("שגיאה", str(e))
            return
        
        status_var.set("⏳ מנתח נתונים...")
        root.update()
        
        try:
            output = run_analysis(from_date)
            status_var.set(f"✅ הקובץ נשמר: {output.name}")
            messagebox.showinfo("הושלם", f"הניתוח הושלם!\n\nקובץ: {output.name}\nתיקייה: {output.parent}")
            # Open file
            try:
                import os
                os.startfile(str(output))
            except Exception:
                pass
        except Exception as e:
            status_var.set("❌ שגיאה")
            messagebox.showerror("שגיאה", str(e))
    
    run_btn = ttk.Button(frame, text="▶  הפעל ניתוח", command=on_run)
    run_btn.pack(pady=5)
    
    root.mainloop()


# ── Entry point ────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) > 1:
        # CLI mode
        from_date = _parse_date(sys.argv[1])
        print(f"Analyzing from {from_date.strftime('%d/%m/%Y')}...")
        output = run_analysis(from_date)
        print(f"✅ Output: {output}")
    else:
        # GUI mode
        run_gui()
