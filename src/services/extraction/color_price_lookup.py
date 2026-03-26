"""
Color / Paint Price Lookup — loads prices from BOM/COLORS.xlsx catalog.

COLORS.xlsx has three sheets:
  PARTS  — col A: external PN, col B: description (with specs), col F: internal PN
  INV    — col A: internal PN, col B: qty in stock
  PRICES — col A: internal PN, col B: supplier, col C: last price, col D: currency

Lookup flow:
  1. Extract spec numbers (MIL-xxx, AMS xxx, etc.) from merged_specs
  2. Match each spec against PARTS col B (substring)
  3. Via internal PN → return ALL suppliers + prices from PRICES
"""

import logging
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

_COLORS_PATH = Path(__file__).resolve().parents[3] / "BOM" / "COLORS.xlsx"

# ── Cached data (loaded once) ──────────────────────────────────
_parts_index: List[Tuple[str, str, str]] = []  # (description_upper, internal_pn, external_pn)
_prices: Dict[str, List[Tuple[str, float, str]]] = {}  # internal_pn → [(supplier, price, currency), ...]
_loaded = False

# ── Paint-process trigger keywords ─────────────────────────────
_PAINT_KEYWORDS = re.compile(
    r'צביע|paint|primer|פריימר|topcoat|טופקאוט|'
    r'צבע\b|ציפוי|coating|coat\b|לכה|lacquer|'
    r'אפוקסי|epoxy|פוליאורתן|polyurethane',
    re.IGNORECASE,
)

# ── Color / finish hint words (for narrowing catalog matches) ──
_COLOR_WORDS_HE = ['אפור', 'לבן', 'שחור', 'צהוב', 'ירוק', 'כחול', 'אדום', 'חום', 'כתום', 'סגול']
_COLOR_WORDS_EN = ['GRAY', 'GREY', 'WHITE', 'BLACK', 'YELLOW', 'GREEN', 'BLUE', 'RED', 'BROWN', 'ORANGE']
_FINISH_WORDS_HE = ['מט', 'מבריק', 'משי']
_FINISH_WORDS_EN = ['MATT', 'MATTE', 'GLOSS', 'SEMI', 'SATIN']
_ALL_HINTS = (
    _COLOR_WORDS_HE + _COLOR_WORDS_EN +
    _FINISH_WORDS_HE + _FINISH_WORDS_EN
)

# ── Spec extraction regex ──────────────────────────────────────
_SPEC_RE = re.compile(
    r'((?:MIL|AMS|ASTM|BMS|DMS|RAFDOC|FED|QPL|TT|DOD|SAE)'
    r'[\s\-#]?'
    r'[\w][\w.\-/]*'
    r'|RAL\s*\d{4}(?:/\w+)?)',          # RAL colour codes, e.g. RAL 6003, RAL6003, RAL6031/F9
    re.IGNORECASE,
)


def _load_catalog() -> None:
    """Load BOM/COLORS.xlsx PARTS + PRICES sheets into memory once."""
    global _parts_index, _prices, _loaded
    if _loaded:
        return
    _loaded = True

    if not _COLORS_PATH.exists():
        logger.warning(f"[COLOR-PRICE] Catalog not found: {_COLORS_PATH}")
        return

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(_COLORS_PATH), data_only=True, read_only=True)

        # ── PARTS sheet ──
        ws_parts = wb['PARTS']
        for row in ws_parts.iter_rows(min_row=2, values_only=True):
            desc = str(row[1] or '').strip()
            internal_pn = str(row[5] or '').strip() if len(row) > 5 else ''
            external_pn = str(row[0] or '').strip()
            if desc and internal_pn:
                _parts_index.append((desc.upper(), internal_pn, external_pn))

        # ── PRICES sheet ──
        ws_prices = wb['PRICES']
        for row in ws_prices.iter_rows(min_row=2, values_only=True):
            pn = str(row[0] or '').strip()
            supplier = str(row[1] or '').strip()
            price_val = row[2]
            currency = str(row[3] or '').strip()
            if not pn or price_val is None:
                continue
            try:
                price = float(price_val)
            except (ValueError, TypeError):
                continue
            if price <= 0:
                continue
            if not currency:
                currency = '₪'
            _prices.setdefault(pn, []).append((supplier, price, currency))

        wb.close()
        logger.info(
            f"[COLOR-PRICE] Loaded {len(_parts_index)} parts, "
            f"{len(_prices)} priced items from COLORS.xlsx"
        )
    except Exception as e:
        logger.error(f"[COLOR-PRICE] Failed to load catalog: {e}")


def has_paint_process(merged_processes: str) -> bool:
    """Return True if merged_processes contains a painting/coating keyword."""
    if not merged_processes:
        return False
    return bool(_PAINT_KEYWORDS.search(merged_processes))


def _extract_specs(merged_specs: str) -> List[str]:
    """Extract spec identifiers from merged_specs string.

    Input like: "תמורה - MIL-DTL-5541 Type I Class 1A | צביעה - MIL-PRF-85285"
    Returns:    ["MIL-DTL-5541", "MIL-PRF-85285"]
    """
    if not merged_specs:
        return []
    return [m.group(1).strip().rstrip('.') for m in _SPEC_RE.finditer(merged_specs)]


def _extract_spec_sections(merged_specs: str) -> List[List[str]]:
    """Split merged_specs by | and extract spec ids per section.

    Input:  "תמורה - MIL-DTL-5541 Type I | צביעה - MIL-PRF-85285 FED 37875"
    Returns: [["MIL-DTL-5541"], ["MIL-PRF-85285", "FED 37875"]]

    Each inner list contains the specs from ONE pipe-delimited section.
    When multiple specs appear in the same section (e.g. MIL-PRF-85285
    and FED 37875), they should be matched together (AND) to narrow results.
    """
    if not merged_specs:
        return []
    sections: List[List[str]] = []
    for part in merged_specs.split('|'):
        specs = [m.group(1).strip().rstrip('.') for m in _SPEC_RE.finditer(part)]
        if specs:
            sections.append(specs)
    return sections


def _find_parts_by_spec(spec: str) -> List[Tuple[str, str, str]]:
    """Find PARTS rows whose description contains the given spec.

    Returns list of (description, internal_pn, external_pn).
    """
    spec_upper = spec.upper()
    # Normalize: MIL-PRF-85285 should match "MIL-PRF-85285", "Mil-PRF 85285" etc.
    # Build flexible pattern: allow dash/space/nothing between parts
    parts = re.split(r'[\s\-]+', spec_upper)
    if len(parts) < 2:
        # Single token — direct substring
        return [(d, ipn, epn) for d, ipn, epn in _parts_index if spec_upper in d]

    # Build pattern that allows dash or space between parts
    pattern = r'[\s\-]*'.join(re.escape(p) for p in parts)
    try:
        rx = re.compile(pattern, re.IGNORECASE)
    except re.error:
        return [(d, ipn, epn) for d, ipn, epn in _parts_index if spec_upper in d]

    return [(d, ipn, epn) for d, ipn, epn in _parts_index if rx.search(d)]


def _get_prices_for_pn(internal_pn: str) -> List[Tuple[str, float, str]]:
    """Get all (supplier, price, currency) for an internal PN."""
    return _prices.get(internal_pn, [])


def lookup_color_prices(merged_specs: str, merged_processes: str) -> str:
    """Look up paint/color prices for a drawing item.

    Only triggers when merged_processes contains a painting keyword.
    Extracts specs from merged_specs, matches against COLORS.xlsx,
    and returns a formatted string with all suppliers and prices.

    Strategy:
      1. Split merged_specs into sections (pipe-separated).
      2. For each section, extract ALL spec ids and search PARTS
         using the narrowest combination (AND) first.
      3. If the narrow search finds results, use those; otherwise
         fall back to just the base spec.
      4. Limit results per section to keep output readable.

    Returns
    -------
    str
        Formatted string with matched products and their suppliers/prices.
        Empty string if no matches or no paint process.
    """
    if not has_paint_process(merged_processes):
        return ''

    _load_catalog()

    sections = _extract_spec_sections(merged_specs)
    if not sections:
        return ''

    # Extract colour/finish hints from processes (e.g. "אפור", "מט")
    color_hints = _extract_color_hints(merged_processes)

    MAX_PER_SECTION = 8
    results: list[str] = []
    seen_internal_pns: set[str] = set()

    for spec_ids in sections:
        if not spec_ids:
            continue

        # Try narrowest first: match ALL specs in section simultaneously
        if len(spec_ids) > 1:
            narrow = _find_parts_by_all_specs(spec_ids)
            if narrow:
                matching_parts = narrow
            else:
                # Fall back to just the base spec (first one)
                matching_parts = _find_parts_by_spec(spec_ids[0])
        else:
            matching_parts = _find_parts_by_spec(spec_ids[0])

        # Further narrow by colour/finish hints (with fallback)
        matching_parts = _narrow_by_hints(matching_parts, color_hints)

        section_label = ' + '.join(spec_ids)
        added = 0

        for desc, internal_pn, external_pn in matching_parts:
            if internal_pn in seen_internal_pns:
                continue
            seen_internal_pns.add(internal_pn)

            prices = _get_prices_for_pn(internal_pn)
            if not prices:
                continue

            supplier_parts = []
            for supplier, price, currency in prices:
                supplier_short = supplier[:20].strip()
                price_str = f"{price:g}"
                supplier_parts.append(f"{supplier_short} {price_str}{currency}")

            suppliers_str = ' | '.join(supplier_parts)
            # Use original (non-upper) description from desc
            # _parts_index stores upper case — recover from matching parts
            desc_display = desc[:60].strip()
            results.append(f"{section_label}: [{internal_pn}] {desc_display} → {suppliers_str}")
            added += 1
            if added >= MAX_PER_SECTION:
                remaining = len(matching_parts) - added
                if remaining > 0:
                    results.append(f"  ... +{remaining} more items for {section_label}")
                break

    return '\n'.join(results)


def _extract_color_hints(merged_processes: str) -> List[str]:
    """Extract color/finish hint words from merged_processes.

    Looks for Hebrew and English colour names (אפור, BLACK, …)
    and finish terms (מט, GLOSS, …).  Returns list of matched words
    (upper-cased for English, original for Hebrew).
    """
    if not merged_processes:
        return []
    hints: List[str] = []
    proc_upper = merged_processes.upper()
    for w in _COLOR_WORDS_HE + _FINISH_WORDS_HE:
        if w in merged_processes:
            hints.append(w)
    for w in _COLOR_WORDS_EN + _FINISH_WORDS_EN:
        if w in proc_upper:
            hints.append(w)
    return hints


def _narrow_by_hints(
    candidates: List[Tuple[str, str, str]],
    hints: List[str],
) -> List[Tuple[str, str, str]]:
    """Filter candidates whose description contains ALL hint words.

    Falls back to unfiltered candidates if narrowing yields zero results.
    """
    if not hints or not candidates:
        return candidates
    narrowed = candidates
    for hint in hints:
        hint_upper = hint.upper()
        filtered = [(d, i, e) for d, i, e in narrowed if hint_upper in d]
        if filtered:
            narrowed = filtered
        # else: skip this hint — too restrictive
    return narrowed


def _find_parts_by_all_specs(spec_ids: List[str]) -> List[Tuple[str, str, str]]:
    """Find PARTS rows whose description contains ALL given specs.

    Used for narrow matching: e.g. MIL-PRF-85285 AND FED 37875.
    """
    if not spec_ids:
        return []

    # Start with results for first spec, then intersect
    candidates = _find_parts_by_spec(spec_ids[0])
    for additional_spec in spec_ids[1:]:
        additional_spec_upper = additional_spec.upper()
        parts = re.split(r'[\s\-]+', additional_spec_upper)
        pattern = r'[\s\-]*'.join(re.escape(p) for p in parts)
        try:
            rx = re.compile(pattern, re.IGNORECASE)
        except re.error:
            candidates = [(d, i, e) for d, i, e in candidates if additional_spec_upper in d]
            continue
        candidates = [(d, i, e) for d, i, e in candidates if rx.search(d)]

    return candidates
