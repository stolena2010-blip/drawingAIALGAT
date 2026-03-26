"""
Excel Export – classification report, PL sheet update, results workbook
=======================================================================
Extracted from customer_extractor_v3_dual.py  (Phase 2.5)
"""

from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional

import pandas as pd

from src.core.constants import WARN_FILE_SIZE_MB, MAX_FILE_SIZE_MB
from src.services.extraction.filename_utils import (
    _extract_item_number_from_filename,
    _normalize_item_number,
)
from src.services.reporting.pl_generator import _generate_pl_summary_english
from src.utils.logger import get_logger

logger = get_logger(__name__)

# ── public API ──────────────────────────────────────────────────────────
__all__ = [
    "_save_classification_report",
    "_update_pl_sheet_with_associated_items",
    "_save_results_to_excel",
]


# ────────────────────────────────────────────────────────────────────────
def _save_classification_report(file_classifications: List[Dict], project_folder: Path, 
                                tokens_in: int, tokens_out: int, custom_filename: str = None,
                                drawing_map: Optional[Dict[str, str]] = None, drawing_results: Optional[List[Dict]] = None) -> None:
    """Save file classification report to Excel with metadata"""
    try:
        if custom_filename:
            output_path = project_folder / custom_filename
        else:
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            output_path = project_folder / f"file_classification_{timestamp}.xlsx"
        
        # Lazy imports to avoid circular dependency
        from src.services.file.file_utils import _get_file_metadata, _find_associated_drawing

        # Prepare data with metadata
        report_data = []
        large_files_count = 0
        too_large_files_count = 0
        
        for fc in file_classifications:
            metadata = _get_file_metadata(fc['file_path'])
            
            if metadata['is_large_file']:
                large_files_count += 1
            if metadata['is_too_large']:
                too_large_files_count += 1
            
            # Add warning to description for large files
            description = fc['description']
            if metadata['is_too_large']:
                description = f" TOO LARGE ({metadata['file_size_mb']}MB) - SKIPPED. {description}"
            elif metadata['is_large_file']:
                description = f" LARGE FILE ({metadata['file_size_mb']}MB). {description}"
            
            # Determine associated item based on file type
            associated_item = ""
            
            # תיקון חזק: אם הקובץ הוא תמונה בפועל (jpg, png וכו'), וודא שהסיווג הוא 3D_IMAGE
            file_type = fc.get('file_type', 'OTHER')
            ext_lower = fc['file_path'].suffix.lower()
            if ext_lower in {'.jpg', '.jpeg', '.png', '.tif', '.tiff', '.bmp', '.gif', '.webp'}:
                # זה תמונה בטוח - תיקון הסיווג אם לא נכון
                if file_type != '3D_IMAGE':
                    logger.info(f"🔧 Correcting file type for {fc['file_path'].name}: {file_type} → 3D_IMAGE")
                    fc['file_type'] = '3D_IMAGE'
                    file_type = '3D_IMAGE'
                    # חשוב: גם לנקות את renamed_filename כדי שיינוצר מחדש עם הסיווג הנכון
                    fc['renamed_filename'] = ''
            
            if file_type == 'DRAWING' and drawing_results:
                # For DRAWING: find part_number from drawing_results
                file_name = fc['file_path'].name
                for dr in drawing_results:
                    if dr.get('file_name') == file_name:
                        associated_item = dr.get('part_number', '')
                        break
            elif fc.get('file_type') in ['3D_MODEL', '3D_IMAGE', 'PARTS_LIST', 'OTHER'] and drawing_map:
                # For MODEL/IMAGE/PL/OTHER: find associated drawing by filename similarity
                associated_item = _find_associated_drawing(fc['file_path'], fc.get('file_type', ''), drawing_map)
                if fc.get('file_type') == '3D_IMAGE' and associated_item:
                    logger.info(f"✓ 3D_IMAGE associated: {fc['file_path'].name} → {associated_item}")
                elif fc.get('file_type') == '3D_IMAGE':
                    item_num = _extract_item_number_from_filename(fc['file_path'].name)
                    logger.warning(f"⚠ 3D_IMAGE NOT associated: {fc['file_path'].name} (extracted: {item_num}, map has: {list(drawing_map.keys())})")
            
            # Save associated_item back to file_classifications for later use (renaming)
            fc['associated_item'] = associated_item
            
            # Extract drawing_number and revision for report
            drawing_number = ''
            revision = ''
            if fc.get('file_type') == 'DRAWING' and drawing_results:
                # For DRAWING: find by filename
                file_name = fc['file_path'].name
                for dr in drawing_results:
                    if dr.get('file_name') == file_name:
                        drawing_number = dr.get('drawing_number', '')
                        revision = dr.get('revision', '')
                        break
            elif associated_item and drawing_results:
                # For non-DRAWING files: find drawing_number and revision by associated_item (part_number)
                for dr in drawing_results:
                    if dr.get('part_number', '') == associated_item:
                        drawing_number = dr.get('drawing_number', '')
                        revision = dr.get('revision', '')
                        break
            
            # Build DISPLAY NAME for any file with associated_item
            # For DRAWING / 3D_MODEL / 3D_IMAGE: drawing_number + TAB + associated_item + TAB + revision + TAB + extension
            # For other files: original_filename + TAB + associated_item + TAB + revision + TAB + extension
            display_name = ''
            if associated_item:
                # תיקון: השתמש בערך file_type המתוקן (לא הערך הישן מ-fc)
                file_extension = fc['file_path'].suffix  # כולל נקודה, למשל ".pdf"
                if file_type in ('DRAWING', '3D_MODEL', '3D_IMAGE') and drawing_number:
                    # For drawings and 3D files, use drawing_number
                    display_name = f"{drawing_number} \t{associated_item} \t{revision} \t{file_extension}"
                elif associated_item:
                    # For non-drawing files, use original filename (without extension)
                    original_filename = fc.get('original_filename', fc['file_path'].name)
                    # Remove extension for cleaner display
                    original_name_no_ext = Path(original_filename).stem
                    display_name = f"{original_name_no_ext} \t{associated_item} \t{revision} \t{file_extension}"
            fc['display_name'] = display_name

            # Ensure RENAMED_FILENAME is populated for report even if rename happened later
            renamed_filename = fc.get('renamed_filename', '')
            if not renamed_filename and associated_item:
                # השתמש בערך file_type המתוקן (לא הערך הישן מ-fc)
                if file_type == 'DRAWING':
                    prefix = 'B2BDraw'
                    type_suffix = '_30'
                elif file_type == '3D_MODEL':
                    prefix = 'B2BModel'
                    type_suffix = '_25'
                elif file_type == '3D_IMAGE':
                    prefix = 'B2BImg'
                    type_suffix = '_16'
                else:  # PARTS_LIST, OTHER, QUOTE, INVOICE, PURCHASE_ORDER
                    prefix = 'B2BDoc'
                    type_suffix = '_99'
                original_name_no_ext = fc['file_path'].stem
                ext = fc['file_path'].suffix
                # Match auto-rename pattern: B2B{Type}_{original_name}{suffix}.{ext}
                renamed_filename = f"{prefix}_{original_name_no_ext}{type_suffix}{ext}"
                fc['renamed_filename'] = renamed_filename
            # תיקון: תמונות תמיד צריכות _16, גם אם אין associated_item
            elif not renamed_filename and file_type == '3D_IMAGE':
                prefix = 'B2BImg'
                type_suffix = '_16'
                original_name_no_ext = fc['file_path'].stem
                ext = fc['file_path'].suffix
                renamed_filename = f"{prefix}_{original_name_no_ext}{type_suffix}{ext}"
                fc['renamed_filename'] = renamed_filename

            display_name_visual = display_name.replace("\t", " | ") if display_name else ''

            # Use stored original filename (before rename happens later)
            original_filename = fc.get('original_filename', fc['file_path'].name)
            
            report_data.append({
                'FILE_NAME': original_filename,
                'RENAMED_FILENAME': renamed_filename,
                'DISPLAY_NAME': display_name,
                'DISPLAY_NAME_VISUAL': display_name_visual,
                'drawing_number': drawing_number,
                'revision': revision,
                'file_title': metadata['file_title'],
                'file_size_mb': metadata['file_size_mb'],
                'file_path': str(fc['file_path']),
                'file_type': fc['file_type'],
                'extension': fc['file_path'].suffix.lower(),
                'quote_number': fc.get('quote_number', ''),
                'order_number': fc.get('order_number', ''),
                'associated_item': associated_item,
                'description': description,
                'created_date': metadata['created_date'],
                'modified_date': metadata['modified_date']
            })
        
        df = pd.DataFrame(report_data)
        df.to_excel(output_path, index=False)
        
        logger.info(f"Classification report saved: {output_path}")
        logger.info(f"{len(file_classifications)} files classified")
        logger.info(f"Metadata included: file names, titles, sizes, dates")
        
        if large_files_count > 0:
            logger.info(f"Large files detected:")
            logger.warning(f"- Files > {WARN_FILE_SIZE_MB}MB: {large_files_count}")
        if too_large_files_count > 0:
            logger.info(f"- Files > {MAX_FILE_SIZE_MB}MB (TOO LARGE, will be skipped): {too_large_files_count}")
        
    except Exception as e:
        logger.error(f"Failed to save classification report: {e}")


# ────────────────────────────────────────────────────────────────────────
def _update_pl_sheet_with_associated_items(drawing_results_path: Path, file_classification_path: Path) -> None:
    """
    Post-process: Update Parts_List_Items sheet with associated_item data from file_classifications.
    
    Reads:
    1. file_classifications Excel (עמודה A = filename, עמודה associated_item)
    2. drawing_results Excel Parts_List_Items sheet (עמודה A = PL filename)
    
    Updates:
    - drawing_results Parts_List_Items, עמודה D (Associated Item) with values from file_classifications
    """
    try:
        # Defensive None checking
        if drawing_results_path is None or file_classification_path is None:
            logger.info(f"Skipping PL update: paths are None (results={drawing_results_path is not None}, classification={file_classification_path is not None})")
            return
        
        if not drawing_results_path.exists():
            logger.info(f"drawing_results file not found: {drawing_results_path.name if hasattr(drawing_results_path, 'name') else drawing_results_path}")
            return
        if not file_classification_path.exists():
            logger.info(f"file_classification file not found: {file_classification_path.name if hasattr(file_classification_path, 'name') else file_classification_path}")
            return
        
        # Read file_classifications Excel
        try:
            df_classifications = pd.read_excel(file_classification_path)
            # Build map: filename -> associated_item
            classification_map = {}
            for _, row in df_classifications.iterrows():
                filename = str(row.get('FILE_NAME', '')).strip()
                associated = str(row.get('associated_item', '')).strip()
                if filename and associated and associated != 'nan':
                    classification_map[filename] = associated
            logger.info(f"Built classification map with {len(classification_map)} entries")
        except Exception as e:
            logger.error(f"Error reading file_classifications: {e}")
            return
        
        if not classification_map:
            logger.info(f"No associated_items found in file_classifications")
            return
        
        # Update drawing_results Excel
        try:
            from openpyxl import load_workbook
            
            if not drawing_results_path or not isinstance(drawing_results_path, Path):
                logger.info(f"Invalid drawing_results_path: {drawing_results_path}")
                return
            
            wb = load_workbook(drawing_results_path)
            if wb is None:
                logger.error(f"Failed to load workbook from {drawing_results_path.name}")
                return
            
            if "Parts_List_Items" not in wb.sheetnames:
                logger.info(f"No Parts_List_Items sheet found in drawing_results")
                wb.close()
                return
            
            ws_pl = wb["Parts_List_Items"]
            if ws_pl is None:
                logger.error(f"Failed to access Parts_List_Items sheet")
                wb.close()
                return
            
            # Find column indices by header
            header_row = ws_pl[1]
            col_a_idx = None  # PL Filename
            col_d_idx = None  # Associated Item
            
            for col_idx, cell in enumerate(header_row, 1):
                if cell.value == 'PL Filename':
                    col_a_idx = col_idx
                elif cell.value == 'Associated Item':
                    col_d_idx = col_idx
            
            if not col_a_idx or not col_d_idx:
                logger.info(f"Missing required columns: PL Filename={col_a_idx}, Associated Item={col_d_idx}")
                return
            
            # Update rows
            updated_count = 0
            for row_idx in range(2, ws_pl.max_row + 1):
                pl_filename = ws_pl.cell(row=row_idx, column=col_a_idx).value
                if pl_filename and isinstance(pl_filename, str):
                    pl_filename = pl_filename.strip()
                    if pl_filename in classification_map:
                        ws_pl.cell(row=row_idx, column=col_d_idx).value = classification_map[pl_filename]
                        updated_count += 1
            
            if updated_count > 0:
                wb.save(drawing_results_path)
                logger.info(f"✓ Updated {updated_count} rows with associated_items")
            else:
                logger.info(f"No matching filenames found between sheets (checked {ws_pl.max_row - 1} data rows)")
            
            wb.close()
        except Exception as e:
            logger.error(f"Error updating drawing_results: {e}")
            import traceback
            traceback.print_exc()
    
    except Exception as e:
        logger.error(f"Error in _update_pl_sheet_with_associated_items: {e}")
        import traceback
        traceback.print_exc()


# ────────────────────────────────────────────────────────────────────────
def _save_results_to_excel(results: List[Dict], output_path: Path, pl_items: List[Dict] = None) -> int:
    """Save results to Excel file with optional second sheet for PL items.
    
    Args:
        results: List of result dictionaries (drawings)
        output_path: Path to Excel file
        pl_items: Optional list of PL item dictionaries
    
    Returns:
        Number of RAFAEL rows found
    """
    if not results:
        return 0
    
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        
        # Add PL summary columns if pl_items provided
        if pl_items:
            try:
                for result in results:
                    if not result:  # Skip None results
                        continue
                    part_number = result.get('part_number', '')
                    # If part_number was overridden by PL, also try matching with original OCR PN
                    ocr_original = result.get('part_number_ocr_original', '')
                    pl_summary_english = _generate_pl_summary_english(part_number, pl_items, ocr_original)
                    # AI Summary: Get all summaries from matched items
                    ai_summaries = []
                    part_num_norm = _normalize_item_number(str(part_number))
                    ocr_orig_norm = _normalize_item_number(str(ocr_original)) if ocr_original else ''
                    
                    def _match_associated(associated_norm, part_norm, ocr_norm) -> bool:
                        """Check if associated_item matches part_number or ocr_original."""
                        if (associated_norm == part_norm or
                            associated_norm in part_norm or
                            part_norm in associated_norm):
                            return True
                        if ocr_norm and (associated_norm == ocr_norm or
                            associated_norm in ocr_norm or
                            ocr_norm in associated_norm):
                            return True
                        return False
                    
                    # PL Main Part Number: find from matching PL items
                    pl_main_pn = result.get('pl_main_part_number', '')
                    if not pl_main_pn:
                        for pl_item in pl_items:
                            if not pl_item:
                                continue
                            associated = pl_item.get('associated_item', '')
                            associated_norm = _normalize_item_number(str(associated))
                            if _match_associated(associated_norm, part_num_norm, ocr_orig_norm):
                                pl_main_pn = pl_item.get('pl_main_part_number', '')
                                if pl_main_pn:
                                    break
                    for pl_item in pl_items:
                        if not pl_item:  # Skip None pl_items
                            continue
                        associated = pl_item.get('associated_item', '')
                        associated_norm = _normalize_item_number(str(associated))
                        if _match_associated(associated_norm, part_num_norm, ocr_orig_norm):
                            if pl_item.get('summary'):
                                ai_summaries.append(pl_item.get('summary', ''))
                    result['PL Part Number'] = pl_main_pn
                    result['PL Summary'] = pl_summary_english
                    result['PL Summary (AI)'] = ' | '.join(ai_summaries) if ai_summaries else ""
                    # PL Override Note — shows if part_number was corrected by PL
                    result['PL Override Note'] = result.get('pl_override_note', '')

                    # Structured PL fields (from AI extraction)
                    pl_proc_list = []
                    pl_paint_list = []
                    pl_hw_list = []
                    pl_mat_list = []
                    pl_heb_list = []
                    for pl_item in pl_items:
                        if not pl_item:
                            continue
                        associated = pl_item.get('associated_item', '')
                        associated_norm = _normalize_item_number(str(associated))
                        if _match_associated(associated_norm, part_num_norm, ocr_orig_norm):
                            if pl_item.get('pl_processes'):
                                pl_proc_list.append(pl_item['pl_processes'])
                            if pl_item.get('pl_paint'):
                                pl_paint_list.append(pl_item['pl_paint'])
                            if pl_item.get('pl_hardware'):
                                pl_hw_list.append(pl_item['pl_hardware'])
                            if pl_item.get('pl_material'):
                                pl_mat_list.append(pl_item['pl_material'])
                            if pl_item.get('pl_summary_hebrew'):
                                pl_heb_list.append(pl_item['pl_summary_hebrew'])
                    
                    result['PL Processes'] = ' | '.join(dict.fromkeys(pl_proc_list)) if pl_proc_list else ''
                    result['PL Paint'] = ' | '.join(dict.fromkeys(pl_paint_list)) if pl_paint_list else ''
                    result['PL Hardware'] = ' | '.join(dict.fromkeys(pl_hw_list)) if pl_hw_list else ''
                    result['PL Material'] = ' | '.join(dict.fromkeys(pl_mat_list)) if pl_mat_list else ''
                    result['PL Summary Hebrew'] = ' | '.join(dict.fromkeys(pl_heb_list)) if pl_heb_list else ''
            except Exception as add_pl_error:
                logger.error(f"WARNING: Error adding PL summaries: {str(add_pl_error)}")
                # Continue without PL summaries if there's an error
        
        # Format inserts_hardware list → readable string for Excel
        for result in results:
            hw_list = result.get('inserts_hardware')
            if isinstance(hw_list, list) and hw_list:
                hw_parts = []
                for hw in hw_list:
                    if isinstance(hw, dict):
                        cat = hw.get('cat_no', '')
                        qty = hw.get('qty', '')
                        desc = hw.get('description', '')
                        price = hw.get('unit_price')
                        currency = hw.get('currency', '')
                        # Format: cat_no ×qty ×price currency
                        entry = f"{cat}×{qty}" if cat and qty else cat or desc
                        if price is not None:
                            entry += f" ×{price}{currency}"
                        elif desc and cat:
                            entry += f" ({desc})"
                        hw_parts.append(entry)
                result['inserts_hardware'] = ' | '.join(hw_parts)
            else:
                result['inserts_hardware'] = ''

        df = pd.DataFrame(results)

        # Reorder: put merged_* columns after work description columns
        merged_cols = ['merged_description', 'merged_processes', 'merged_specs', 'merged_bom', 'merged_notes', 'color_prices']
        existing_merged = [c for c in merged_cols if c in df.columns]
        if existing_merged:
            other_cols = [c for c in df.columns if c not in existing_merged]
            # Find insertion point: after work_description_doc or work_description_email
            insert_after = None
            for anchor in ('work_description_doc', 'work_description_email', 'PL Summary Hebrew', 'process_summary_hebrew'):
                if anchor in other_cols:
                    insert_after = other_cols.index(anchor) + 1
                    break
            if insert_after is not None:
                new_order = other_cols[:insert_after] + existing_merged + other_cols[insert_after:]
            else:
                new_order = other_cols + existing_merged
            df = df[new_order]

        # Add accuracy_score column based on confidence_level and weights from .env
        try:
            import os
            _acc_weights = {
                "full":   float(os.getenv("ACCURACY_WEIGHT_FULL", "1.0")),
                "high":   float(os.getenv("ACCURACY_WEIGHT_HIGH", "1.0")),
                "medium": float(os.getenv("ACCURACY_WEIGHT_MEDIUM", "0.8")),
                "low":    float(os.getenv("ACCURACY_WEIGHT_LOW", "0.5")),
                "none":   float(os.getenv("ACCURACY_WEIGHT_NONE", "0.0")),
            }
            if 'confidence_level' in df.columns:
                df['accuracy_score'] = df['confidence_level'].apply(
                    lambda c: _acc_weights.get(str(c).strip().lower(), _acc_weights['none'])
                )
        except Exception as _acc_err:
            logger.warning(f"WARNING: Could not add accuracy_score column: {_acc_err}")

        # Save to Excel
        df.to_excel(output_path, index=False)
        
        # Apply formatting to RAFAEL rows
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Define RAFAEL highlight color (light blue)
        rafael_fill = PatternFill(start_color="B3D9FF", end_color="B3D9FF", fill_type="solid")
        
        # Find customer_name column index
        headers = [cell.value for cell in ws[1]]
        try:
            customer_col_idx = headers.index('customer_name') + 1  # openpyxl is 1-indexed
        except ValueError:
            customer_col_idx = None
        
        # Find part_number and part_number_ocr_original column indices for PL override coloring
        try:
            pn_col_idx = headers.index('part_number') + 1
        except ValueError:
            pn_col_idx = None
        try:
            ocr_orig_col_idx = headers.index('part_number_ocr_original') + 1
        except ValueError:
            ocr_orig_col_idx = None
        
        # Dark blue font for PL-overridden part numbers
        from openpyxl.styles import Font as _Font
        pl_override_font = _Font(color="00008B", bold=True)  # Dark blue
        
        # Color RAFAEL rows + PL override part numbers
        rafael_count = 0
        for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
            # RAFAEL row highlighting
            if customer_col_idx:
                customer_cell = ws.cell(row=row_idx, column=customer_col_idx)
                customer_value = customer_cell.value
                
                if customer_value and isinstance(customer_value, str) and "RAFAEL" in customer_value.upper():
                    # Color entire row
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = rafael_fill
                    rafael_count += 1
            
            # PL override: dark blue font on part_number cell
            if pn_col_idx and ocr_orig_col_idx:
                ocr_orig_value = ws.cell(row=row_idx, column=ocr_orig_col_idx).value
                if ocr_orig_value and str(ocr_orig_value).strip():
                    ws.cell(row=row_idx, column=pn_col_idx).font = pl_override_font
        
        # Save workbook (but keep open if we need to add PL sheet)
        wb.save(output_path)
        
        # Add PL items sheet if provided (Stage 6 output)
        if pl_items:
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                wb_pl = load_workbook(output_path)
                
                # Check if sheet exists, if not create it
                if "Parts_List_Items" in wb_pl.sheetnames:
                    logger.info(f"Parts_List_Items sheet exists, appending {len(pl_items)} items")
                    ws_pl = wb_pl["Parts_List_Items"]
                    # Get the last row with data
                    next_row = ws_pl.max_row + 1
                else:
                    logger.info(f"Creating new Parts_List_Items sheet with {len(pl_items)} items")
                    ws_pl = wb_pl.create_sheet(title="Parts_List_Items")
                    next_row = 1
                    
                    # Add headers only if sheet is new
                    headers = ['PL Filename', 'Item Number', 'Description', 'Associated Item', 'Matched Item', 'Drawing Part Number',
                              'Quantity', 'Processes', 'Specifications', 'Product Tree', 'Item Type', 'Summary (AI)']
                    ws_pl.append(headers)
                    
                    # Format header row
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    for cell in ws_pl[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
                    next_row = 2
                
                # Add data rows
                for item in pl_items:
                    # Convert lists to comma-separated strings
                    processes_str = ', '.join(item.get('processes', [])) if isinstance(item.get('processes'), list) else str(item.get('processes', ''))
                    specs_str = ', '.join(item.get('specifications', [])) if isinstance(item.get('specifications'), list) else str(item.get('specifications', ''))
                    
                    ws_pl.append([
                        item.get('pl_filename', ''),
                        item.get('item_number', ''),
                        item.get('description', ''),
                        item.get('associated_item', ''),
                        item.get('matched_item_name', ''),
                        item.get('matched_drawing_part_number', ''),
                        item.get('quantity', ''),
                        processes_str,
                        specs_str,
                        item.get('product_tree', ''),
                        item.get('item_type', ''),
                        item.get('summary', '')
                    ])
                
                # Set column widths and wrap text (do this every time)
                col_widths = [20, 15, 25, 18, 20, 18, 12, 25, 30, 35, 15, 50]
                for idx, width in enumerate(col_widths, 1):
                    ws_pl.column_dimensions[chr(64 + idx)].width = width
                    ws_pl.column_dimensions[chr(64 + idx)].alignment = Alignment(wrap_text=True)
                
                wb_pl.save(output_path)
                wb_pl.close()
                logger.info(f"Added PL Items sheet with {len(pl_items)} items to {output_path.name}")
            except Exception as pl_error:
                logger.error(f"Error adding PL Items sheet: {pl_error}")
                wb.close()
        else:
            wb.close()
        
        return rafael_count
        
    except Exception as e:
        logger.error(f"Error saving Excel: {e}")
        return 0
