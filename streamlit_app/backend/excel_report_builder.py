"""Shared Excel report builder used by Dashboard and Scheduler exports."""
from __future__ import annotations

import os
from collections import Counter, defaultdict
from datetime import datetime
from typing import Any, Dict, Optional

_WEIGHTS = {
    "full": float(os.getenv("ACCURACY_WEIGHT_FULL", "1.0")),
    "high": float(os.getenv("ACCURACY_WEIGHT_HIGH", "1.0")),
    "medium": float(os.getenv("ACCURACY_WEIGHT_MEDIUM", "0.8")),
    "low": float(os.getenv("ACCURACY_WEIGHT_LOW", "0.5")),
    "none": float(os.getenv("ACCURACY_WEIGHT_NONE", "0.0")),
}

_EMAIL_COLS = [
    "תאריך", "שולח", "לקוחות", "קבצים", "שורות",
    "עלות $", "זמן (שניות)", "נשלח", "דיוק %",
    "מלא (Full)", "גבוה (High)", "בינוני (Medium)", "נמוך (Low)", "ללא (None)",
    "PL Overrides", "שגיאות",
]


def _is_email_entry(e: Dict[str, Any]) -> bool:
    return bool(e.get("accuracy_data") or e.get("files_processed"))


def _entry_message_id(e: Dict[str, Any]) -> str:
    return str(e.get("message_id") or "").strip()


def _entry_files(e: Dict[str, Any]) -> int:
    return int(e.get("files_processed", 0) or 0)


def _unique_message_keys(entries, prefix: str):
    keys = set()
    for idx, entry in enumerate(entries):
        mid = _entry_message_id(entry)
        if mid:
            keys.add(mid)
        else:
            keys.add(f"{prefix}:{idx}:{entry.get('timestamp') or entry.get('start_time') or ''}")
    return keys


def _is_no_draw_email_entry(entry: Dict[str, Any]) -> bool:
    return (
        _get_items(entry) == 0
        and _entry_files(entry) == 0
        and float(entry.get("cost_usd", 0) or 0) == 0.0
        and str(entry.get("type") or "").upper() != "RERUN"
    )


def _run_type_for_entry(entry: Dict[str, Any], heavy_message_ids) -> str:
    run_type = str(entry.get("run_type") or "").strip().lower()
    if run_type in {"regular", "heavy"}:
        return run_type
    return "heavy" if _entry_message_id(entry) in heavy_message_ids else "regular"


def _operations_metrics(entries):
    event_entries = [entry for entry in entries if entry.get("event")]
    rerun_entries = [entry for entry in entries if str(entry.get("type") or "").upper() == "RERUN"]
    heavy_marked_events = [entry for entry in event_entries if entry.get("event") == "heavy_email_skipped"]
    no_draw_events = [
        entry for entry in event_entries
        if entry.get("event") in {"no_draw_skipped", "no_attachment_skipped"}
    ]
    skip_sender_events = [entry for entry in event_entries if entry.get("event") == "skip_sender"]
    skip_category_events = [entry for entry in event_entries if entry.get("event") == "skip_category"]

    heavy_marked_ids = _unique_message_keys(heavy_marked_events, "heavy")
    processable_entries = [
        entry for entry in entries
        if _is_email_entry(entry) and str(entry.get("type") or "").upper() != "RERUN"
    ]
    no_draw_processed_entries = [entry for entry in processable_entries if _is_no_draw_email_entry(entry)]
    processed_entries = [entry for entry in processable_entries if not _is_no_draw_email_entry(entry)]
    heavy_processed_entries = [
        entry for entry in processed_entries
        if _run_type_for_entry(entry, heavy_marked_ids) == "heavy"
    ]
    regular_processed_entries = [
        entry for entry in processed_entries
        if _run_type_for_entry(entry, heavy_marked_ids) == "regular"
    ]

    processed_keys = _unique_message_keys(processed_entries, "processed")
    no_draw_keys = _unique_message_keys(no_draw_processed_entries, "no-draw-processed") | _unique_message_keys(no_draw_events, "no-draw-event")
    rerun_keys = _unique_message_keys(rerun_entries, "rerun")
    skip_sender_keys = _unique_message_keys(skip_sender_events, "skip-sender")
    skip_category_keys = _unique_message_keys(skip_category_events, "skip-category")
    skip_keys = skip_sender_keys | skip_category_keys

    all_unique = processed_keys | no_draw_keys | rerun_keys | skip_keys

    regular_rows = sum(_get_items(entry) for entry in regular_processed_entries)
    heavy_rows = sum(_get_items(entry) for entry in heavy_processed_entries)
    regular_count = len(_unique_message_keys(regular_processed_entries, "regular"))
    heavy_count = len(_unique_message_keys(heavy_processed_entries, "heavy-processed"))
    regular_time = sum(float(entry.get("processing_time_seconds", 0) or 0) for entry in regular_processed_entries)
    heavy_time = sum(float(entry.get("processing_time_seconds", 0) or 0) for entry in heavy_processed_entries)

    arithmetic_gap = max(0, len(all_unique) - (len(processed_keys) + len(no_draw_keys) + len(rerun_keys)))

    return {
        "all_unique": len(all_unique),
        "processed": len(processed_keys),
        "regular": regular_count,
        "heavy": heavy_count,
        "no_draw": len(no_draw_keys),
        "rerun": len(rerun_keys),
        "skip_total": len(skip_keys),
        "skip_sender": len(skip_sender_keys),
        "skip_category": len(skip_category_keys),
        "arithmetic_gap": arithmetic_gap,
        "regular_rows": regular_rows,
        "heavy_rows": heavy_rows,
        "regular_avg_rows": (regular_rows / regular_count) if regular_count else 0,
        "heavy_avg_rows": (heavy_rows / heavy_count) if heavy_count else 0,
        "regular_avg_time": (regular_time / regular_count) if regular_count else 0,
        "heavy_avg_time": (heavy_time / heavy_count) if heavy_count else 0,
    }


def _get_items(e: Dict[str, Any]) -> int:
    ic = e.get("items_count", 0)
    return ic if ic and ic > 0 else e.get("accuracy_data", {}).get("total", 0)


def _calc_accuracy(acc_data: Dict[str, Any]) -> Optional[float]:
    total = acc_data.get("total", 0)
    if not total:
        return None
    return sum(acc_data.get(k, 0) * _WEIGHTS.get(k, 0) for k in _WEIGHTS) / total * 100


def _local_ts(e: Dict[str, Any]) -> str:
    ts = e.get("timestamp") or e.get("start_time") or ""
    try:
        dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
        return dt.astimezone().strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return ts[:19] if ts else ""


def _make_styles():
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    return {
        "header_fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "header_font": Font(bold=True, color="FFFFFF", size=11),
        "header_align": Alignment(horizontal="center", vertical="center"),
        "green": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "yellow": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "red": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
        "thin": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        ),
        "bold_font": Font(bold=True, size=11),
    }


def _write_header(ws, columns, styles, row=1):
    for ci, name in enumerate(columns, 1):
        cell = ws.cell(row=row, column=ci, value=name)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["header_align"]
        cell.border = styles["thin"]


def _auto_width(ws, min_w=10, max_w=35):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max(min(max_len + 3, max_w), min_w)


def _color_accuracy(cell, val, styles):
    if val is None:
        return
    if val >= 85:
        cell.fill = styles["green"]
    elif val >= 70:
        cell.fill = styles["yellow"]
    else:
        cell.fill = styles["red"]


def _write_no_data(ws, message="אין נתונים"):
    from openpyxl.styles import Font, Alignment
    cell = ws.cell(row=2, column=1, value=message)
    cell.font = Font(italic=True, color="888888", size=11)
    cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 40


def _build_summary_sheet(ws, entries, run_label, styles, title_label, run_start=None, total_unique_override=None):
    ws.sheet_view.rightToLeft = True
    thin = styles["thin"]

    unique_entries = []
    seen_message_keys = set()
    for idx, entry in enumerate(entries):
        mid = _entry_message_id(entry)
        key = mid if mid else f"row:{idx}:{entry.get('timestamp') or entry.get('start_time') or ''}"
        if key in seen_message_keys:
            continue
        seen_message_keys.add(key)
        unique_entries.append(entry)

    analyzed_total = len(unique_entries)
    total = int(total_unique_override) if total_unique_override is not None else analyzed_total
    sent = sum(1 for e in unique_entries if e.get("sent"))
    costs = [e.get("cost_usd", 0) for e in unique_entries if e.get("cost_usd")]
    total_cost = sum(costs)
    avg_cost = total_cost / len(costs) if costs else 0
    times = [e.get("processing_time_seconds", 0) for e in unique_entries if e.get("processing_time_seconds")]
    avg_time = sum(times) / len(times) if times else 0
    total_time = sum(times)

    acc = Counter()
    for e in unique_entries:
        ad = e.get("accuracy_data", {})
        for lvl in ["full", "high", "medium", "low", "none"]:
            acc[lvl] += ad.get(lvl, 0)
    total_items = sum(acc.values())
    success_rate = (acc["full"] + acc["high"]) / total_items * 100 if total_items else 0
    unique_snd = len({(e.get("sender") or "").lower() for e in unique_entries if e.get("sender")})
    unique_cust = len({c.upper() for e in unique_entries for c in (e.get("customers") or [])})

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    start_str = run_start.strftime("%Y-%m-%d %H:%M:%S") if run_start else "—"

    rows = [
        ["מדד", "ערך"],
        [f"— {title_label} —", ""],
        ["סוג ריצה", run_label],
        ["מתאריך", start_str],
        ["נוצר בתאריך", now_str],
        ["", ""],
        ["סה\"כ מיילים ייחודיים", total],
        ["מיילים לניתוח (ללא RERUN/אירועים)", analyzed_total],
        ["נשלחו בהצלחה", f"{sent} ({sent/analyzed_total*100:.1f}%)" if analyzed_total else "0"],
        ["שולחים ייחודיים", unique_snd],
        ["לקוחות ייחודיים", unique_cust],
        ["סה\"כ פריטים (שורות)", total_items],
        ["סה\"כ קבצים", sum(e.get("files_processed", 0) for e in unique_entries)],
        ["דיוק כולל (full+high)", f"{success_rate:.1f}%"],
        ["מלא (Full)", acc["full"]],
        ["גבוה (High)", acc["high"]],
        ["בינוני (Medium)", acc["medium"]],
        ["נמוך (Low)", acc["low"]],
        ["ללא (None)", acc["none"]],
        ["עלות כוללת", f"${total_cost:.2f}"],
        ["עלות ממוצעת/מייל", f"${avg_cost:.4f}"],
        ["עלות לפריט", f"${total_cost/total_items:.4f}" if total_items else "$0"],
        ["זמן כולל (דקות)", f"{total_time/60:.1f}"],
        ["זמן ממוצע/מייל (שניות)", f"{avg_time:.0f}"],
        ["זמן לפריט (שניות)", f"{total_time/total_items:.1f}" if total_items else "0"],
        ["PL Overrides", sum(e.get("pl_overrides", 0) for e in unique_entries)],
    ]

    for ri, row in enumerate(rows, 1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin
            if ri == 1:
                cell.fill = styles["header_fill"]
                cell.font = styles["header_font"]
                cell.alignment = styles["header_align"]

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 24


def _build_email_sheet(ws, entries, styles):
    ws.sheet_view.rightToLeft = True
    _write_header(ws, _EMAIL_COLS, styles)

    if not entries:
        _write_no_data(ws, "אין מיילים לתקופה זו")
        _auto_width(ws)
        return

    thin = styles["thin"]
    for ri, e in enumerate(entries, 2):
        ad = e.get("accuracy_data", {})
        acc_val = _calc_accuracy(ad)
        items = _get_items(e)
        vals = [
            _local_ts(e)[:19],
            e.get("sender", ""),
            ", ".join(e.get("customers", [])),
            e.get("files_processed", 0),
            items,
            e.get("cost_usd", 0),
            e.get("processing_time_seconds", 0),
            "כן" if e.get("sent") else "לא",
            f"{acc_val:.1f}%" if acc_val is not None else "—",
            ad.get("full", 0), ad.get("high", 0), ad.get("medium", 0),
            ad.get("low", 0), ad.get("none", 0),
            e.get("pl_overrides", 0),
            ", ".join(e.get("error_types", [])),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = thin
            if ci == 9:
                _color_accuracy(cell, acc_val, styles)
    _auto_width(ws)


def _build_daily_sheet(ws, entries, styles):
    ws.sheet_view.rightToLeft = True
    cols = ["תאריך", "מיילים", "נשלחו", "קבצים", "שורות", "עלות $", "זמן (דקות)", "דיוק %"]
    _write_header(ws, cols, styles)

    if not entries:
        _write_no_data(ws, "אין נתונים")
        _auto_width(ws)
        return

    thin = styles["thin"]
    daily = defaultdict(lambda: {"count": 0, "cost": 0, "sent": 0, "items": 0, "score": 0, "time": 0, "files": 0})
    for e in entries:
        day = _local_ts(e)[:10]
        if not day:
            continue
        daily[day]["count"] += 1
        daily[day]["cost"] += e.get("cost_usd", 0)
        daily[day]["sent"] += 1 if e.get("sent") else 0
        daily[day]["time"] += e.get("processing_time_seconds", 0)
        daily[day]["files"] += e.get("files_processed", 0)
        ad = e.get("accuracy_data", {})
        itm = sum(ad.get(k, 0) for k in ["full", "high", "medium", "low", "none"])
        daily[day]["items"] += itm
        daily[day]["score"] += sum(ad.get(k, 0) * _WEIGHTS[k] for k in _WEIGHTS)

    for ri, (day, d) in enumerate(sorted(daily.items()), 2):
        rate = d["score"] / d["items"] * 100 if d["items"] else 0
        row_vals = [day, d["count"], d["sent"], d["files"], d["items"], round(d["cost"], 2), round(d["time"] / 60, 1), f"{rate:.1f}%"]
        for ci, v in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = thin
            if ci == 8:
                _color_accuracy(cell, rate, styles)
    _auto_width(ws)


def _build_customers_sheet(ws, entries, styles):
    ws.sheet_view.rightToLeft = True
    cols = [
        "לקוח", "מיילים", "% מיילים", "קבצים", "שורות", "% שורות",
        "דיוק %", "מלא (Full)", "גבוה (High)", "בינוני (Medium)", "נמוך (Low)", "ללא (None)",
        "עלות $", "עלות/פריט $", "זמן (דקות)", "זמן/פריט (שניות)",
        "שולחים ייחודיים", "שיעור הצלחה",
    ]
    _write_header(ws, cols, styles)

    if not entries:
        _write_no_data(ws, "אין נתונים")
        _auto_width(ws)
        return

    thin = styles["thin"]
    bold_font = styles["bold_font"]
    cust_data = defaultdict(lambda: {
        "emails": 0, "files": 0, "items": 0, "score": 0,
        "cost": 0, "time": 0, "sent": 0,
        "full": 0, "high": 0, "medium": 0, "low": 0, "none": 0,
        "senders": set(),
    })
    for e in entries:
        customers = e.get("customers") or ["לא ידוע"]
        ad = e.get("accuracy_data", {})
        itm = _get_items(e)
        score = sum(ad.get(k, 0) * _WEIGHTS.get(k, 0) for k in _WEIGHTS)
        sender = (e.get("sender") or "").lower()
        for cname in customers:
            c = cname.upper()
            cust_data[c]["emails"] += 1
            cust_data[c]["files"] += e.get("files_processed", 0)
            cust_data[c]["items"] += itm
            cust_data[c]["score"] += score
            cust_data[c]["cost"] += e.get("cost_usd", 0)
            cust_data[c]["time"] += e.get("processing_time_seconds", 0)
            cust_data[c]["sent"] += 1 if e.get("sent") else 0
            for lvl in ["full", "high", "medium", "low", "none"]:
                cust_data[c][lvl] += ad.get(lvl, 0)
            if sender:
                cust_data[c]["senders"].add(sender)

    grand_emails = sum(d["emails"] for d in cust_data.values())
    grand_items = sum(d["items"] for d in cust_data.values())
    sorted_custs = sorted(cust_data.items(), key=lambda x: x[1]["items"], reverse=True)

    for ri, (cname, d) in enumerate(sorted_custs, 2):
        acc_pct = d["score"] / d["items"] * 100 if d["items"] else 0
        row_vals = [
            cname,
            d["emails"],
            f"{d['emails']/grand_emails*100:.1f}%" if grand_emails else "0%",
            d["files"],
            d["items"],
            f"{d['items']/grand_items*100:.1f}%" if grand_items else "0%",
            f"{acc_pct:.1f}%",
            d["full"], d["high"], d["medium"], d["low"], d["none"],
            round(d["cost"], 2),
            round(d["cost"] / d["items"], 4) if d["items"] else 0,
            round(d["time"] / 60, 1),
            round(d["time"] / d["items"], 1) if d["items"] else 0,
            len(d["senders"]),
            f"{d['sent']/d['emails']*100:.1f}%" if d["emails"] else "0%",
        ]
        for ci, v in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = thin
            if ci == 7:
                _color_accuracy(cell, acc_pct, styles)

    tot_row = len(sorted_custs) + 2
    for ci, v in [(1, 'סה"כ'), (2, grand_emails), (5, grand_items), (13, round(sum(d["cost"] for d in cust_data.values()), 2))]:
        cell = ws.cell(row=tot_row, column=ci, value=v)
        cell.font = bold_font
        cell.border = thin
    _auto_width(ws)


def _build_senders_sheet(ws, entries, styles):
    ws.sheet_view.rightToLeft = True
    cols = [
        "שולח", "לקוחות", "מיילים", "קבצים", "שורות",
        "דיוק %", "עלות $", "עלות/פריט $",
        "זמן (דקות)", "זמן/פריט (שניות)", "שיעור הצלחה",
    ]
    _write_header(ws, cols, styles)

    if not entries:
        _write_no_data(ws, "אין נתונים")
        _auto_width(ws)
        return

    thin = styles["thin"]
    sender_data = defaultdict(lambda: {
        "emails": 0, "files": 0, "items": 0,
        "score": 0, "cost": 0, "time": 0,
        "sent": 0, "customers": set(),
    })
    for e in entries:
        snd = (e.get("sender") or "לא ידוע").lower()
        ad = e.get("accuracy_data", {})
        itm = _get_items(e)
        score = sum(ad.get(k, 0) * _WEIGHTS.get(k, 0) for k in _WEIGHTS)
        sender_data[snd]["emails"] += 1
        sender_data[snd]["files"] += e.get("files_processed", 0)
        sender_data[snd]["items"] += itm
        sender_data[snd]["score"] += score
        sender_data[snd]["cost"] += e.get("cost_usd", 0)
        sender_data[snd]["time"] += e.get("processing_time_seconds", 0)
        sender_data[snd]["sent"] += 1 if e.get("sent") else 0
        for c in (e.get("customers") or []):
            sender_data[snd]["customers"].add(c.upper())

    sorted_senders = sorted(sender_data.items(), key=lambda x: x[1]["emails"], reverse=True)
    for ri, (snd, d) in enumerate(sorted_senders, 2):
        acc_pct = d["score"] / d["items"] * 100 if d["items"] else 0
        row_vals = [
            snd,
            ", ".join(sorted(d["customers"])),
            d["emails"],
            d["files"],
            d["items"],
            f"{acc_pct:.1f}%",
            round(d["cost"], 2),
            round(d["cost"] / d["items"], 4) if d["items"] else 0,
            round(d["time"] / 60, 1),
            round(d["time"] / d["items"], 1) if d["items"] else 0,
            f"{d['sent']/d['emails']*100:.1f}%" if d["emails"] else "0%",
        ]
        for ci, v in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = thin
            if ci == 6:
                _color_accuracy(cell, acc_pct, styles)
    _auto_width(ws)


def _build_operations_sheet(ws, entries, styles):
    ws.sheet_view.rightToLeft = True
    thin = styles["thin"]
    metrics = _operations_metrics(entries)

    rows = [
        ["מדד תפעולי", "ערך"],
        ["סה\"כ מיילים ייחודיים", metrics["all_unique"]],
        ["מיילים שעובדו", metrics["processed"]],
        ["רגילים", metrics["regular"]],
        ["כבדים", metrics["heavy"]],
        ["NO DRAW", metrics["no_draw"]],
        ["RERUN", metrics["rerun"]],
        ["דילוגים (סה\"כ)", metrics["skip_total"]],
        ["דילוג שולח", metrics["skip_sender"]],
        ["דילוג קטגוריה", metrics["skip_category"]],
        ["הפרש חשבוני", metrics["arithmetic_gap"]],
        ["", ""],
        ["שורות רגיל", metrics["regular_rows"]],
        ["שורות כבד", metrics["heavy_rows"]],
        ["ממוצע שורות רגיל", round(metrics["regular_avg_rows"], 2)],
        ["ממוצע שורות כבד", round(metrics["heavy_avg_rows"], 2)],
        ["ממוצע זמן רגיל (שניות)", round(metrics["regular_avg_time"], 1)],
        ["ממוצע זמן כבד (שניות)", round(metrics["heavy_avg_time"], 1)],
    ]

    for ri, row in enumerate(rows, 1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin
            if ri == 1:
                cell.fill = styles["header_fill"]
                cell.font = styles["header_font"]
                cell.alignment = styles["header_align"]

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 20


def build_workbook_dashboard(entries, run_label="dashboard", all_entries=None):
    """Build dashboard workbook including operational metrics."""
    import openpyxl

    styles = _make_styles()
    wb = openpyxl.Workbook()
    email_entries = [e for e in entries if _is_email_entry(e)]
    operational_entries = all_entries if all_entries is not None else entries

    ws1 = wb.active
    ws1.title = "סיכום כללי"
    total_unique = _operations_metrics(operational_entries)["all_unique"]
    _build_summary_sheet(
        ws1,
        email_entries,
        run_label,
        styles,
        "תקופה נבחרת",
        total_unique_override=total_unique,
    )

    ws2 = wb.create_sheet("נתוני מיילים")
    _build_email_sheet(ws2, email_entries, styles)

    ws3 = wb.create_sheet("סיכום יומי")
    _build_daily_sheet(ws3, email_entries, styles)

    ws4 = wb.create_sheet("לקוחות")
    _build_customers_sheet(ws4, email_entries, styles)

    ws5 = wb.create_sheet("שולחים")
    _build_senders_sheet(ws5, email_entries, styles)

    ws6 = wb.create_sheet("סטטוס תפעולי")
    _build_operations_sheet(ws6, operational_entries, styles)

    return wb


def build_workbook_scheduler(all_entries, run_entries, run_label, run_start=None):
    """Build scheduler workbook: current-run + full-history including operations sheets."""
    import openpyxl

    styles = _make_styles()
    wb = openpyxl.Workbook()
    run_email_entries = [e for e in run_entries if _is_email_entry(e)]
    all_email_entries = [e for e in all_entries if _is_email_entry(e)]

    ws1 = wb.active
    ws1.title = "יום נוכחי - סיכום"
    run_total_unique = _operations_metrics(run_entries)["all_unique"]
    _build_summary_sheet(
        ws1,
        run_email_entries,
        run_label,
        styles,
        "יום נוכחי",
        run_start,
        total_unique_override=run_total_unique,
    )

    ws2 = wb.create_sheet("יום נוכחי - מיילים")
    _build_email_sheet(ws2, run_email_entries, styles)

    ws3 = wb.create_sheet("יום נוכחי - יומי")
    _build_daily_sheet(ws3, run_email_entries, styles)

    ws4 = wb.create_sheet("יום נוכחי - לקוחות")
    _build_customers_sheet(ws4, run_email_entries, styles)

    ws5 = wb.create_sheet("יום נוכחי - שולחים")
    _build_senders_sheet(ws5, run_email_entries, styles)

    ws6 = wb.create_sheet("יום נוכחי - תפעול")
    _build_operations_sheet(ws6, run_entries, styles)

    ws7 = wb.create_sheet("היסטוריה - סיכום")
    all_total_unique = _operations_metrics(all_entries)["all_unique"]
    _build_summary_sheet(
        ws7,
        all_email_entries,
        run_label,
        styles,
        "כל ההיסטוריה",
        total_unique_override=all_total_unique,
    )

    ws8 = wb.create_sheet("היסטוריה - מיילים")
    _build_email_sheet(ws8, all_email_entries, styles)

    ws9 = wb.create_sheet("היסטוריה - יומי")
    _build_daily_sheet(ws9, all_email_entries, styles)

    ws10 = wb.create_sheet("היסטוריה - לקוחות")
    _build_customers_sheet(ws10, all_email_entries, styles)

    ws11 = wb.create_sheet("היסטוריה - שולחים")
    _build_senders_sheet(ws11, all_email_entries, styles)

    ws12 = wb.create_sheet("היסטוריה - תפעול")
    _build_operations_sheet(ws12, all_entries, styles)

    return wb


def workbook_to_bytes(wb) -> bytes:
    import io

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
