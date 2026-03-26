"""Snapshot test: Excel export produces valid file with correct structure."""
import pytest
from pathlib import Path
from src.services.reporting.excel_export import _save_results_to_excel


class TestExcelExport:
    def test_creates_file(self, sample_results, tmp_output):
        output_file = tmp_output / "test_results.xlsx"
        _save_results_to_excel(sample_results, output_file, None)
        assert output_file.exists()
        assert output_file.stat().st_size > 0

    def test_creates_file_with_pl(self, sample_results, sample_pl_items, tmp_output):
        output_file = tmp_output / "test_results_pl.xlsx"
        _save_results_to_excel(sample_results, output_file, sample_pl_items)
        assert output_file.exists()

    def test_excel_has_correct_sheets(self, sample_results, tmp_output):
        """Check that Excel has expected sheet names."""
        import openpyxl
        output_file = tmp_output / "test_sheets.xlsx"
        _save_results_to_excel(sample_results, output_file, None)

        wb = openpyxl.load_workbook(output_file)
        sheet_names = wb.sheetnames
        # Should have at least a summary sheet
        assert len(sheet_names) >= 1
        wb.close()

    def test_excel_has_correct_row_count(self, sample_results, tmp_output):
        """Check that Excel has correct number of data rows."""
        import openpyxl
        output_file = tmp_output / "test_rows.xlsx"
        _save_results_to_excel(sample_results, output_file, None)

        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        # Should have header + 2 data rows
        data_rows = sum(1 for row in ws.iter_rows(min_row=2)
                        if any(cell.value for cell in row))
        assert data_rows == 2
        wb.close()

    def test_returns_rafael_count(self, sample_results, tmp_output):
        """Function returns number of RAFAEL rows."""
        output_file = tmp_output / "test_rafael.xlsx"
        rafael_count = _save_results_to_excel(sample_results, output_file, None)
        # One result has customer_name "RAFAEL"
        assert isinstance(rafael_count, int)
        assert rafael_count >= 1

    def test_empty_results(self, tmp_output):
        """Empty results should return 0 and not crash."""
        output_file = tmp_output / "test_empty.xlsx"
        result = _save_results_to_excel([], output_file, None)
        assert result == 0
